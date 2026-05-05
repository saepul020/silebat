from io import BytesIO

from django.contrib import messages
from django.db import transaction
from django.db.models import Count
from django.db.models.functions import ExtractYear
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.contrib.auth import password_validation, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from apps.core.file_cleanup import delete_file_if_unused, delete_instance_files
from apps.core.list_pagination import paginate_list
from apps.core.excel_utils import build_excel_response
from apps.core.import_utils import (
    import_cell as _import_cell,
    json_response_or_none as _import_json_response,
    load_import_worksheet as _load_shared_import_worksheet,
    string_cell as _string_cell,
)
from apps.core.permissions import (
    can_access_pengguna_app,
    can_edit_user,
    can_view_user_detail,
    deny_access,
    is_super_admin,
)

from .forms import PelatihanForm, UserForm, UserProfileForm, UserUpdateForm
from .models import Pelatihan, Role, User, UserProfile, get_default_role_queryset

IMPORT_PENGGUNA_SESSION_KEY = "pengguna_import_validated_rows"
IMPORT_PENGGUNA_HEADERS = [
    "Username",
    "Nama Lengkap dan Gelar",
    "Email",
    "NIP / NIK",
    "Nomor HP",
    "Kata Sandi",
    "Peran",
]
IMPORT_PENGGUNA_REQUIRED_HEADERS = IMPORT_PENGGUNA_HEADERS[:]
IMPORT_PENGGUNA_MAX_SIZE = 7 * 1024 * 1024


def _deny_import_access(request):
    messages.error(request, 'Fitur import daftar pengguna hanya dapat diakses oleh Super Admin.')
    return redirect('pengguna:daftar')


def _load_import_worksheet(file_obj, headers, required_headers):
    return _load_shared_import_worksheet(
        file_obj,
        headers,
        required_headers,
        max_size_bytes=IMPORT_PENGGUNA_MAX_SIZE,
    )


def _validate_duplicate_value(*, row_errors, seen, value, excel_row_number, label):
    key = (value or '').strip().lower()
    if not key:
        return
    if key in seen:
        row_errors.append(f'{label} duplikat dengan baris {seen[key]}.')
    else:
        seen[key] = excel_row_number


def _validate_pengguna_import(file_obj):
    worksheet, normalized_headers, header_aliases, initial_errors = _load_import_worksheet(
        file_obj,
        IMPORT_PENGGUNA_HEADERS,
        IMPORT_PENGGUNA_REQUIRED_HEADERS,
    )
    if initial_errors:
        return [], initial_errors

    role_queryset = get_default_role_queryset()
    role_map = {role.nama.lower(): role.nama for role in role_queryset}
    rows = []
    errors = []
    seen_username = {}
    seen_email = {}
    seen_nip = {}
    seen_no_hp = {}

    def cell(row, header):
        return _import_cell(row, header, normalized_headers, header_aliases)

    for excel_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_string_cell(value) for value in row):
            continue

        row_errors = []
        data = {
            'username': cell(row, 'Username'),
            'nama_lengkap': cell(row, 'Nama Lengkap dan Gelar'),
            'email': cell(row, 'Email'),
            'nip': cell(row, 'NIP / NIK'),
            'no_hp': cell(row, 'Nomor HP'),
            'password': cell(row, 'Kata Sandi'),
            'role': cell(row, 'Peran'),
        }

        for header in IMPORT_PENGGUNA_REQUIRED_HEADERS:
            if not cell(row, header):
                row_errors.append(f'{header} wajib diisi.')

        if data['email']:
            try:
                validate_email(data['email'])
            except ValidationError:
                row_errors.append('Email harus menggunakan format email yang valid.')

        if data['nip'] and not data['nip'].isdigit():
            row_errors.append('NIP / NIK hanya boleh berisi angka.')

        if data['no_hp'] and not data['no_hp'].isdigit():
            row_errors.append('Nomor HP hanya boleh berisi angka.')

        if data['password']:
            try:
                password_validation.validate_password(data['password'])
            except ValidationError as exc:
                row_errors.append('Kata Sandi tidak valid: ' + ' '.join(exc.messages))

        role_key = data['role'].lower()
        if role_key:
            if role_key not in role_map:
                row_errors.append('Peran tidak sesuai pilihan yang tersedia.')
            else:
                data['role'] = role_map[role_key]

        _validate_duplicate_value(row_errors=row_errors, seen=seen_username, value=data['username'], excel_row_number=excel_row_number, label='Username')
        _validate_duplicate_value(row_errors=row_errors, seen=seen_email, value=data['email'], excel_row_number=excel_row_number, label='Email')
        _validate_duplicate_value(row_errors=row_errors, seen=seen_nip, value=data['nip'], excel_row_number=excel_row_number, label='NIP / NIK')
        _validate_duplicate_value(row_errors=row_errors, seen=seen_no_hp, value=data['no_hp'], excel_row_number=excel_row_number, label='Nomor HP')

        if data['username'] and User.objects.filter(username__iexact=data['username']).exists():
            row_errors.append('Username sudah digunakan.')
        if data['email'] and User.objects.filter(email__iexact=data['email']).exists():
            row_errors.append('Email sudah digunakan.')
        if data['nip'] and User.objects.filter(nip__iexact=data['nip']).exists():
            row_errors.append('NIP / NIK sudah digunakan.')
        if data['no_hp'] and User.objects.filter(no_hp__iexact=data['no_hp']).exists():
            row_errors.append('Nomor HP sudah digunakan.')

        if row_errors:
            errors.append(f'Baris {excel_row_number}: ' + ' '.join(row_errors))
            continue

        data['nama_lengkap'] = ' '.join(data['nama_lengkap'].split())
        rows.append(data)

    if not rows and not errors:
        errors.append('File Excel tidak memiliki data pengguna untuk diimport.')

    return rows, errors


def _save_pengguna_import(rows):
    duplicate_errors = []
    for index, row in enumerate(rows, start=1):
        if User.objects.filter(username__iexact=row['username']).exists():
            duplicate_errors.append(f'Data valid nomor {index}: Username sudah digunakan.')
        if User.objects.filter(email__iexact=row['email']).exists():
            duplicate_errors.append(f'Data valid nomor {index}: Email sudah digunakan.')
        if User.objects.filter(nip__iexact=row['nip']).exists():
            duplicate_errors.append(f'Data valid nomor {index}: NIP / NIK sudah digunakan.')
        if User.objects.filter(no_hp__iexact=row['no_hp']).exists():
            duplicate_errors.append(f'Data valid nomor {index}: Nomor HP sudah digunakan.')

    if duplicate_errors:
        return 0, duplicate_errors

    with transaction.atomic():
        for row in rows:
            role = Role.objects.get(nama=row['role'])
            user = User(
                username=row['username'],
                first_name=row['nama_lengkap'],
                last_name='',
                email=row['email'],
                nip=row['nip'],
                no_hp=row['no_hp'],
            )
            user.set_password(row['password'])
            user.full_clean()
            user.save()
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.role = role
            profile.save()

    return len(rows), []


def _handle_pengguna_import_post(request):
    if request.method != 'POST':
        return redirect('pengguna:daftar')

    action = request.POST.get('import_action')

    if action == 'cancel':
        request.session.pop(IMPORT_PENGGUNA_SESSION_KEY, None)
        response = _import_json_response(request, {'ok': True, 'cancelled': True})
        if response:
            return response
        return redirect('pengguna:daftar')

    if action == 'validate':
        rows, errors = _validate_pengguna_import(request.FILES.get('file_import'))
        import_context = {
            'show_modal': True,
            'validated': not errors,
            'can_save': bool(rows) and not errors,
            'total_rows': len(rows),
            'errors': errors,
        }
        if errors:
            request.session.pop(IMPORT_PENGGUNA_SESSION_KEY, None)
        else:
            request.session[IMPORT_PENGGUNA_SESSION_KEY] = rows
            request.session.modified = True

        response = _import_json_response(request, {
            'ok': not bool(errors),
            'validated': not bool(errors),
            'can_save': bool(rows) and not errors,
            'total_rows': len(rows),
            'errors': errors,
            'message': f'Validasi berhasil. {len(rows)} data siap disimpan.' if not errors else 'Validasi belum berhasil.',
        })
        if response:
            return response
        return _render_daftar_pengguna(request, import_context=import_context)

    if action == 'save':
        rows = request.session.get(IMPORT_PENGGUNA_SESSION_KEY) or []
        if not rows:
            payload = {
                'ok': False,
                'saved': False,
                'can_save': False,
                'errors': ['Data validasi tidak ditemukan atau sudah kedaluwarsa. Lakukan Validasi Data terlebih dahulu.'],
                'message': 'Data belum dapat disimpan.',
            }
            response = _import_json_response(request, payload)
            if response:
                return response
            return _render_daftar_pengguna(request, import_context={'show_modal': True, 'errors': payload['errors']})

        total_saved, save_errors = _save_pengguna_import(rows)
        if save_errors:
            request.session.pop(IMPORT_PENGGUNA_SESSION_KEY, None)
            payload = {
                'ok': False,
                'saved': False,
                'can_save': False,
                'errors': save_errors,
                'message': 'Data tidak disimpan karena ditemukan duplikasi terbaru di database.',
            }
            response = _import_json_response(request, payload)
            if response:
                return response
            return _render_daftar_pengguna(request, import_context={'show_modal': True, 'errors': save_errors})

        request.session.pop(IMPORT_PENGGUNA_SESSION_KEY, None)
        response = _import_json_response(request, {
            'ok': True,
            'saved': True,
            'can_save': False,
            'total_rows': total_saved,
            'errors': [],
            'message': f'{total_saved} data pengguna berhasil diimport.',
            'redirect_url': reverse('pengguna:daftar'),
        })
        if response:
            return response
        return redirect('pengguna:daftar')

    payload = {'ok': False, 'errors': ['Aksi import tidak dikenali.'], 'message': 'Aksi import tidak dikenali.'}
    response = _import_json_response(request, payload, status=400)
    if response:
        return response
    return redirect('pengguna:daftar')


def _download_import_template(*, headers, sample, references, validations, worksheet_title, filename):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError('Library openpyxl belum tersedia. Jalankan: pip install openpyxl') from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_title

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='EAF2FF')
        worksheet.column_dimensions[cell.column_letter].width = 28

    for column_index, value in enumerate(sample, start=1):
        worksheet.cell(row=2, column=column_index, value=value)

    if references:
        lists_sheet = workbook.create_sheet('Referensi Pilihan')
        for col_index, (title, values) in enumerate(references, start=1):
            lists_sheet.cell(row=1, column=col_index, value=title).font = Font(bold=True)
            lists_sheet.column_dimensions[lists_sheet.cell(row=1, column=col_index).column_letter].width = 34
            for row_index, value in enumerate(values, start=2):
                lists_sheet.cell(row=row_index, column=col_index, value=value)

    for header, values in (validations or {}).items():
        if header not in headers:
            continue
        column_index = headers.index(header) + 1
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        formula = '"' + ','.join(values) + '"'
        validator = DataValidation(type='list', formula1=formula, allow_blank=False)
        worksheet.add_data_validation(validator)
        validator.add(f'{column_letter}2:{column_letter}1000')

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response




@login_required
def export_pengguna(request):
    if not can_access_pengguna_app(request.user):
        return _deny_import_access(request)

    queryset = User.objects.select_related('profile', 'profile__role', 'profile__nama_tim').order_by('username')
    rows = []
    for user in queryset:
        profile = user.safe_profile
        rows.append([
            user.username,
            user.get_full_name() or '-',
            user.email or '-',
            user.nip or '-',
            user.no_hp or '-',
            profile.role.nama if profile.role else '-',
            profile.jabatan or '-',
            str(profile.nama_tim) if profile.nama_tim else '-',
            profile.alamat or '-',
            'Aktif' if user.is_active else 'Nonaktif',
            user.date_joined,
            user.last_login,
        ])

    return build_excel_response(
        'export_daftar_pengguna.xlsx',
        [
            {
                'title': 'Daftar Pengguna',
                'headers': [
                    'Username',
                    'Nama Lengkap dan Gelar',
                    'Email',
                    'NIP / NIK',
                    'Nomor HP',
                    'Peran',
                    'Jabatan',
                    'Nama Tim',
                    'Alamat',
                    'Status Akun',
                    'Tanggal Bergabung',
                    'Login Terakhir',
                ],
                'rows': rows,
            }
        ],
    )

@login_required
def import_pengguna(request):
    if not can_access_pengguna_app(request.user):
        return _deny_import_access(request)
    return _handle_pengguna_import_post(request)


@login_required
def download_format_import_pengguna(request):
    if not can_access_pengguna_app(request.user):
        return _deny_import_access(request)

    role_values = [role.nama for role in get_default_role_queryset()]
    return _download_import_template(
        headers=IMPORT_PENGGUNA_HEADERS,
        sample=[
            'budi.santoso',
            'Budi Santoso, S.T.',
            'budi.santoso@example.com',
            '199001012020011001',
            '081234567890',
            'PasswordRahasia123',
            'User',
        ],
        references=[
            ('Peran', role_values),
        ],
        validations={
            'Peran': role_values,
        },
        worksheet_title='Format Import Pengguna',
        filename='format_import_daftar_pengguna.xlsx',
    )



def _render_daftar_pengguna(request, import_context=None):
    queryset = User.objects.select_related('profile', 'profile__role', 'profile__nama_tim').order_by('username')
    pagination_context = paginate_list(request, queryset)
    users = pagination_context["items"]
    for user in users:
        user.safe_profile

    context = {
        'users': users,
        'can_manage_user_app': True,
        'import_context': import_context or {},
    }
    context.update(pagination_context)
    return render(request, 'pengguna/daftar_pengguna.html', context)


@login_required
def daftar_pengguna(request):
    if not can_access_pengguna_app(request.user):
        return deny_access(request, 'App "Pengguna" hanya dapat diakses oleh Super Admin.')

    return _render_daftar_pengguna(request)


@login_required
def tambah_pengguna(request):
    if not can_access_pengguna_app(request.user):
        return deny_access(request, 'Fitur "Tambah Pengguna" hanya dapat diakses oleh Super Admin.')

    if request.method == 'POST':
        form_user = UserForm(request.POST)
        form_profile = UserProfileForm(request.POST, request.FILES, allow_role_edit=True, require_role=True)

        if form_user.is_valid() and form_profile.is_valid():
            with transaction.atomic():
                user = form_user.save()
                profile = UserProfile.objects.select_for_update().get(user=user)
                form_profile = UserProfileForm(
                    request.POST,
                    request.FILES,
                    instance=profile,
                    allow_role_edit=True,
                    require_role=True,
                )
                form_profile.is_valid()
                form_profile.save()

            messages.success(request, 'Pengguna berhasil ditambahkan.')
            return redirect('pengguna:daftar')
    else:
        form_user = UserForm()
        form_profile = UserProfileForm(allow_role_edit=True, require_role=True)

    context = {
        'form_user': form_user,
        'form_profile': form_profile,
        'page_title': 'Tambah Pengguna',
        'page_subtitle': 'Masukkan data akun dan profil pengguna baru.',
        'submit_label': 'Simpan',
        'is_edit': False,
        'obj': None,
    }
    return render(request, 'pengguna/tambah_pengguna.html', context)


@login_required
def edit_pengguna(request, pk):
    user = get_object_or_404(
        User.objects.select_related('profile', 'profile__role', 'profile__nama_tim'),
        pk=pk,
    )

    if not can_edit_user(request.user, user):
        return deny_access(
            request,
            'Anda hanya dapat mengubah profil Anda sendiri, kecuali jika Anda adalah Super Admin.',
        )

    profile, _ = UserProfile.objects.get_or_create(user=user)
    allow_admin_fields = is_super_admin(request.user)

    if request.method == 'POST':
        form_user = UserUpdateForm(
            request.POST,
            instance=user,
            allow_username_edit=allow_admin_fields,
        )
        form_profile = UserProfileForm(
            request.POST,
            request.FILES,
            instance=profile,
            allow_role_edit=allow_admin_fields,
        )

        if form_user.is_valid() and form_profile.is_valid():
            updated_user = form_user.save()
            form_profile.save()

            if request.user.pk == updated_user.pk and form_user.password_changed():
                update_session_auth_hash(request, updated_user)

            if form_user.password_changed():
                messages.success(request, 'Data pengguna dan password berhasil diperbarui.')
            else:
                messages.success(request, 'Data pengguna berhasil diperbarui.')

            return redirect('pengguna:detail', pk=updated_user.pk)
    else:
        form_user = UserUpdateForm(instance=user, allow_username_edit=allow_admin_fields)
        form_profile = UserProfileForm(instance=profile, allow_role_edit=allow_admin_fields)

    context = {
        'form_user': form_user,
        'form_profile': form_profile,
        'page_title': 'Edit Pengguna',
        'page_subtitle': 'Perbarui data akun, foto profil, dan informasi pengguna.',
        'submit_label': 'Update',
        'is_edit': True,
        'obj': user,
    }
    return render(request, 'pengguna/tambah_pengguna.html', context)


@login_required
def hapus_pengguna(request, pk):
    if not can_access_pengguna_app(request.user):
        return deny_access(request, 'Fitur "Hapus Pengguna" hanya dapat diakses oleh Super Admin.')

    user = get_object_or_404(User, pk=pk)

    if request.user.pk == user.pk:
        messages.error(request, 'Akun yang sedang digunakan tidak dapat dihapus.')
        return redirect('pengguna:daftar')

    if request.method == 'POST':
        nama = user.get_full_name() or user.username
        user.delete()
        messages.success(request, f'Pengguna "{nama}" berhasil dihapus.')
        return redirect('pengguna:daftar')

    return redirect('pengguna:daftar')


@login_required
def detail_pengguna(request, pk):
    user = get_object_or_404(
        User.objects.select_related('profile', 'profile__role', 'profile__nama_tim'),
        pk=pk,
    )

    if not can_view_user_detail(request.user, user):
        return deny_access(
            request,
            'Anda hanya dapat melihat profil Anda sendiri, kecuali jika Anda adalah Super Admin.',
        )

    UserProfile.objects.get_or_create(user=user)
    return render(
        request,
        'pengguna/detail_pengguna.html',
        {
            'obj': user,
            'can_edit_obj': can_edit_user(request.user, user),
        },
    )


@login_required
def dashboard_sdm(request):
    pelatihan_qs = Pelatihan.objects.select_related('user')
    if not is_super_admin(request.user):
        pelatihan_qs = pelatihan_qs.filter(user=request.user)

    total_pelatihan = pelatihan_qs.count()
    pelatihan_terbaru = pelatihan_qs.order_by('-tanggal_mulai', '-created_at')[:5]
    current_year = timezone.localtime().year

    tipe_rows = list(
        pelatihan_qs.annotate(year=ExtractYear('tanggal_mulai'))
        .values('year', 'tipe_pelatihan')
        .annotate(total=Count('id'))
        .order_by('year', 'tipe_pelatihan')
    )
    jenis_rows = list(
        pelatihan_qs.annotate(year=ExtractYear('tanggal_mulai'))
        .values('year', 'jenis_pelatihan')
        .annotate(total=Count('id'))
        .order_by('year', 'jenis_pelatihan')
    )
    available_years = sorted(
        {
            int(row['year'])
            for row in tipe_rows + jenis_rows
            if row.get('year')
        }
        | {current_year}
    )
    sdm_tipe_chart = {
        'rows': [
            {
                'year': int(row['year']),
                'key': row['tipe_pelatihan'],
                'total': int(row['total'] or 0),
            }
            for row in tipe_rows
            if row.get('year') and row.get('tipe_pelatihan')
        ],
        'availableYears': available_years,
        'defaultYear': current_year,
        'categories': [
            {'id': Pelatihan.TIPE_INTERNAL, 'label': Pelatihan.TIPE_INTERNAL},
            {'id': Pelatihan.TIPE_EKSTERNAL, 'label': Pelatihan.TIPE_EKSTERNAL},
        ],
    }
    sdm_jenis_chart = {
        'rows': [
            {
                'year': int(row['year']),
                'key': row['jenis_pelatihan'],
                'total': int(row['total'] or 0),
            }
            for row in jenis_rows
            if row.get('year') and row.get('jenis_pelatihan')
        ],
        'availableYears': available_years,
        'defaultYear': current_year,
        'categories': [
            {'id': Pelatihan.JENIS_LABORATORIUM, 'label': Pelatihan.JENIS_LABORATORIUM},
            {'id': Pelatihan.JENIS_NON_LABORATORIUM, 'label': Pelatihan.JENIS_NON_LABORATORIUM},
        ],
    }

    return render(
        request,
        'pengguna/dashboard_sdm.html',
        {
            'total_pelatihan': total_pelatihan,
            'pelatihan_terbaru': pelatihan_terbaru,
            'sdm_tipe_chart': sdm_tipe_chart,
            'sdm_jenis_chart': sdm_jenis_chart,
            'is_config_placeholder': True,
        },
    )


def _get_pelatihan_queryset_for_user(user):
    queryset = Pelatihan.objects.select_related('user', 'user__profile', 'user__profile__role')
    if is_super_admin(user):
        return queryset.order_by('-tanggal_mulai', '-created_at')
    return queryset.filter(user=user).order_by('-tanggal_mulai', '-created_at')


def _get_pelatihan_for_user_or_404(request, pk):
    return get_object_or_404(_get_pelatihan_queryset_for_user(request.user), pk=pk)


@login_required
def daftar_pelatihan(request):
    queryset = _get_pelatihan_queryset_for_user(request.user)
    pagination_context = paginate_list(request, queryset)
    context = {
        'items': pagination_context['items'],
        'can_view_all_pelatihan': is_super_admin(request.user),
    }
    context.update(pagination_context)
    return render(request, 'pengguna/daftar_pelatihan.html', context)


@login_required
def tambah_pelatihan(request):
    if request.method == 'POST':
        form = PelatihanForm(request.POST, request.FILES)
        if form.is_valid():
            pelatihan = form.save(commit=False)
            pelatihan.user = request.user
            pelatihan.save()
            messages.success(request, 'Data pelatihan berhasil ditambahkan.')
            return redirect('pengguna:pelatihan_daftar')
    else:
        form = PelatihanForm()

    return render(
        request,
        'pengguna/pelatihan_form.html',
        {
            'form': form,
            'page_title': 'Tambah Pelatihan',
            'page_subtitle': 'Masukkan data pelatihan pengguna.',
            'submit_label': 'Simpan',
            'is_edit': False,
            'obj': None,
        },
    )


@login_required
def edit_pelatihan(request, pk):
    pelatihan = _get_pelatihan_for_user_or_404(request, pk)
    old_file_sertifikat = pelatihan.file_sertifikat
    old_file_materi = pelatihan.file_materi

    if request.method == 'POST':
        form = PelatihanForm(request.POST, request.FILES, instance=pelatihan)
        if form.is_valid():
            updated = form.save(commit=False)
            updated.user = pelatihan.user
            updated.save()

            if old_file_sertifikat and getattr(old_file_sertifikat, 'name', '') != getattr(updated.file_sertifikat, 'name', ''):
                delete_file_if_unused(Pelatihan, 'file_sertifikat', old_file_sertifikat, exclude_pk=updated.pk)
            if old_file_materi and getattr(old_file_materi, 'name', '') != getattr(updated.file_materi, 'name', ''):
                delete_file_if_unused(Pelatihan, 'file_materi', old_file_materi, exclude_pk=updated.pk)

            messages.success(request, 'Data pelatihan berhasil diperbarui.')
            return redirect('pengguna:pelatihan_detail', pk=updated.pk)
    else:
        form = PelatihanForm(instance=pelatihan)

    return render(
        request,
        'pengguna/pelatihan_form.html',
        {
            'form': form,
            'page_title': 'Edit Pelatihan',
            'page_subtitle': 'Perbarui data pelatihan pengguna.',
            'submit_label': 'Update',
            'is_edit': True,
            'obj': pelatihan,
        },
    )


@login_required
def detail_pelatihan(request, pk):
    pelatihan = _get_pelatihan_for_user_or_404(request, pk)
    return render(
        request,
        'pengguna/detail_pelatihan.html',
        {
            'obj': pelatihan,
            'can_view_all_pelatihan': is_super_admin(request.user),
        },
    )


@login_required
def hapus_pelatihan(request, pk):
    pelatihan = _get_pelatihan_for_user_or_404(request, pk)
    if request.method == 'POST':
        nama_pelatihan = pelatihan.nama_pelatihan
        delete_instance_files(pelatihan, ['file_sertifikat', 'file_materi'])
        pelatihan.delete()
        messages.success(request, f'Data pelatihan "{nama_pelatihan}" berhasil dihapus.')
        return redirect('pengguna:pelatihan_daftar')

    return redirect('pengguna:pelatihan_daftar')
