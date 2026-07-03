from datetime import datetime
from io import BytesIO
from typing import Mapping, Sequence

from django.http import HttpResponse


EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _safe_sheet_title(title: str) -> str:
    title = str(title or "Sheet1").strip() or "Sheet1"
    for char in "[]:*?/\\":
        title = title.replace(char, "-")
    return title[:31]


def _normalize_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def build_excel_response(filename: str, sheets: Sequence[Mapping]) -> HttpResponse:
    """
    Membuat response file Excel sederhana untuk kebutuhan export data list.

    sheets: [
        {
            "title": "Nama Sheet",
            "headers": ["Kolom 1", "Kolom 2"],
            "rows": [[...], [...]],
        }
    ]
    """
    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("Library openpyxl belum tersedia. Jalankan: pip install openpyxl") from exc

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    if not sheets:
        sheets = [{"title": "Data", "headers": [], "rows": []}]

    for sheet in sheets:
        worksheet = workbook.create_sheet(_safe_sheet_title(sheet.get("title", "Data")))
        headers = list(sheet.get("headers") or [])
        rows = list(sheet.get("rows") or [])

        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="EAF2FF")

        for row_index, row in enumerate(rows, start=2):
            for column_index, value in enumerate(row, start=1):
                worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=_normalize_cell_value(value),
                )

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_index, header in enumerate(headers, start=1):
            letter = get_column_letter(column_index)
            values = [str(header or "")]
            values.extend(
                str(_normalize_cell_value(row[column_index - 1]))
                for row in rows[:200]
                if len(row) >= column_index
            )
            width = min(max((len(value) for value in values), default=12) + 2, 55)
            worksheet.column_dimensions[letter].width = max(width, 12)

        chart_config = sheet.get("chart") or {}
        if chart_config and rows and len(headers) > 1:
            chart_type = str(chart_config.get("type") or "bar").lower()
            chart = PieChart() if chart_type == "pie" else BarChart()
            data = Reference(
                worksheet,
                min_col=2,
                max_col=len(headers),
                min_row=1,
                max_row=len(rows) + 1,
            )
            categories = Reference(
                worksheet,
                min_col=1,
                min_row=2,
                max_row=len(rows) + 1,
            )
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            chart.title = str(chart_config.get("title") or sheet.get("title") or "Grafik")
            chart.height = 10
            chart.width = 18
            if isinstance(chart, BarChart):
                chart.type = "col"
                chart.y_axis.title = str(chart_config.get("value_title") or "Total")
                chart.x_axis.title = str(chart_config.get("category_title") or "Kategori")
                if chart_config.get("stacked"):
                    chart.grouping = "stacked"
                    chart.overlap = 100
            worksheet.add_chart(chart, f"{get_column_letter(len(headers) + 2)}2")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type=EXCEL_CONTENT_TYPE)
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
