from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Iterable, Sequence

from django.http import JsonResponse

EXCEL_EXTENSIONS = (".xlsx",)
EXCEL_READ_ERROR = "File Excel tidak dapat dibaca. Pastikan file sesuai format .xlsx dan tidak rusak."
OPENPYXL_MISSING_ERROR = "Library openpyxl belum tersedia. Jalankan: pip install openpyxl"


def normalize_import_header(value) -> str:
    """Normalisasi nama kolom Excel agar cocok dengan header template."""
    return str(value or "").strip().lower().replace(" ", "_").replace("/", "_").replace("-", "_")


def string_cell(value, *, date_to_iso: bool = False) -> str:
    """Ambil nilai cell Excel sebagai string bersih tanpa akhiran .0 yang tidak perlu."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if date_to_iso and isinstance(value, datetime):
        return value.date().isoformat()
    if date_to_iso and isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def choice_values(choices: Iterable[tuple]) -> set:
    """Ambil kumpulan value valid dari Django choices."""
    return {value for value, _label in choices}


def wants_json(request) -> bool:
    """Deteksi request AJAX/import modal yang mengharapkan JSON."""
    return (
        request.headers.get("x-requested-with") == "XMLHttpRequest"
        or "application/json" in request.headers.get("accept", "")
    )


def json_response_or_none(request, payload, status: int = 200):
    """Kembalikan JsonResponse hanya untuk request yang memang meminta JSON."""
    if not wants_json(request):
        return None
    return JsonResponse(payload, status=status)


def load_import_workbook(file_obj, *, max_size_bytes: int, max_size_label: str = "7 MB"):
    """Validasi dasar dan baca workbook Excel upload."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(OPENPYXL_MISSING_ERROR) from exc

    if not file_obj:
        return None, ["File Excel wajib diupload."]

    filename = str(getattr(file_obj, "name", "") or "")
    if not filename.lower().endswith(EXCEL_EXTENSIONS):
        return None, ["Format file harus berupa Excel .xlsx."]

    if getattr(file_obj, "size", 0) > int(max_size_bytes):
        return None, [f"Ukuran file import maksimal {max_size_label}."]

    try:
        return load_workbook(filename=BytesIO(file_obj.read()), data_only=True), []
    except Exception:
        return None, [EXCEL_READ_ERROR]


def get_header_maps(worksheet, headers: Sequence[str], required_headers: Sequence[str]):
    """Bangun map header Excel dan daftar kolom wajib yang hilang."""
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return None, None, ["Baris header pada file Excel tidak ditemukan."]

    normalized_headers = {
        normalize_import_header(value): index
        for index, value in enumerate(header_row)
        if string_cell(value)
    }
    header_aliases = {header: normalize_import_header(header) for header in headers}
    missing_headers = [header for header in required_headers if header_aliases[header] not in normalized_headers]
    if missing_headers:
        return None, None, [f'Kolom wajib belum tersedia: {", ".join(missing_headers)}.']

    return normalized_headers, header_aliases, []


def load_import_worksheet(
    file_obj,
    headers: Sequence[str],
    required_headers: Sequence[str],
    *,
    max_size_bytes: int,
    max_size_label: str = "7 MB",
):
    """Baca worksheet utama beserta peta header yang siap dipakai validator."""
    workbook, errors = load_import_workbook(
        file_obj,
        max_size_bytes=max_size_bytes,
        max_size_label=max_size_label,
    )
    if errors:
        return None, None, None, errors

    worksheet = workbook.active
    normalized_headers, header_aliases, header_errors = get_header_maps(
        worksheet,
        headers,
        required_headers,
    )
    if header_errors:
        return None, None, None, header_errors

    return worksheet, normalized_headers, header_aliases, []


def import_cell(row, header: str, normalized_headers: dict, header_aliases: dict) -> str:
    """Ambil nilai cell berdasarkan nama header template."""
    key = header_aliases[header]
    index = normalized_headers.get(key)
    return string_cell(row[index]) if index is not None and index < len(row) else ""
