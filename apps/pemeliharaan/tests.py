from datetime import date, datetime, timedelta
from io import BytesIO
from unittest.mock import patch

from django.core.files.uploadedfile import SimpleUploadedFile
from django.http import QueryDict
from django.test import TestCase
from django.urls import reverse
from django.utils.datastructures import MultiValueDict
from django.utils import timezone
from PIL import Image
from openpyxl import load_workbook

from apps.master_data.forms import BarangLaboratoriumForm
from apps.master_data.models import (
    BarangLaboratorium,
    KategoriBarangLaboratoriumChoices,
    KondisiBarangChoices,
    StatusBarangChoices,
)
from apps.notifikasi.models import Notification, NotificationCategory
from apps.operasional.models import TIM_LAYANAN_TEKNIS_NAME, TimKegiatan
from apps.pengguna.models import Role, User

from apps.pemeliharaan.forms import PemeliharaanForm, PemeliharaanVendorForm
from apps.pemeliharaan.models import (
    JenisFotoPemeliharaanChoices,
    KondisiPemeliharaanChoices,
    PemeliharaanFoto,
    PemeliharaanItem,
    PemeliharaanPengajuan,
    PemeliharaanVendor,
    StepPemeliharaanChoices,
    TindakanPerbaikanChoices,
)
from apps.pemeliharaan.pdf_utils import (
    TABLE_GAP,
    _repair_table,
    _signers,
    render_pemeliharaan_pdf,
)


class LaporanPemeliharaanExportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        role, _ = Role.objects.get_or_create(nama="Super Admin")
        cls.user = User.objects.create_user(
            username="super-admin-report-pemeliharaan",
            password="test-pass-123",
        )
        cls.user.safe_profile.role = role
        cls.user.safe_profile.save(update_fields=["role"])
        with patch("apps.master_data.signals.ensure_master_qr_code"):
            cls.alat = BarangLaboratorium.objects.create(
                nama_barang="Alat Laporan Pemeliharaan",
                tipe_merek_barang="Tipe Laporan",
                jenis_barang="Alat Uji",
                status_barang=StatusBarangChoices.NON_BMN,
                kategori_barang=KategoriBarangLaboratoriumChoices.DRONE,
            )
        cls.pengajuan = PemeliharaanPengajuan.objects.create(
            pemohon=cls.user,
            alat=cls.alat,
            current_step=StepPemeliharaanChoices.SELESAI,
            kepala_lab_at=timezone.now(),
        )
        PemeliharaanItem.objects.create(
            pengajuan=cls.pengajuan,
            komponen="Kamera",
            kondisi=KondisiPemeliharaanChoices.BAIK,
        )

    def setUp(self):
        self.client.force_login(self.user)

    def test_filter_default_tahun_berjalan_dan_export_workbook(self):
        current_year = str(timezone.localdate().year)
        response = self.client.get(reverse("pemeliharaan:laporan"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_report_year"], current_year)
        self.assertContains(response, self.pengajuan.nomor_pengajuan)

        export_response = self.client.get(
            reverse("pemeliharaan:laporan_export"),
            {"tahun": current_year},
        )
        self.assertEqual(export_response.status_code, 200)
        workbook = load_workbook(BytesIO(export_response.content), data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["Laporan Pemeliharaan", "Detail Komponen", "Data Vendor"],
        )
        self.assertEqual(
            workbook["Laporan Pemeliharaan"].cell(row=2, column=1).value,
            self.pengajuan.nomor_pengajuan,
        )


class PemeliharaanVerifikasiTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.users = {}
        for role_name in ("Super Admin", "Admin Lab", "Kepala Lab", "Pimpinan"):
            role, _ = Role.objects.get_or_create(nama=role_name)
            user = User.objects.create_user(
                username=f"pemeliharaan-{role_name.lower().replace(' ', '-')}",
                password="test-pass-123",
            )
            user.safe_profile.role = role
            profile_fields = ["role"]
            if role_name == "Pimpinan":
                user.safe_profile.jabatan = "Ketua Tim Layanan Teknis"
                profile_fields.append("jabatan")
            user.safe_profile.save(update_fields=profile_fields)
            cls.users[role_name] = user

        with patch("apps.master_data.signals.ensure_master_qr_code"):
            cls.alat = BarangLaboratorium.objects.create(
                nama_barang="Drone Verifikasi Pemeliharaan",
                tipe_merek_barang="DJI Uji",
                jenis_barang="Drone",
                status_barang=StatusBarangChoices.NON_BMN,
                kode_laboratorium="LAB-PEM-VER-001",
                lokasi_barang="Gudang Uji",
                kategori_barang=KategoriBarangLaboratoriumChoices.DRONE,
                komponen_pemeliharaan=["Kamera"],
            )

    def _login(self, role_name):
        self.client.force_login(self.users[role_name])

    def _create_pengajuan(self, tindakan):
        self.alat.refresh_from_db()
        pengajuan = PemeliharaanPengajuan.objects.create(
            pemohon=self.users["Admin Lab"],
            alat=self.alat,
        )
        PemeliharaanItem.objects.create(
            pengajuan=pengajuan,
            komponen="Kamera",
            kondisi=KondisiPemeliharaanChoices.PERLU_PERBAIKAN,
            tindakan_perbaikan=tindakan,
            tanggal_selesai_perbaikan=timezone.now()
            if tindakan == TindakanPerbaikanChoices.MANDIRI
            else None,
        )
        pengajuan.tandai_alat_dalam_pemeliharaan()
        return pengajuan

    def _create_pengajuan_baik(self):
        self.alat.refresh_from_db()
        pengajuan = PemeliharaanPengajuan.objects.create(
            pemohon=self.users["Admin Lab"],
            alat=self.alat,
        )
        PemeliharaanItem.objects.create(
            pengajuan=pengajuan,
            komponen="Kamera",
            kondisi=KondisiPemeliharaanChoices.BAIK,
        )
        pengajuan.tandai_alat_dalam_pemeliharaan()
        return pengajuan

    def _master_form_data(self, alat, components):
        data = QueryDict(mutable=True)
        data.update(
            {
                "status_barang": alat.status_barang,
                "nama_barang": alat.nama_barang,
                "tipe_merek_barang": alat.tipe_merek_barang,
                "jenis_barang": alat.jenis_barang,
                "kode_laboratorium": alat.kode_laboratorium,
                "satuan": alat.satuan,
                "tahun_perolehan": alat.tahun_perolehan or "",
                "kondisi_barang": alat.kondisi_barang or KondisiBarangChoices.BAIK,
                "lokasi_barang": alat.lokasi_barang,
                "kategori_barang": alat.kategori_barang,
                "catatan": alat.catatan,
            }
        )
        data.setlist("komponen_pemeliharaan", components)
        return data

    def _image_upload(self, name):
        output = BytesIO()
        Image.new("RGB", (120, 80), "white").save(output, format="PNG")
        return SimpleUploadedFile(name, output.getvalue(), content_type="image/png")

    def test_simpan_pengajuan_menandai_alat_dalam_pemeliharaan_dan_filter_dropdown(self):
        self._login("Admin Lab")
        response = self.client.post(
            reverse("pemeliharaan:tambah"),
            {
                "pilih_alat": str(self.alat.pk),
                "komponen_0": "Kamera",
                "kondisi_0": KondisiPemeliharaanChoices.BAIK,
                "dokumentasi_pemeriksaan": SimpleUploadedFile(
                    "pemeriksaan.png",
                    b"fake-image",
                    content_type="image/png",
                ),
            },
        )
        self.assertEqual(response.status_code, 302)

        self.alat.refresh_from_db()
        self.assertEqual(
            self.alat.kondisi_barang,
            KondisiBarangChoices.DALAM_PEMELIHARAAN,
        )
        self.assertIsNone(self.alat.tanggal_pemeliharaan)
        self.assertIsNone(self.alat.tanggal_perbaikan)

        response = self.client.get(reverse("pemeliharaan:tambah"))
        self.assertNotContains(response, self.alat.nama_barang)

    def test_alat_sedang_dipinjam_tidak_muncul_di_dropdown_pengajuan(self):
        with patch("apps.master_data.signals.ensure_master_qr_code"):
            alat_dipinjam = BarangLaboratorium.objects.create(
                nama_barang="Drone Sedang Dipinjam",
                tipe_merek_barang="DJI Dipinjam",
                jenis_barang="Drone",
                status_barang=StatusBarangChoices.NON_BMN,
                kode_laboratorium="LAB-PEM-DIPINJAM",
                lokasi_barang="Gudang Uji",
                kategori_barang=KategoriBarangLaboratoriumChoices.DRONE,
                komponen_pemeliharaan=["Kamera"],
                sedang_dipinjam=True,
            )

        self._login("Admin Lab")
        response = self.client.get(reverse("pemeliharaan:tambah"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, alat_dipinjam.nama_barang)
        self.assertNotContains(response, f'"{alat_dipinjam.pk}":')

    def test_form_menolak_post_manual_alat_sedang_dipinjam(self):
        with patch("apps.master_data.signals.ensure_master_qr_code"):
            alat_dipinjam = BarangLaboratorium.objects.create(
                nama_barang="Drone Manual Dipinjam",
                tipe_merek_barang="DJI Manual",
                jenis_barang="Drone",
                status_barang=StatusBarangChoices.NON_BMN,
                kode_laboratorium="LAB-PEM-MANUAL-DIPINJAM",
                lokasi_barang="Gudang Uji",
                kategori_barang=KategoriBarangLaboratoriumChoices.DRONE,
                komponen_pemeliharaan=["Kamera"],
                sedang_dipinjam=True,
            )

        form = PemeliharaanForm(
            data={"pilih_alat": str(alat_dipinjam.pk)},
            actor=self.users["Admin Lab"],
        )

        self.assertFalse(form.is_valid())
        self.assertIn("pilih_alat", form.errors)

    def test_form_menerima_tanggal_perbaikan_format_picker(self):
        tanggal_pemeriksaan = timezone.make_aware(datetime(2026, 6, 12, 9, 0))
        form = PemeliharaanForm(
            data={
                "pilih_alat": str(self.alat.pk),
                "komponen_0": "Kamera",
                "kondisi_0": KondisiPemeliharaanChoices.PERLU_PERBAIKAN,
                "tindakan_0": TindakanPerbaikanChoices.MANDIRI,
                "uraian_perbaikan_0": "Kalibrasi ulang kamera.",
                "tanggal_selesai_perbaikan_0": "12 Jun 2026",
            },
            files=MultiValueDict(
                {
                    "dokumentasi_pemeriksaan": [
                        SimpleUploadedFile(
                            "pemeriksaan.png",
                            b"fake-image",
                            content_type="image/png",
                        )
                    ],
                    "dokumentasi_perbaikan_0": [
                        SimpleUploadedFile(
                            "perbaikan.png",
                            b"fake-image",
                            content_type="image/png",
                        )
                    ],
                }
            ),
            actor=self.users["Admin Lab"],
            tanggal_pemeriksaan=tanggal_pemeriksaan,
        )

        self.assertTrue(form.is_valid(), form.errors)
        selesai = form.cleaned_items[0]["tanggal_selesai_perbaikan"]
        self.assertEqual(selesai.hour, 23)
        self.assertEqual(selesai.minute, 59)

    def test_edit_form_menampilkan_dokumentasi_tersimpan(self):
        pengajuan = self._create_pengajuan_baik()
        item = pengajuan.items.get()
        foto = PemeliharaanFoto.objects.create(
            item=item,
            jenis=JenisFotoPemeliharaanChoices.PEMERIKSAAN,
            foto=self._image_upload("edit-pemeriksaan.png"),
        )

        self._login("Admin Lab")
        response = self.client.get(reverse("pemeliharaan:edit", args=[pengajuan.pk]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "data-gallery-existing-item")
        self.assertContains(response, "data-gallery-remove")
        self.assertContains(response, 'title="Hapus foto"')
        self.assertContains(response, f'value="{foto.pk}"')
        self.assertContains(response, foto.foto.url)
        self.assertContains(response, "edit-pemeriksaan")
        self.assertContains(response, 'data-gallery-empty hidden', html=False)

    def test_edit_form_menghapus_dokumentasi_tersimpan(self):
        pengajuan = self._create_pengajuan_baik()
        item = pengajuan.items.get()
        foto_hapus = PemeliharaanFoto.objects.create(
            item=item,
            jenis=JenisFotoPemeliharaanChoices.PEMERIKSAAN,
            foto=self._image_upload("hapus-pemeriksaan.png"),
            urutan=1,
        )
        foto_sisa = PemeliharaanFoto.objects.create(
            item=item,
            jenis=JenisFotoPemeliharaanChoices.PEMERIKSAAN,
            foto=self._image_upload("sisa-pemeriksaan.png"),
            urutan=2,
        )
        deleted_name = foto_hapus.foto.name
        storage = foto_hapus.foto.storage

        self._login("Admin Lab")
        response = self.client.post(
            reverse("pemeliharaan:edit", args=[pengajuan.pk]),
            {
                "pilih_alat": str(self.alat.pk),
                "komponen_0": "Kamera",
                "kondisi_0": KondisiPemeliharaanChoices.BAIK,
                "hapus_foto_ids": str(foto_hapus.pk),
            },
        )

        self.assertRedirects(
            response,
            reverse("pemeliharaan:detail", args=[pengajuan.pk]),
            fetch_redirect_response=False,
        )
        self.assertFalse(PemeliharaanFoto.objects.filter(pk=foto_hapus.pk).exists())
        foto_sisa.refresh_from_db()
        self.assertEqual(foto_sisa.urutan, 1)
        self.assertFalse(storage.exists(deleted_name))

    def test_edit_form_menolak_hapus_foto_di_luar_pengajuan(self):
        pengajuan = self._create_pengajuan_baik()
        item = pengajuan.items.get()
        PemeliharaanFoto.objects.create(
            item=item,
            jenis=JenisFotoPemeliharaanChoices.PEMERIKSAAN,
            foto=self._image_upload("pemeriksaan-valid.png"),
        )
        form = PemeliharaanForm(
            data={
                "pilih_alat": str(self.alat.pk),
                "komponen_0": "Kamera",
                "kondisi_0": KondisiPemeliharaanChoices.BAIK,
                "hapus_foto_ids": ["999999999"],
            },
            actor=self.users["Admin Lab"],
            instance=pengajuan,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("tidak valid", str(form.errors["hapus_foto_ids"]))

    def test_edit_form_mewajibkan_pengganti_foto_terakhir_yang_dihapus(self):
        pengajuan = self._create_pengajuan_baik()
        item = pengajuan.items.get()
        foto = PemeliharaanFoto.objects.create(
            item=item,
            jenis=JenisFotoPemeliharaanChoices.PEMERIKSAAN,
            foto=self._image_upload("pemeriksaan-terakhir.png"),
        )
        form = PemeliharaanForm(
            data={
                "pilih_alat": str(self.alat.pk),
                "komponen_0": "Kamera",
                "kondisi_0": KondisiPemeliharaanChoices.BAIK,
                "hapus_foto_ids": [str(foto.pk)],
            },
            actor=self.users["Admin Lab"],
            instance=pengajuan,
        )

        self.assertFalse(form.is_valid())
        self.assertIn(
            "wajib diupload",
            str(form.errors["dokumentasi_pemeriksaan"]),
        )

    def test_edit_form_menolak_total_dokumentasi_lebih_dari_tiga(self):
        pengajuan = self._create_pengajuan_baik()
        item = pengajuan.items.get()
        for index in range(3):
            PemeliharaanFoto.objects.create(
                item=item,
                jenis=JenisFotoPemeliharaanChoices.PEMERIKSAAN,
                foto=self._image_upload(f"pemeriksaan-lama-{index}.png"),
            )

        form = PemeliharaanForm(
            data={
                "pilih_alat": str(self.alat.pk),
                "komponen_0": "Kamera",
                "kondisi_0": KondisiPemeliharaanChoices.BAIK,
            },
            files=MultiValueDict(
                {
                    "dokumentasi_pemeriksaan": [
                        self._image_upload("pemeriksaan-baru.png")
                    ],
                }
            ),
            actor=self.users["Admin Lab"],
            instance=pengajuan,
        )

        self.assertFalse(form.is_valid())
        self.assertIn("maksimal 3 foto", str(form.errors["dokumentasi_pemeriksaan"]))

    def test_master_data_menolak_hapus_komponen_dan_barang_aktif(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        pengajuan.tandai_alat_dalam_pemeliharaan()

        form = BarangLaboratoriumForm(
            data=self._master_form_data(self.alat, []),
            instance=self.alat,
        )
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data["komponen_pemeliharaan"], ["Kamera"])

        self._login("Admin Lab")
        response = self.client.post(
            reverse("master_data:hapus_barang_laboratorium", args=[self.alat.pk])
        )
        self.assertRedirects(
            response,
            reverse("master_data:data_barang_laboratorium"),
            fetch_redirect_response=False,
        )
        self.assertTrue(BarangLaboratorium.objects.filter(pk=self.alat.pk).exists())

    def test_alur_eksternal_melewati_verifikasi_vendor_dan_update_master(self):
        tanggal_pemeliharaan_awal = date(2025, 12, 20)
        self.alat.tanggal_pemeliharaan = tanggal_pemeliharaan_awal
        self.alat.save(update_fields=["tanggal_pemeliharaan", "updated_at"])
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.EKSTERNAL)
        self.assertEqual(pengajuan.jenis_pengajuan_label, "Perbaikan")
        item = pengajuan.items.get()
        item.uraian_kerusakan = "Kamera tidak dapat merekam."
        item.save(update_fields=["uraian_kerusakan"])
        PemeliharaanFoto.objects.create(
            item=item,
            jenis=JenisFotoPemeliharaanChoices.KERUSAKAN,
            foto=self._image_upload("kerusakan-kamera.png"),
        )
        PemeliharaanItem.objects.create(
            pengajuan=pengajuan,
            komponen="Sensor",
            kondisi=KondisiPemeliharaanChoices.PERLU_PERBAIKAN,
            tindakan_perbaikan=TindakanPerbaikanChoices.EKSTERNAL,
            uraian_kerusakan="Sensor tidak membaca data.",
        )
        PemeliharaanItem.objects.create(
            pengajuan=pengajuan,
            komponen="Baterai",
            kondisi=KondisiPemeliharaanChoices.PERLU_PERBAIKAN,
            tindakan_perbaikan=TindakanPerbaikanChoices.MANDIRI,
            tanggal_selesai_perbaikan=timezone.now() + timedelta(days=30),
        )

        self._login("Admin Lab")
        response = self.client.post(reverse("pemeliharaan:kirim", args=[pengajuan.pk]))
        self.assertRedirects(
            response,
            reverse("pemeliharaan:list"),
            fetch_redirect_response=False,
        )
        pengajuan.refresh_from_db()
        self.assertEqual(pengajuan.current_step, StepPemeliharaanChoices.KEPALA_LAB)
        self.alat.refresh_from_db()
        self.assertEqual(self.alat.tanggal_pemeliharaan, tanggal_pemeliharaan_awal)
        self.assertIsNone(self.alat.tanggal_perbaikan)

        self._login("Kepala Lab")
        response = self.client.get(
            reverse("verifikasi:detail_pemeliharaan", args=[pengajuan.pk])
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Pemeriksaan")
        self.assertContains(response, "Perbaikan")
        self.assertContains(response, "Tolak")
        self.assertContains(response, "Perbaiki")
        self.assertContains(response, "Setujui")
        self.assertContains(response, 'data-action="setujui"')
        self.assertContains(response, 'class="btn btn-primary js-verify-action"')
        self.assertContains(response, "btn-warning")

        response = self.client.post(
            reverse("verifikasi:detail_pemeliharaan", args=[pengajuan.pk]),
            {"aksi": "setujui", "catatan": ""},
        )
        self.assertRedirects(response, reverse("verifikasi:index"))
        pengajuan.refresh_from_db()
        self.assertEqual(pengajuan.current_step, StepPemeliharaanChoices.PIMPINAN)

        self._login("Pimpinan")
        response = self.client.post(
            reverse("verifikasi:detail_pemeliharaan", args=[pengajuan.pk]),
            {"aksi": "setujui", "catatan": ""},
        )
        self.assertRedirects(response, reverse("verifikasi:index"))
        pengajuan.refresh_from_db()
        self.assertEqual(pengajuan.current_step, StepPemeliharaanChoices.VENDOR_DRAFT)
        self.alat.refresh_from_db()
        self.assertEqual(
            self.alat.kondisi_barang,
            KondisiBarangChoices.DALAM_PEMELIHARAAN,
        )
        self.assertEqual(self.alat.tanggal_pemeliharaan, tanggal_pemeliharaan_awal)
        self.assertIsNone(self.alat.tanggal_perbaikan)

        self._login("Admin Lab")
        detail_response = self.client.get(
            reverse("pemeliharaan:detail", args=[pengajuan.pk])
        )
        self.assertContains(detail_response, "Input Data Vendor")
        self.assertContains(detail_response, "Edit Pengajuan")
        self.assertNotContains(
            detail_response,
            reverse("pemeliharaan:edit", args=[pengajuan.pk]),
        )
        edit_response = self.client.get(
            reverse("pemeliharaan:edit", args=[pengajuan.pk])
        )
        self.assertRedirects(
            edit_response,
            reverse("dashboard:index"),
            fetch_redirect_response=False,
        )

        vendor_response = self.client.get(
            reverse("pemeliharaan:vendor", args=[pengajuan.pk])
        )
        self.assertEqual(vendor_response.status_code, 200)
        self.assertContains(
            vendor_response,
            f'href="{reverse("pemeliharaan:list")}" class="btn btn-secondary">Kembali</a>',
            html=False,
        )
        self.assertContains(vendor_response, "Kamera tidak dapat merekam.")
        self.assertContains(vendor_response, "Sensor tidak membaca data.")
        self.assertContains(vendor_response, "kerusakan-kamera")

        tanggal_mulai = timezone.localdate()
        tanggal_selesai = tanggal_mulai + timedelta(days=5)
        response = self.client.post(
            reverse("pemeliharaan:vendor", args=[pengajuan.pk]),
            {
                "nama_vendor": "PT Servis Survei",
                "nama_pic": "Budi Teknisi",
                "nomor_hp_pic": "081234567890",
                "alamat": "Jl. Pengujian No. 1 Bandung",
                "tanggal_mulai": tanggal_mulai.isoformat(),
                "tanggal_selesai": tanggal_selesai.isoformat(),
            },
        )
        self.assertRedirects(
            response,
            reverse("pemeliharaan:detail", args=[pengajuan.pk]),
            fetch_redirect_response=False,
        )
        pengajuan.refresh_from_db()
        self.assertEqual(pengajuan.current_step, StepPemeliharaanChoices.VENDOR_DRAFT)
        vendor = PemeliharaanVendor.objects.get(pengajuan=pengajuan)
        self.assertEqual(vendor.nama_vendor, "PT Servis Survei")
        self.assertEqual(vendor.tanggal_selesai, tanggal_selesai)
        detail_response = self.client.get(
            reverse("pemeliharaan:detail", args=[pengajuan.pk])
        )
        detail_html = detail_response.content.decode()
        self.assertContains(
            detail_response,
            "<label>Kategori Pengajuan</label><p>Perbaikan</p>",
            html=True,
        )
        self.assertContains(detail_response, "pemeliharaan-vendor-grid__third")
        self.assertContains(detail_response, "pemeliharaan-vendor-grid__full")
        self.assertContains(detail_response, "pemeliharaan-vendor-grid__half")
        self.assertLess(detail_html.index("Nama Vendor"), detail_html.index("Alamat Vendor"))
        self.assertLess(
            detail_html.index("Alamat Vendor"),
            detail_html.index("Tanggal Perbaikan Mulai"),
        )

        response = self.client.post(
            reverse("pemeliharaan:vendor_kirim", args=[pengajuan.pk])
        )
        self.assertRedirects(
            response,
            reverse("pemeliharaan:list"),
            fetch_redirect_response=False,
        )
        pengajuan.refresh_from_db()
        self.assertEqual(
            pengajuan.current_step,
            StepPemeliharaanChoices.VENDOR_KEPALA_LAB,
        )
        self.assertTrue(
            Notification.objects.filter(
                recipient=self.users["Kepala Lab"],
                source_pemeliharaan=pengajuan,
                category=NotificationCategory.VERIFICATION,
                dedupe_key=(
                    f"pemeliharaan:{pengajuan.pk}:"
                    f"step:{StepPemeliharaanChoices.VENDOR_KEPALA_LAB.value}:flow:pending"
                ),
            ).exists()
        )

        self._login("Kepala Lab")
        self.client.post(
            reverse("verifikasi:detail_pemeliharaan", args=[pengajuan.pk]),
            {"aksi": "setujui", "catatan": ""},
        )
        pengajuan.refresh_from_db()
        self.assertEqual(
            pengajuan.current_step,
            StepPemeliharaanChoices.VENDOR_PIMPINAN,
        )

        self._login("Pimpinan")
        self.client.post(
            reverse("verifikasi:detail_pemeliharaan", args=[pengajuan.pk]),
            {"aksi": "setujui", "catatan": ""},
        )
        pengajuan.refresh_from_db()
        vendor.refresh_from_db()
        self.assertEqual(pengajuan.current_step, StepPemeliharaanChoices.SELESAI)
        self.assertEqual(vendor.kepala_lab_status, "approved")
        self.assertEqual(vendor.pimpinan_status, "approved")
        self.alat.refresh_from_db()
        self.assertEqual(self.alat.kondisi_barang, KondisiBarangChoices.BAIK)
        self.assertEqual(self.alat.tanggal_pemeliharaan, tanggal_pemeliharaan_awal)
        self.assertEqual(self.alat.tanggal_perbaikan, tanggal_selesai)

    def test_revisi_vendor_kembali_ke_draft_tanpa_membuka_edit_pengajuan(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.EKSTERNAL)
        pengajuan.current_step = StepPemeliharaanChoices.VENDOR_KEPALA_LAB
        pengajuan.save(update_fields=["current_step", "updated_at"])
        PemeliharaanVendor.objects.create(
            pengajuan=pengajuan,
            nama_vendor="Vendor Revisi",
            nama_pic="PIC Revisi",
            nomor_hp_pic="081200000000",
            alamat="Bandung",
            tanggal_mulai=timezone.localdate(),
            tanggal_selesai=timezone.localdate() + timedelta(days=2),
            submitted_at=timezone.now(),
        )

        self._login("Kepala Lab")
        response = self.client.post(
            reverse("verifikasi:detail_pemeliharaan", args=[pengajuan.pk]),
            {"aksi": "perbaiki", "catatan": "Perbaiki identitas vendor."},
        )
        self.assertRedirects(response, reverse("verifikasi:index"))
        pengajuan.refresh_from_db()
        self.assertEqual(pengajuan.current_step, StepPemeliharaanChoices.VENDOR_DRAFT)

        self._login("Admin Lab")
        response = self.client.get(reverse("pemeliharaan:edit", args=[pengajuan.pk]))
        self.assertRedirects(
            response,
            reverse("dashboard:index"),
            fetch_redirect_response=False,
        )
        response = self.client.get(reverse("pemeliharaan:vendor", args=[pengajuan.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Vendor Revisi")

    def test_form_vendor_mengunci_tanggal_selesai_dan_nomor_hp_hanya_angka(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.EKSTERNAL)
        pengajuan.current_step = StepPemeliharaanChoices.VENDOR_DRAFT
        pengajuan.save(update_fields=["current_step", "updated_at"])

        form = PemeliharaanVendorForm(pengajuan=pengajuan)
        selesai_attrs = form.fields["tanggal_selesai"].widget.attrs
        hp_attrs = form.fields["nomor_hp_pic"].widget.attrs
        self.assertEqual(selesai_attrs["disabled"], "disabled")
        self.assertEqual(selesai_attrs["aria-disabled"], "true")
        self.assertEqual(selesai_attrs["data-maint-min-source"], "id_tanggal_mulai")
        self.assertEqual(hp_attrs["inputmode"], "numeric")
        self.assertEqual(hp_attrs["pattern"], "[0-9]*")
        self.assertEqual(hp_attrs["data-maint-digits"], "true")

        tanggal_mulai = timezone.localdate()
        form = PemeliharaanVendorForm(
            data={
                "nama_vendor": "Vendor Validasi",
                "nama_pic": "PIC Validasi",
                "nomor_hp_pic": "0812-ABC",
                "alamat": "Bandung",
                "tanggal_mulai": tanggal_mulai.isoformat(),
                "tanggal_selesai": (tanggal_mulai - timedelta(days=1)).isoformat(),
            },
            pengajuan=pengajuan,
        )
        self.assertFalse(form.is_valid())
        self.assertIn(
            "Nomor HP PIC Vendor hanya boleh berisi angka.",
            form.errors["nomor_hp_pic"],
        )
        self.assertIn(
            "Tanggal perbaikan selesai tidak boleh lebih awal dari tanggal perbaikan mulai.",
            form.errors["tanggal_selesai"],
        )
        self.assertNotIn("disabled", form.fields["tanggal_selesai"].widget.attrs)

    def test_list_dan_detail_memakai_badge_seragam_dan_modal_hapus_custom(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.EKSTERNAL)
        self._login("Super Admin")

        response = self.client.get(reverse("pemeliharaan:list"))
        self.assertContains(
            response,
            '<span class="status-badge badge-warning">Perlu Perbaikan</span>',
            html=False,
        )
        self.assertContains(response, 'data-delete-modal="pemeliharaan"')
        self.assertContains(response, 'id="pemeliharaanDeleteModal"')
        self.assertNotContains(response, "onclick=\"return confirm(")

        response = self.client.get(
            reverse("pemeliharaan:detail", args=[pengajuan.pk])
        )
        self.assertContains(response, 'data-delete-modal="pemeliharaan"')
        self.assertContains(response, 'id="pemeliharaanDeleteModal"')
        self.assertContains(response, "Konfirmasi Hapus Pengajuan")
        self.assertNotContains(response, "onclick=\"return confirm(")

    def test_kirim_pengajuan_membuat_notifikasi_kepala_lab(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)

        self._login("Admin Lab")
        self.client.post(reverse("pemeliharaan:kirim", args=[pengajuan.pk]))

        notification = Notification.objects.get(
            recipient=self.users["Kepala Lab"],
            source_pemeliharaan=pengajuan,
            category=NotificationCategory.VERIFICATION,
            dedupe_key=(
                f"pemeliharaan:{pengajuan.pk}:"
                f"step:{StepPemeliharaanChoices.KEPALA_LAB.value}:flow:pending"
            ),
        )
        self.assertEqual(notification.title, f"Verifikasi Pemeliharaan {pengajuan.nomor_pengajuan}")
        self.assertIn("menunggu proses", notification.message)
        self.assertEqual(
            notification.link_url,
            reverse("verifikasi:detail_pemeliharaan", args=[pengajuan.pk]),
        )
        self.assertFalse(notification.is_read)

    def test_kirim_pengajuan_semua_baik_langsung_selesai(self):
        tanggal_pemeriksaan = timezone.make_aware(datetime(2026, 7, 5, 9, 30))
        tanggal_perbaikan_awal = date(2025, 11, 15)
        self.alat.tanggal_perbaikan = tanggal_perbaikan_awal
        self.alat.save(update_fields=["tanggal_perbaikan", "updated_at"])
        pengajuan = self._create_pengajuan_baik()
        pengajuan.tanggal_pemeriksaan = tanggal_pemeriksaan
        pengajuan.save(update_fields=["tanggal_pemeriksaan", "updated_at"])
        self.assertEqual(pengajuan.jenis_pengajuan_label, "Pemeliharaan")

        self._login("Admin Lab")
        response = self.client.post(reverse("pemeliharaan:kirim", args=[pengajuan.pk]))

        self.assertRedirects(
            response,
            reverse("pemeliharaan:list"),
            fetch_redirect_response=False,
        )
        pengajuan.refresh_from_db()
        self.assertEqual(pengajuan.current_step, StepPemeliharaanChoices.SELESAI)
        self.assertIsNotNone(pengajuan.submitted_at)
        self.assertFalse(
            Notification.objects.filter(
                recipient=self.users["Kepala Lab"],
                source_pemeliharaan=pengajuan,
                category=NotificationCategory.VERIFICATION,
            ).exists()
        )
        self.assertTrue(
            Notification.objects.filter(
                recipient=self.users["Admin Lab"],
                source_pemeliharaan=pengajuan,
                category=NotificationCategory.STATUS,
                dedupe_key=(
                    f"pemeliharaan:{pengajuan.pk}:"
                    f"status:{StepPemeliharaanChoices.SELESAI.value}"
                ),
            ).exists()
        )
        self.alat.refresh_from_db()
        self.assertEqual(self.alat.kondisi_barang, KondisiBarangChoices.BAIK)
        self.assertEqual(self.alat.tanggal_pemeliharaan, date(2026, 7, 5))
        self.assertEqual(self.alat.tanggal_perbaikan, tanggal_perbaikan_awal)
        detail_response = self.client.get(
            reverse("pemeliharaan:detail", args=[pengajuan.pk])
        )
        self.assertContains(
            detail_response,
            "<label>Kategori Pengajuan</label><p>Pemeliharaan</p>",
            html=True,
        )

    def test_persetujuan_final_menutup_notifikasi_lama_dan_mengabari_pemohon(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)

        self._login("Admin Lab")
        self.client.post(reverse("pemeliharaan:kirim", args=[pengajuan.pk]))

        self._login("Kepala Lab")
        self.client.post(
            reverse("verifikasi:detail_pemeliharaan", args=[pengajuan.pk]),
            {"aksi": "setujui", "catatan": ""},
        )

        kepala_notification = Notification.objects.get(
            recipient=self.users["Kepala Lab"],
            source_pemeliharaan=pengajuan,
            category=NotificationCategory.VERIFICATION,
        )
        self.assertTrue(kepala_notification.is_read)

        status_notification = Notification.objects.get(
            recipient=self.users["Admin Lab"],
            source_pemeliharaan=pengajuan,
            category=NotificationCategory.STATUS,
            dedupe_key=(
                f"pemeliharaan:{pengajuan.pk}:"
                f"status:{StepPemeliharaanChoices.SELESAI.value}"
            ),
        )
        self.assertEqual(status_notification.title, "Pemeliharaan Selesai")
        self.assertIn("telah selesai", status_notification.message)
        self.assertEqual(
            status_notification.link_url,
            reverse("pemeliharaan:detail", args=[pengajuan.pk]),
        )

    def test_alur_mandiri_selesai_di_kepala_lab(self):
        tanggal_pemeliharaan_awal = date(2025, 10, 10)
        tanggal_selesai = timezone.make_aware(datetime(2026, 7, 8, 15, 45))
        self.alat.kondisi_barang = KondisiBarangChoices.RUSAK
        self.alat.tanggal_pemeliharaan = tanggal_pemeliharaan_awal
        self.alat.save(
            update_fields=["kondisi_barang", "tanggal_pemeliharaan", "updated_at"]
        )
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        item = pengajuan.items.get()
        item.tanggal_selesai_perbaikan = tanggal_selesai
        item.save(update_fields=["tanggal_selesai_perbaikan"])
        self.assertEqual(pengajuan.jenis_pengajuan_label, "Perbaikan")

        self._login("Admin Lab")
        self.client.post(reverse("pemeliharaan:kirim", args=[pengajuan.pk]))

        self._login("Kepala Lab")
        response = self.client.post(
            reverse("verifikasi:detail_pemeliharaan", args=[pengajuan.pk]),
            {"aksi": "setujui", "catatan": ""},
        )
        self.assertRedirects(response, reverse("verifikasi:index"))
        pengajuan.refresh_from_db()
        self.assertEqual(pengajuan.current_step, StepPemeliharaanChoices.SELESAI)
        self.alat.refresh_from_db()
        self.assertEqual(self.alat.kondisi_barang, KondisiBarangChoices.BAIK)
        self.assertEqual(self.alat.tanggal_pemeliharaan, tanggal_pemeliharaan_awal)
        self.assertEqual(self.alat.tanggal_perbaikan, date(2026, 7, 8))

    def test_edit_detail_terkunci_setelah_dikirim_dan_aktif_saat_dikembalikan(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        self._login("Admin Lab")

        self.client.post(reverse("pemeliharaan:kirim", args=[pengajuan.pk]))
        response = self.client.get(reverse("pemeliharaan:detail", args=[pengajuan.pk]))

        self.assertContains(
            response,
            'title="Pengajuan tidak dapat diedit pada tahap ini" disabled>Edit Pengajuan</button>',
            html=False,
        )
        self.assertNotContains(response, reverse("pemeliharaan:edit", args=[pengajuan.pk]))

        pengajuan.current_step = StepPemeliharaanChoices.DIKEMBALIKAN
        pengajuan.save(update_fields=["current_step", "updated_at"])
        response = self.client.get(reverse("pemeliharaan:detail", args=[pengajuan.pk]))

        self.assertContains(response, reverse("pemeliharaan:edit", args=[pengajuan.pk]))
        self.assertNotContains(response, 'disabled>Edit Pengajuan</button>')

    def test_perbaiki_mengembalikan_ke_pemohon(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        self._login("Admin Lab")
        self.client.post(reverse("pemeliharaan:kirim", args=[pengajuan.pk]))

        self._login("Kepala Lab")
        response = self.client.post(
            reverse("verifikasi:detail_pemeliharaan", args=[pengajuan.pk]),
            {"aksi": "perbaiki", "catatan": "Lengkapi dokumentasi."},
        )
        self.assertRedirects(response, reverse("verifikasi:index"))
        pengajuan.refresh_from_db()
        self.assertEqual(pengajuan.current_step, StepPemeliharaanChoices.DIKEMBALIKAN)

    def test_tolak_kepala_lab_selesai_dan_memulihkan_kondisi_awal(self):
        self.alat.kondisi_barang = KondisiBarangChoices.RUSAK
        self.alat.tanggal_pemeliharaan = date(2025, 1, 10)
        self.alat.tanggal_perbaikan = date(2025, 1, 11)
        self.alat.save(
            update_fields=[
                "kondisi_barang",
                "tanggal_pemeliharaan",
                "tanggal_perbaikan",
                "updated_at",
            ]
        )
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        self._login("Admin Lab")
        self.client.post(reverse("pemeliharaan:kirim", args=[pengajuan.pk]))

        self._login("Kepala Lab")
        response = self.client.post(
            reverse("verifikasi:detail_pemeliharaan", args=[pengajuan.pk]),
            {"aksi": "tolak", "catatan": "Tidak layak diproses."},
        )

        self.assertRedirects(response, reverse("verifikasi:index"))
        pengajuan.refresh_from_db()
        self.assertEqual(pengajuan.current_step, StepPemeliharaanChoices.DITOLAK)
        self.assertEqual(pengajuan.kepala_lab_status, "rejected")
        self.alat.refresh_from_db()
        self.assertEqual(self.alat.kondisi_barang, KondisiBarangChoices.RUSAK)
        self.assertEqual(self.alat.tanggal_pemeliharaan, date(2025, 1, 10))
        self.assertEqual(self.alat.tanggal_perbaikan, date(2025, 1, 11))

    def test_hapus_pengajuan_memulihkan_seluruh_data_master_awal(self):
        self.alat.kondisi_barang = KondisiBarangChoices.RUSAK
        self.alat.tanggal_pemeliharaan = date(2025, 1, 10)
        self.alat.tanggal_perbaikan = date(2025, 1, 11)
        self.alat.save(
            update_fields=[
                "kondisi_barang",
                "tanggal_pemeliharaan",
                "tanggal_perbaikan",
                "updated_at",
            ]
        )
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        pengajuan.add_timeline("Pelaksana Pemeliharaan", "Data uji dibuat")

        self._login("Admin Lab")
        self.client.post(reverse("pemeliharaan:kirim", args=[pengajuan.pk]))
        self._login("Kepala Lab")
        self.client.post(
            reverse("verifikasi:detail_pemeliharaan", args=[pengajuan.pk]),
            {"aksi": "setujui", "catatan": ""},
        )
        self.alat.refresh_from_db()
        self.assertEqual(self.alat.kondisi_barang, KondisiBarangChoices.BAIK)
        self.assertEqual(self.alat.tanggal_pemeliharaan, date(2025, 1, 10))
        self.assertNotEqual(self.alat.tanggal_perbaikan, date(2025, 1, 11))

        self._login("Super Admin")
        response = self.client.post(reverse("pemeliharaan:hapus", args=[pengajuan.pk]))

        self.assertRedirects(
            response,
            reverse("pemeliharaan:list"),
            fetch_redirect_response=False,
        )
        self.assertFalse(PemeliharaanPengajuan.objects.filter(pk=pengajuan.pk).exists())
        self.assertFalse(PemeliharaanItem.objects.filter(pengajuan_id=pengajuan.pk).exists())
        self.alat.refresh_from_db()
        self.assertEqual(self.alat.kondisi_barang, KondisiBarangChoices.RUSAK)
        self.assertEqual(self.alat.tanggal_pemeliharaan, date(2025, 1, 10))
        self.assertEqual(self.alat.tanggal_perbaikan, date(2025, 1, 11))

    def test_hapus_pengajuan_lama_tidak_mengubah_hasil_pengajuan_lebih_baru(self):
        pengajuan_lama = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        pengajuan_lama.current_step = StepPemeliharaanChoices.SELESAI
        pengajuan_lama.save(update_fields=["current_step", "updated_at"])

        pengajuan_baru = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        pengajuan_baru.current_step = StepPemeliharaanChoices.SELESAI
        pengajuan_baru.save(update_fields=["current_step", "updated_at"])
        pengajuan_baru.catat_riwayat_alat_disetujui()
        pengajuan_baru.tandai_alat_baik_jika_selesai()
        self.alat.refresh_from_db()
        kondisi_terbaru = self.alat.kondisi_barang
        pemeliharaan_terbaru = self.alat.tanggal_pemeliharaan
        perbaikan_terbaru = self.alat.tanggal_perbaikan

        self._login("Super Admin")
        self.client.post(reverse("pemeliharaan:hapus", args=[pengajuan_lama.pk]))

        self.alat.refresh_from_db()
        self.assertEqual(self.alat.kondisi_barang, kondisi_terbaru)
        self.assertEqual(self.alat.tanggal_pemeliharaan, pemeliharaan_terbaru)
        self.assertEqual(self.alat.tanggal_perbaikan, perbaikan_terbaru)

    def test_hapus_laporan_hanya_super_admin_dan_kembali_ke_laporan(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        pengajuan.current_step = StepPemeliharaanChoices.SELESAI
        pengajuan.save(update_fields=["current_step", "updated_at"])
        delete_url = reverse("pemeliharaan:hapus", args=[pengajuan.pk])

        self._login("Admin Lab")
        response = self.client.post(delete_url)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(PemeliharaanPengajuan.objects.filter(pk=pengajuan.pk).exists())

        self._login("Super Admin")
        response = self.client.post(
            delete_url,
            HTTP_REFERER=f"http://testserver{reverse('pemeliharaan:laporan')}",
        )
        self.assertRedirects(
            response,
            reverse("pemeliharaan:laporan"),
            fetch_redirect_response=False,
        )
        self.assertFalse(PemeliharaanPengajuan.objects.filter(pk=pengajuan.pk).exists())

    def test_kondisi_dalam_perbaikan_tidak_lagi_muncul_di_pilihan_master_data(self):
        values = [value for value, _label in KondisiBarangChoices.choices]
        self.assertNotIn("Dalam Perbaikan", values)
        self.assertIn(KondisiBarangChoices.DALAM_PEMELIHARAAN, values)

    def test_pengajuan_final_masuk_laporan_dan_tidak_masuk_daftar_berjalan(self):
        selesai = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        selesai.current_step = StepPemeliharaanChoices.SELESAI
        selesai.kepala_lab_by = self.users["Kepala Lab"]
        selesai.kepala_lab_at = timezone.now()
        selesai.save(
            update_fields=[
                "current_step",
                "kepala_lab_by",
                "kepala_lab_at",
                "updated_at",
            ]
        )
        ditolak = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        ditolak.current_step = StepPemeliharaanChoices.DITOLAK
        ditolak.kepala_lab_by = self.users["Kepala Lab"]
        ditolak.kepala_lab_at = timezone.now()
        ditolak.save(
            update_fields=[
                "current_step",
                "kepala_lab_by",
                "kepala_lab_at",
                "updated_at",
            ]
        )
        berjalan = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)

        self._login("Admin Lab")
        response = self.client.get(reverse("pemeliharaan:list"))
        self.assertContains(response, berjalan.nomor_pengajuan)
        self.assertNotContains(response, selesai.nomor_pengajuan)
        self.assertNotContains(response, ditolak.nomor_pengajuan)

        response = self.client.get(reverse("pemeliharaan:laporan"))
        self.assertContains(response, selesai.nomor_pengajuan)
        self.assertContains(response, ditolak.nomor_pengajuan)
        self.assertContains(response, "Disetujui")
        self.assertContains(response, "Ditolak")
        self.assertNotContains(response, berjalan.nomor_pengajuan)
        self.assertNotContains(
            response,
            reverse("pemeliharaan:hapus", args=[selesai.pk]),
        )

        self._login("Super Admin")
        response = self.client.get(reverse("pemeliharaan:laporan"))
        self.assertContains(
            response,
            reverse("pemeliharaan:hapus", args=[selesai.pk]),
        )
        self.assertContains(response, 'title="Hapus laporan"', html=False)

    def test_pdf_laporan_hanya_tersedia_setelah_proses_final(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        self._login("Admin Lab")

        response = self.client.get(reverse("pemeliharaan:download_pdf", args=[pengajuan.pk]))
        self.assertEqual(response.status_code, 404)

        pengajuan.current_step = StepPemeliharaanChoices.SELESAI
        pengajuan.kepala_lab_by = self.users["Kepala Lab"]
        pengajuan.kepala_lab_at = timezone.now()
        pengajuan.add_timeline(
            "Kepala Lab",
            "Pengajuan pemeliharaan disetujui Kepala Lab dan dinyatakan selesai",
            self.users["Kepala Lab"],
        )
        item = pengajuan.items.get()
        PemeliharaanFoto.objects.create(
            item=item,
            jenis=JenisFotoPemeliharaanChoices.PEMERIKSAAN,
            foto=self._image_upload("pdf-pemeriksaan.png"),
        )
        PemeliharaanFoto.objects.create(
            item=item,
            jenis=JenisFotoPemeliharaanChoices.PERBAIKAN,
            foto=self._image_upload("pdf-perbaikan.png"),
        )
        pengajuan.save(
            update_fields=[
                "current_step",
                "kepala_lab_by",
                "kepala_lab_at",
                "updated_at",
            ]
        )

        response = self.client.get(reverse("pemeliharaan:download_pdf", args=[pengajuan.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertIn(
            f"laporan-pemeliharaan-{pengajuan.nomor_pengajuan}.pdf",
            response["Content-Disposition"],
        )
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_pdf_memakai_penanda_tangan_dari_data_master(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.EKSTERNAL)
        pengajuan.current_step = StepPemeliharaanChoices.SELESAI
        pengajuan.kepala_lab_by = self.users["Admin Lab"]
        pengajuan.pimpinan_by = self.users["Admin Lab"]
        pengajuan.save(
            update_fields=[
                "current_step",
                "kepala_lab_by",
                "pimpinan_by",
                "updated_at",
            ]
        )
        TimKegiatan.objects.update_or_create(
            nama_tim=TIM_LAYANAN_TEKNIS_NAME,
            defaults={"ketua_tim": self.users["Pimpinan"]},
        )

        signers = _signers(pengajuan)

        self.assertEqual(
            [user for _title, user in signers],
            [
                self.users["Admin Lab"],
                self.users["Kepala Lab"],
                self.users["Pimpinan"],
            ],
        )
        self.assertEqual(
            [title for title, _user in signers],
            [
                "Pelaksana Pemeliharaan,",
                "Kepala Laboratorium,",
                "Ketua Tim Layanan Teknis,",
            ],
        )

    def test_laporan_memakai_snapshot_status_barang(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.MANDIRI)
        self.assertEqual(pengajuan.snapshot_status_barang, StatusBarangChoices.NON_BMN)

        BarangLaboratorium.objects.filter(pk=self.alat.pk).update(
            status_barang=StatusBarangChoices.BMN
        )
        pengajuan.refresh_from_db()

        self.assertEqual(pengajuan.status_barang_label, StatusBarangChoices.NON_BMN)

    def test_pdf_tanpa_riwayat_proses_dan_foto_sesuai_bagian(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.EKSTERNAL)
        item = pengajuan.items.get()
        mandiri = PemeliharaanItem.objects.create(
            pengajuan=pengajuan,
            komponen="Baterai",
            kondisi=KondisiPemeliharaanChoices.PERLU_PERBAIKAN,
            tindakan_perbaikan=TindakanPerbaikanChoices.MANDIRI,
            uraian_perbaikan="Kalibrasi baterai.",
            tanggal_selesai_perbaikan=timezone.now(),
        )
        PemeliharaanFoto.objects.create(
            item=item,
            jenis=JenisFotoPemeliharaanChoices.PEMERIKSAAN,
            foto=self._image_upload("pemeriksaan.png"),
        )
        PemeliharaanFoto.objects.create(
            item=item,
            jenis=JenisFotoPemeliharaanChoices.KERUSAKAN,
            foto=self._image_upload("kerusakan.png"),
        )
        PemeliharaanFoto.objects.create(
            item=mandiri,
            jenis=JenisFotoPemeliharaanChoices.PERBAIKAN,
            foto=self._image_upload("perbaikan.png"),
        )

        with (
            patch("apps.pemeliharaan.pdf_utils.section") as section_mock,
            patch("apps.pemeliharaan.pdf_utils.info_table", return_value="") as info_mock,
            patch("apps.pemeliharaan.pdf_utils.data_table", return_value="") as data_mock,
            patch("apps.pemeliharaan.pdf_utils.photo_grid", return_value=[]) as photo_mock,
            patch("apps.pemeliharaan.pdf_utils._repair_table", return_value=[]) as repair_mock,
            patch("apps.pemeliharaan.pdf_utils.signature_list", return_value=[]) as sign_mock,
            patch("apps.pemeliharaan.pdf_utils.build_pdf"),
        ):
            render_pemeliharaan_pdf(BytesIO(), pengajuan, lambda value: "-")

        section_titles = [call.args[0] for call in section_mock.call_args_list]
        self.assertNotIn("D. Riwayat Proses", section_titles)
        self.assertNotIn("Dokumentasi Pemeriksaan", section_titles)
        self.assertIn("C. Tindak Lanjut Perbaikan", section_titles)
        self.assertEqual(photo_mock.call_count, 1)
        self.assertEqual(photo_mock.call_args.kwargs["header"], "Dokumentasi Pemeriksaan")
        self.assertEqual(info_mock.call_args.kwargs["valign"], "MIDDLE")
        self.assertEqual(data_mock.call_args.kwargs["valign"], "MIDDLE")
        self.assertEqual(
            [call.args[0] for call in repair_mock.call_args_list],
            ["Perbaikan Mandiri", "Perbaikan Eksternal"],
        )
        self.assertTrue(sign_mock.call_args.kwargs["centered"])
        self.assertEqual(sign_mock.call_args.kwargs["approval_label"], "Menyetujui:")

    def test_pdf_memberi_jarak_antar_tabel(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.EKSTERNAL)
        item = pengajuan.items.get()
        PemeliharaanItem.objects.create(
            pengajuan=pengajuan,
            komponen="Baterai",
            kondisi=KondisiPemeliharaanChoices.PERLU_PERBAIKAN,
            tindakan_perbaikan=TindakanPerbaikanChoices.MANDIRI,
            uraian_perbaikan="Kalibrasi baterai.",
            tanggal_selesai_perbaikan=timezone.now(),
        )
        PemeliharaanFoto.objects.create(
            item=item,
            jenis=JenisFotoPemeliharaanChoices.PEMERIKSAAN,
            foto=self._image_upload("pemeriksaan-gap.png"),
        )
        foto_table = object()
        mandiri_table = object()
        eksternal_table = object()

        with (
            patch("apps.pemeliharaan.pdf_utils.info_table", return_value=object()),
            patch("apps.pemeliharaan.pdf_utils.data_table", return_value=object()),
            patch(
                "apps.pemeliharaan.pdf_utils.photo_grid",
                return_value=[foto_table],
            ),
            patch(
                "apps.pemeliharaan.pdf_utils._repair_table",
                side_effect=[[mandiri_table], [eksternal_table]],
            ),
            patch("apps.pemeliharaan.pdf_utils.signature_list", return_value=[]),
            patch("apps.pemeliharaan.pdf_utils.build_pdf") as build_mock,
        ):
            render_pemeliharaan_pdf(BytesIO(), pengajuan, lambda value: "-")

        story = build_mock.call_args.args[2]
        foto_index = story.index(foto_table)
        mandiri_index = story.index(mandiri_table)
        eksternal_index = story.index(eksternal_table)
        self.assertEqual(story[foto_index - 1].height, TABLE_GAP)
        self.assertEqual(story[mandiri_index + 1].height, TABLE_GAP)
        self.assertEqual(eksternal_index, mandiri_index + 2)

    def test_tabel_perbaikan_terpisah_tanpa_kolom_tindakan(self):
        pengajuan = self._create_pengajuan(TindakanPerbaikanChoices.EKSTERNAL)
        eksternal = pengajuan.items.get()
        mandiri = PemeliharaanItem.objects.create(
            pengajuan=pengajuan,
            komponen="Baterai",
            kondisi=KondisiPemeliharaanChoices.PERLU_PERBAIKAN,
            tindakan_perbaikan=TindakanPerbaikanChoices.MANDIRI,
            uraian_perbaikan="Kalibrasi baterai.",
            tanggal_selesai_perbaikan=timezone.now(),
        )

        with patch("apps.pemeliharaan.pdf_utils.photo_cell", return_value="-"):
            tabel_mandiri = _repair_table(
                "Perbaikan Mandiri",
                [mandiri],
                JenisFotoPemeliharaanChoices.PERBAIKAN,
                lambda value: "-",
            )[0]
            tabel_eksternal = _repair_table(
                "Perbaikan Eksternal",
                [eksternal],
                JenisFotoPemeliharaanChoices.KERUSAKAN,
                lambda value: "-",
            )[0]

        judul_mandiri = tabel_mandiri._cellvalues[0][0].getPlainText()
        judul_eksternal = tabel_eksternal._cellvalues[0][0].getPlainText()
        header_mandiri = [cell.getPlainText() for cell in tabel_mandiri._cellvalues[1]]
        header_eksternal = [cell.getPlainText() for cell in tabel_eksternal._cellvalues[1]]
        self.assertEqual(judul_mandiri, "Perbaikan Mandiri")
        self.assertEqual(judul_eksternal, "Perbaikan Eksternal")
        self.assertNotIn("Tindakan", header_mandiri)
        self.assertNotIn("Tindakan", header_eksternal)
        self.assertIn("Dokumentasi Perbaikan", header_mandiri)
        self.assertIn("Dokumentasi Kerusakan", header_eksternal)
        self.assertEqual(tabel_mandiri.repeatRows, 2)
        self.assertEqual(tabel_eksternal.repeatRows, 2)
        self.assertIsNone(tabel_mandiri._argH[0])
        self.assertIsNone(tabel_mandiri._argH[1])
        self.assertIsNone(tabel_eksternal._argH[0])
        self.assertIsNone(tabel_eksternal._argH[1])
        self.assertTrue(
            all(
                cell.valign == "MIDDLE"
                for row in tabel_mandiri._cellStyles
                for cell in row
            )
        )
        self.assertTrue(
            all(
                cell.valign == "MIDDLE"
                for row in tabel_eksternal._cellStyles
                for cell in row
            )
        )
