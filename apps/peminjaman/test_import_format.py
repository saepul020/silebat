from datetime import date
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from apps.operasional.models import InstansiKlien, LayananKegiatan, SurveiKegiatan, TimKegiatan
from apps.peminjaman.import_riwayat import PEMINJAMAN_HEADERS, _validate_peminjaman_sheet
from apps.peminjaman.models import (
    PeminjamanBarangPenunjang,
    PeminjamanRequest,
    PengembalianBarangPenunjang,
    ReturnStepChoices,
    StepChoices,
)
from apps.pengguna.models import Role, User


class PeminjamanImportFormatTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.layanan = LayananKegiatan.objects.create(jenis_layanan="Layanan Uji")
        cls.tim = TimKegiatan.objects.create(nama_tim="Tim Layanan Teknis")
        cls.instansi = InstansiKlien.objects.create(
            nama_instansi="Instansi Uji",
            organisasi=InstansiKlien.OrganisasiChoices.EKSTERNAL_PU,
            alamat_instansi="Alamat Uji",
        )
        cls.survei = SurveiKegiatan.objects.create(jenis_survei="Borehole Camera")

    def _workbook(self, values):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Peminjaman"
        worksheet.append(PEMINJAMAN_HEADERS)
        worksheet.append([values.get(header, "") for header in PEMINJAMAN_HEADERS])
        return workbook

    def _base_values(self):
        return {
            "Nomor Pengajuan": "PMJ-260714-901",
            "Nama Peminjam": "Pengguna Historis",
            "Layanan Kegiatan": self.layanan.jenis_layanan,
            "Tim Kegiatan": self.tim.nama_tim,
            "Instansi Tujuan": self.instansi.nama_instansi,
            "Organisasi Instansi": self.instansi.organisasi,
            "Alamat Instansi": self.instansi.alamat_instansi,
            "Tanggal Mulai": date(2026, 7, 1),
            "Tanggal Selesai": date(2026, 7, 2),
            "Tanggal Pengembalian": date(2026, 7, 3),
            "Kegiatan Survei": self.survei.jenis_survei,
        }

    def test_field_relasi_dan_lainnya_terpisah_dan_valid(self):
        rows, errors = _validate_peminjaman_sheet(self._workbook(self._base_values()))

        self.assertEqual(errors, [])
        self.assertEqual(rows[0]["layanan_kegiatan"], self.layanan.jenis_layanan)
        self.assertEqual(rows[0]["layanan_kegiatan_lainnya"], "")
        self.assertEqual(rows[0]["instansi_tujuan_lainnya"], "")

    def test_pilihan_master_dan_lainnya_tidak_boleh_diisi_bersamaan(self):
        values = self._base_values()
        values["Layanan Kegiatan Lainnya"] = "Layanan Manual"
        _rows, errors = _validate_peminjaman_sheet(self._workbook(values))

        self.assertTrue(any("tepat salah satu antara Layanan Kegiatan" in error for error in errors))

    def test_export_dapat_dipakai_sebagai_format_import_dan_memuat_volume_pinjam(self):
        role, _ = Role.objects.get_or_create(nama="Super Admin")
        user = User.objects.create_user(username="export-import-peminjaman", password="pass-123")
        user.safe_profile.role = role
        user.safe_profile.save(update_fields=["role"])
        pengajuan = PeminjamanRequest.objects.create(
            nomor_pengajuan="PMJ-260714-902",
            peminjam=user,
            nama_peminjam="Pengguna Export",
            layanan_kegiatan=self.layanan,
            tim_kegiatan=self.tim,
            instansi_tujuan=self.instansi,
            tanggal_mulai=date(2026, 7, 1),
            tanggal_selesai=date(2026, 7, 2),
            current_step=StepChoices.APPROVED,
            return_current_step=ReturnStepChoices.COMPLETED,
            return_completed_at=timezone.now(),
        )
        pengajuan.kegiatan_survei.add(self.survei)
        PeminjamanBarangPenunjang.objects.create(
            pengajuan=pengajuan,
            barang=None,
            volume=4,
            snapshot_nama_barang="Payung Historis",
            snapshot_tipe_merek_barang="Tipe Uji",
            snapshot_kategori_barang="Penunjang Operasional Lapangan",
            snapshot_satuan="Buah",
        )
        PengembalianBarangPenunjang.objects.create(
            pengajuan=pengajuan,
            barang=None,
            qty_dikembalikan=4,
            snapshot_nama_barang="Payung Historis",
            snapshot_tipe_merek_barang="Tipe Uji",
            snapshot_kategori_barang="Penunjang Operasional Lapangan",
            snapshot_satuan="Buah",
        )
        self.client.force_login(user)

        response = self.client.get(reverse("peminjaman:export_laporan_peminjaman"))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        penunjang = workbook["Barang Penunjang"]
        header_map = {
            cell.value: cell.column for cell in penunjang[1]
        }
        self.assertEqual(
            penunjang.cell(row=2, column=header_map["Volume Dipinjam"]).value,
            4,
        )

        workbook["Peminjaman"].cell(row=2, column=1).value = "PMJ-260714-999"
        rows, errors = _validate_peminjaman_sheet(workbook)
        self.assertEqual(errors, [])
        self.assertEqual(rows[0]["nomor_pengajuan"], "PMJ-260714-999")
