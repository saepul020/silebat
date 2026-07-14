from datetime import date, datetime, time
import re
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.urls import reverse
from django.utils import timezone

from apps.core.permissions import ROLE_SUPER_ADMIN, get_role_name
from apps.master_data.models import (
    BahanOperasional,
    BarangLaboratorium,
    BarangPenunjangOperasional,
    PeralatanLaboratorium,
    KategoriBahanOperasionalChoices,
    KategoriBarangPenunjangChoices,
    KondisiBarangChoices,
    SatuanAsetChoices,
    SatuanBahanChoices,
    StatusBarangChoices,
)
from apps.operasional.models import (
    InstansiKlien,
    LayananKegiatan,
    SurveiKegiatan,
    TimKegiatan,
    normalize_survei_name,
    normalize_tim_kegiatan_name,
)

from .models import (
    DecisionChoices,
    PeminjamanBahanOperasional,
    PeminjamanBarangLaboratorium,
    PeminjamanBarangPenunjang,
    PeminjamanPeralatanLaboratorium,
    PeminjamanRequest,
    PeminjamanTimeline,
    PengembalianBahanOperasional,
    PengembalianBarangLaboratorium,
    PengembalianBarangPenunjang,
    PengembalianPeralatanLaboratorium,
    ReturnItemStatusChoices,
    ReturnStepChoices,
    StepChoices,
)
from .views import format_date_id, format_optional_numeric_display


IMPORT_RIWAYAT_SESSION_KEY = "riwayat_peminjaman_import_validated_rows"
IMPORT_RIWAYAT_MAX_SIZE = 7 * 1024 * 1024
NOMOR_LAMA_RE = re.compile(r"^PMJ-(\d{4})(\d{2})(\d{2})-(\d+)$", re.IGNORECASE)


def _normalize_nomor_pengajuan(value):
    nomor = (value or "").strip()
    match = NOMOR_LAMA_RE.match(nomor)
    if not match:
        return nomor
    tahun, bulan, tanggal, urut = match.groups()
    try:
        urut_baru = f"{int(urut):03d}"
    except (TypeError, ValueError):
        urut_baru = urut
    return f"PMJ-{tahun[-2:]}{bulan}{tanggal}-{urut_baru}"


def _clean_asset_value(value):
    text = str(value or "").strip()
    return text if text and text != "-" else ""


def _find_master_asset(model, row):
    """Coba hubungkan riwayat import ke master aktif tanpa mengubah stok."""
    kode_lab = _clean_asset_value(row.get("kode_laboratorium"))
    if kode_lab:
        found = model.objects.filter(kode_laboratorium__iexact=kode_lab).order_by("id").first()
        if found:
            return found

    kode_bmn = _clean_asset_value(row.get("kode_aset_bmn"))
    if kode_bmn:
        found = model.objects.filter(kode_aset_bmn__iexact=kode_bmn).order_by("id").first()
        if found:
            return found

    nama = _clean_asset_value(row.get("nama_barang"))
    tipe = _clean_asset_value(row.get("tipe_merek_barang"))
    jenis = _clean_asset_value(row.get("jenis_barang"))
    if nama and tipe:
        qs = model.objects.filter(
            nama_barang__iexact=nama,
            tipe_merek_barang__iexact=tipe,
        )
        if jenis:
            qs = qs.filter(jenis_barang__iexact=jenis)
        return qs.order_by("id").first()

    return None

PEMINJAMAN_SHEET = "Peminjaman"
LAB_SHEET = "Peralatan Survei"
PENUNJANG_SHEET = "Barang Penunjang"
BAHAN_SHEET = "Bahan Operasional"
PERALATAN_LAB_SHEET = "Peralatan Lab"

PEMINJAMAN_HEADERS = [
    "Nomor Pengajuan",
    "Username Peminjam",
    "Nama Peminjam",
    "Nomor Telepon",
    "Email",
    "NIP / NIK",
    "Alamat",
    "Layanan Kegiatan",
    "Tim Kegiatan",
    "Instansi Tujuan",
    "Organisasi Instansi",
    "Alamat Instansi",
    "Tanggal Mulai",
    "Tanggal Selesai",
    "Tanggal Pengembalian",
    "Kegiatan Survei",
    "Survei Lainnya",
    "Titik Geolistrik 1D",
    "Lintasan Geolistrik 2D",
    "Titik Kualitas Air",
    "Titik MAT",
    "Titik Pumping Test",
    "Titik Infiltrasi",
    "Titik Debit Air",
    "Lokasi Topografi",
    "Titik Borehole Camera",
    "Titik Logging",
]
PEMINJAMAN_REQUIRED_HEADERS = [
    "Nomor Pengajuan",
    "Nama Peminjam",
    "Tanggal Mulai",
    "Tanggal Selesai",
    "Tanggal Pengembalian",
]

LAB_HEADERS = [
    "Nomor Pengajuan",
    "Nama Barang",
    "Tipe / Merek Barang",
    "Jenis Barang",
    "Status Barang",
    "Kode Aset BMN",
    "Kode Laboratorium",
    "Volume",
    "Satuan",
    "Kondisi Barang",
    "Tahun Perolehan",
    "Asal Peminjaman",
    "Status Pengembalian",
    "Tujuan Transfer",
    "Catatan Pengembalian",
]
LAB_REQUIRED_HEADERS = ["Nomor Pengajuan", "Nama Barang", "Status Pengembalian"]

PENUNJANG_HEADERS = [
    "Nomor Pengajuan",
    "Nama Barang",
    "Tipe / Merek Barang",
    "Kategori Barang",
    "Volume Dipinjam",
    "Satuan",
    "Asal Peminjaman",
    "Dikembalikan",
    "Rusak",
    "Hilang",
    "Transfer",
    "Tujuan Transfer",
    "Catatan Pengembalian",
]
PENUNJANG_REQUIRED_HEADERS = ["Nomor Pengajuan", "Nama Barang", "Volume Dipinjam"]

BAHAN_HEADERS = [
    "Nomor Pengajuan",
    "Nama Barang",
    "Kategori Barang",
    "Volume Dipinjam",
    "Satuan",
    "Asal Peminjaman",
    "Sisa",
    "Transfer",
    "Tujuan Transfer",
    "Catatan Pengembalian",
]
BAHAN_REQUIRED_HEADERS = ["Nomor Pengajuan", "Nama Barang", "Volume Dipinjam"]

PERALATAN_LAB_HEADERS = [
    "Nomor Pengajuan",
    "Nama Barang",
    "Tipe / Merek Barang",
    "Jenis Barang",
    "Status Barang",
    "Kode Aset BMN",
    "Kode Laboratorium",
    "Volume Dipinjam",
    "Satuan",
    "Kondisi Barang",
    "Tahun Perolehan",
    "Asal Peminjaman",
    "Dikembalikan",
    "Rusak",
    "Hilang",
    "Transfer",
    "Tujuan Transfer",
    "Catatan Pengembalian",
]
PERALATAN_LAB_REQUIRED_HEADERS = ["Nomor Pengajuan", "Nama Barang", "Volume Dipinjam"]

PENGUKURAN_CONFIG = [
    ("titik_geolistrik_1d", "Titik Geolistrik 1D"),
    ("lintasan_geolistrik_2d", "Lintasan Geolistrik 2D"),
    ("titik_kualitas_air", "Titik Kualitas Air"),
    ("titik_mat", "Titik MAT"),
    ("titik_pumping_test", "Titik Pumping Test"),
    ("titik_infiltrasi", "Titik Infiltrasi"),
    ("titik_debit_air", "Titik Debit Air"),
    ("lokasi_topografi", "Lokasi Topografi"),
    ("titik_borehole", "Titik Borehole Camera"),
    ("titik_logging", "Titik Logging"),
]

PENGUKURAN_HEADER_TO_FIELD = {
    "Titik Geolistrik 1D": "titik_geolistrik_1d",
    "Lintasan Geolistrik 2D": "lintasan_geolistrik_2d",
    "Titik Kualitas Air": "titik_kualitas_air",
    "Titik MAT": "titik_mat",
    "Titik Pumping Test": "titik_pumping_test",
    "Titik Infiltrasi": "titik_infiltrasi",
    "Titik Debit Air": "titik_debit_air",
    "Lokasi Topografi": "lokasi_topografi",
    "Titik Borehole Camera": "titik_borehole",
    "Titik Logging": "titik_logging",
}


# ============================================================
# Generic import helpers: dibuat mengikuti pola validasi-simpan
# pada import Data Peralatan Survei Lapangan.
# ============================================================


def _is_super_admin_user(user):
    return get_role_name(user) == ROLE_SUPER_ADMIN


def _wants_import_json(request):
    return (
        request.headers.get("x-requested-with") == "XMLHttpRequest"
        or "application/json" in request.headers.get("accept", "")
    )


def _json_or_redirect(request, payload, redirect_to="peminjaman:laporan", status=200):
    if _wants_import_json(request):
        return JsonResponse(payload, status=status)
    if payload.get("ok"):
        messages.success(request, payload.get("message") or "Proses import berhasil.")
    else:
        for error in payload.get("errors") or [payload.get("message") or "Proses import belum berhasil."]:
            messages.error(request, error)
    return redirect(redirect_to)


def _deny_import_access(request):
    messages.error(request, "Fitur import riwayat peminjaman hanya dapat diakses oleh Super Admin.")
    return redirect("peminjaman:laporan")


def _normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_").replace("/", "_").replace("-", "_")


def _string_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _sheet_headers(worksheet):
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return {}, {}
    normalized = {
        _normalize_header(value): index
        for index, value in enumerate(header_row)
        if _string_cell(value)
    }
    raw = {index: _string_cell(value) for index, value in enumerate(header_row)}
    return normalized, raw


def _cell(row, header, normalized_headers, header_aliases):
    key = header_aliases[header]
    index = normalized_headers.get(key)
    return _string_cell(row[index]) if index is not None and index < len(row) else ""


def _missing_headers(normalized_headers, headers, required_headers):
    header_aliases = {header: _normalize_header(header) for header in headers}
    missing = [header for header in required_headers if header_aliases[header] not in normalized_headers]
    return missing, header_aliases


def _parse_int(value, label, *, min_value=0, blank_value=None):
    raw = _string_cell(value)
    if raw == "":
        return blank_value
    if raw.endswith(".0"):
        raw = raw[:-2]
    if not raw.isdigit():
        raise ValueError(f"{label} harus berupa angka bulat yang valid.")
    number = int(raw)
    if number < min_value:
        raise ValueError(f"{label} minimal {min_value}.")
    return number


def _parse_year(value, label="Tahun Perolehan"):
    raw = _string_cell(value)
    if not raw:
        return None
    if raw.endswith(".0"):
        raw = raw[:-2]
    if not raw.isdigit():
        raise ValueError(f"{label} harus berupa angka tahun yang valid.")
    year = int(raw)
    if year < 1900 or year > 2100:
        raise ValueError(f"{label} harus berada pada rentang 1900 sampai 2100.")
    return year


def _parse_date(value, label):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _string_cell(value)
    if not raw:
        raise ValueError(f"{label} wajib diisi.")
    raw = raw.replace(".", "/").strip()
    for date_format in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(raw, date_format).date()
        except ValueError:
            continue
    raise ValueError(f"{label} harus berupa tanggal valid. Contoh: 2025-01-31 atau 31/01/2025.")


def _aware_datetime_from_iso(date_iso, at_time=time(8, 0)):
    parsed = date.fromisoformat(date_iso)
    value = datetime.combine(parsed, at_time)
    if timezone.is_naive(value):
        return timezone.make_aware(value, timezone.get_current_timezone())
    return value


def _split_list(value):
    raw = _string_cell(value)
    if not raw:
        return []
    values = []
    for chunk in raw.replace(";", ",").split(","):
        text = chunk.strip()
        if text:
            values.append(text)
    return values


def _choice_values(choices):
    return {value for value, _label in choices}


def _choice_label_map(choices):
    data = {}
    for value, label in choices:
        data[str(value).strip().lower()] = value
        data[str(label).strip().lower()] = value
    return data


def _return_status_value(value):
    raw = _string_cell(value)
    if not raw:
        raise ValueError("Status Pengembalian wajib diisi.")
    aliases = _choice_label_map(ReturnItemStatusChoices.choices)
    key = raw.strip().lower()
    if key not in aliases:
        raise ValueError("Status Pengembalian harus salah satu dari: Dikembalikan, Rusak, Hilang, Transfer.")
    return aliases[key]


def _return_status_label(value):
    return dict(ReturnItemStatusChoices.choices).get(value, value or "-")


def _quantity_status_text(*, dikembalikan=0, rusak=0, hilang=0, transfer=0, sisa=0):
    tags = []
    if dikembalikan or sisa:
        tags.append("Dikembalikan")
    if transfer:
        tags.append("Transfer")
    if hilang:
        tags.append("Hilang")
    if rusak:
        tags.append("Rusak")
    return ", ".join(tags) if tags else "-"


def _load_workbook(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Library openpyxl belum tersedia. Jalankan: pip install openpyxl") from exc

    if not file_obj:
        return None, ["File Excel wajib diupload."]
    filename = getattr(file_obj, "name", "")
    if not filename.lower().endswith(".xlsx"):
        return None, ["Format file harus berupa Excel .xlsx."]
    if getattr(file_obj, "size", 0) > IMPORT_RIWAYAT_MAX_SIZE:
        return None, ["Ukuran file import maksimal 7 MB."]
    try:
        return load_workbook(filename=BytesIO(file_obj.read()), data_only=True), []
    except Exception:
        return None, ["File Excel tidak dapat dibaca. Pastikan file sesuai format .xlsx dan tidak rusak."]


def _sheet_or_error(workbook, sheet_name):
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
    if worksheet is None:
        return None, [f'Sheet "{sheet_name}" tidak ditemukan.']
    return worksheet, []


# ============================================================
# Validation helpers.
# ============================================================


def _validate_peminjaman_sheet(workbook):
    worksheet, errors = _sheet_or_error(workbook, PEMINJAMAN_SHEET)
    if errors:
        return [], errors

    normalized_headers, _raw_headers = _sheet_headers(worksheet)
    missing, header_aliases = _missing_headers(
        normalized_headers,
        PEMINJAMAN_HEADERS,
        PEMINJAMAN_REQUIRED_HEADERS,
    )
    if missing:
        return [], [f'Sheet "{PEMINJAMAN_SHEET}": Kolom wajib belum tersedia: {", ".join(missing)}.']

    rows = []
    errors = []
    seen_nomor = {}
    organisasi_values = _choice_values(InstansiKlien.OrganisasiChoices.choices)
    user_model = get_user_model()

    def cell(row, header):
        return _cell(row, header, normalized_headers, header_aliases)

    for excel_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_string_cell(value) for value in row):
            continue

        row_errors = []
        data = {
            "nomor_pengajuan": _normalize_nomor_pengajuan(cell(row, "Nomor Pengajuan")),
            "username_peminjam": cell(row, "Username Peminjam"),
            "nama_peminjam": cell(row, "Nama Peminjam"),
            "no_hp_peminjam": cell(row, "Nomor Telepon"),
            "email_peminjam": cell(row, "Email"),
            "nip_peminjam": cell(row, "NIP / NIK"),
            "alamat_peminjam": cell(row, "Alamat"),
            "layanan_kegiatan": cell(row, "Layanan Kegiatan"),
            "tim_kegiatan": normalize_tim_kegiatan_name(cell(row, "Tim Kegiatan")),
            "instansi_tujuan": cell(row, "Instansi Tujuan"),
            "organisasi_instansi": cell(row, "Organisasi Instansi") or InstansiKlien.OrganisasiChoices.EKSTERNAL_PU,
            "alamat_instansi": cell(row, "Alamat Instansi"),
            "kegiatan_survei": _split_list(cell(row, "Kegiatan Survei")),
            "survei_lainnya": cell(row, "Survei Lainnya"),
        }

        for header in PEMINJAMAN_REQUIRED_HEADERS:
            if not cell(row, header):
                row_errors.append(f"{header} wajib diisi.")

        nomor_key = data["nomor_pengajuan"].lower()
        if nomor_key:
            if nomor_key in seen_nomor:
                row_errors.append(f'Nomor Pengajuan duplikat dengan baris {seen_nomor[nomor_key]}.')
            else:
                seen_nomor[nomor_key] = excel_row_number
            if PeminjamanRequest.objects.filter(nomor_pengajuan__iexact=data["nomor_pengajuan"]).exists():
                row_errors.append("Nomor Pengajuan sudah terdaftar di database.")

        if data["username_peminjam"] and not user_model.objects.filter(username__iexact=data["username_peminjam"]).exists():
            row_errors.append("Username Peminjam tidak ditemukan di database. Kosongkan kolom ini jika peminjam historis belum memiliki akun.")

        if data["organisasi_instansi"] and data["organisasi_instansi"] not in organisasi_values:
            row_errors.append("Organisasi Instansi tidak sesuai pilihan yang tersedia.")

        try:
            tanggal_mulai = _parse_date(row[normalized_headers[header_aliases["Tanggal Mulai"]]], "Tanggal Mulai")
            tanggal_selesai = _parse_date(row[normalized_headers[header_aliases["Tanggal Selesai"]]], "Tanggal Selesai")
            tanggal_pengembalian = _parse_date(row[normalized_headers[header_aliases["Tanggal Pengembalian"]]], "Tanggal Pengembalian")
            data["tanggal_mulai"] = tanggal_mulai.isoformat()
            data["tanggal_selesai"] = tanggal_selesai.isoformat()
            data["tanggal_pengembalian"] = tanggal_pengembalian.isoformat()
            data["total_hari"] = (tanggal_selesai - tanggal_mulai).days + 1
            if tanggal_selesai < tanggal_mulai:
                row_errors.append("Tanggal Selesai tidak boleh lebih awal dari Tanggal Mulai.")
            if data["total_hari"] < 1:
                data["total_hari"] = 1
        except ValueError as exc:
            row_errors.append(str(exc))
            data["tanggal_mulai"] = ""
            data["tanggal_selesai"] = ""
            data["tanggal_pengembalian"] = ""
            data["total_hari"] = 1

        for header, field_name in PENGUKURAN_HEADER_TO_FIELD.items():
            try:
                data[field_name] = _parse_int(cell(row, header), header, min_value=0, blank_value=None)
            except ValueError as exc:
                row_errors.append(str(exc))
                data[field_name] = None

        if row_errors:
            errors.append(f'Sheet "{PEMINJAMAN_SHEET}" baris {excel_row_number}: ' + " ".join(row_errors))
            continue
        rows.append(data)

    if not rows and not errors:
        errors.append(f'Sheet "{PEMINJAMAN_SHEET}" belum memiliki data riwayat peminjaman untuk diimport.')

    return rows, errors


def _validate_lab_sheet(workbook, nomor_set):
    if LAB_SHEET not in workbook.sheetnames:
        return [], []
    worksheet = workbook[LAB_SHEET]
    normalized_headers, _raw_headers = _sheet_headers(worksheet)
    missing, header_aliases = _missing_headers(normalized_headers, LAB_HEADERS, LAB_REQUIRED_HEADERS)
    if missing:
        return [], [f'Sheet "{LAB_SHEET}": Kolom wajib belum tersedia: {", ".join(missing)}.']

    status_values = _choice_values(StatusBarangChoices.choices)
    satuan_values = _choice_values(SatuanAsetChoices.choices)
    kondisi_values = _choice_values(KondisiBarangChoices.choices)
    rows = []
    errors = []

    def cell(row, header):
        return _cell(row, header, normalized_headers, header_aliases)

    for excel_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_string_cell(value) for value in row):
            continue
        row_errors = []
        data = {
            "nomor_pengajuan": _normalize_nomor_pengajuan(cell(row, "Nomor Pengajuan")),
            "nama_barang": cell(row, "Nama Barang"),
            "tipe_merek_barang": cell(row, "Tipe / Merek Barang"),
            "jenis_barang": cell(row, "Jenis Barang"),
            "status_barang": cell(row, "Status Barang") or "-",
            "kode_aset_bmn": cell(row, "Kode Aset BMN"),
            "kode_laboratorium": cell(row, "Kode Laboratorium"),
            "satuan": cell(row, "Satuan") or "Unit",
            "kondisi_barang": cell(row, "Kondisi Barang") or "Baik",
            "asal_peminjaman": cell(row, "Asal Peminjaman") or "Laboratorium",
            "tujuan_transfer": cell(row, "Tujuan Transfer") or "-",
            "catatan_pengembalian": cell(row, "Catatan Pengembalian") or "-",
        }
        for header in LAB_REQUIRED_HEADERS:
            if not cell(row, header):
                row_errors.append(f"{header} wajib diisi.")
        if data["nomor_pengajuan"] and data["nomor_pengajuan"] not in nomor_set:
            row_errors.append("Nomor Pengajuan tidak ditemukan pada sheet Peminjaman.")
        if data["status_barang"] != "-" and data["status_barang"] not in status_values:
            row_errors.append("Status Barang tidak sesuai pilihan yang tersedia.")
        if data["satuan"] and data["satuan"] not in satuan_values:
            row_errors.append("Satuan tidak sesuai pilihan yang tersedia.")
        if data["kondisi_barang"] and data["kondisi_barang"] not in kondisi_values:
            row_errors.append("Kondisi Barang tidak sesuai pilihan yang tersedia.")
        try:
            data["volume"] = _parse_int(cell(row, "Volume"), "Volume", min_value=0, blank_value=1)
        except ValueError as exc:
            row_errors.append(str(exc))
            data["volume"] = 1
        try:
            data["tahun_perolehan"] = _parse_year(cell(row, "Tahun Perolehan"))
        except ValueError as exc:
            row_errors.append(str(exc))
            data["tahun_perolehan"] = None
        try:
            data["status_pengembalian"] = _return_status_value(cell(row, "Status Pengembalian"))
            data["status_pengembalian_display"] = _return_status_label(data["status_pengembalian"])
        except ValueError as exc:
            row_errors.append(str(exc))
            data["status_pengembalian"] = ReturnItemStatusChoices.DIKEMBALIKAN
            data["status_pengembalian_display"] = "Dikembalikan"

        if row_errors:
            errors.append(f'Sheet "{LAB_SHEET}" baris {excel_row_number}: ' + " ".join(row_errors))
            continue
        rows.append(data)
    return rows, errors


def _validate_penunjang_sheet(workbook, nomor_set):
    return _validate_qty_sheet(
        workbook,
        sheet_name=PENUNJANG_SHEET,
        headers=PENUNJANG_HEADERS,
        required_headers=PENUNJANG_REQUIRED_HEADERS,
        item_kind="penunjang",
        nomor_set=nomor_set,
    )


def _validate_bahan_sheet(workbook, nomor_set):
    return _validate_qty_sheet(
        workbook,
        sheet_name=BAHAN_SHEET,
        headers=BAHAN_HEADERS,
        required_headers=BAHAN_REQUIRED_HEADERS,
        item_kind="bahan",
        nomor_set=nomor_set,
    )


def _validate_peralatan_lab_sheet(workbook, nomor_set):
    return _validate_qty_sheet(
        workbook,
        sheet_name=PERALATAN_LAB_SHEET,
        headers=PERALATAN_LAB_HEADERS,
        required_headers=PERALATAN_LAB_REQUIRED_HEADERS,
        item_kind="peralatan_lab",
        nomor_set=nomor_set,
    )


def _validate_qty_sheet(workbook, *, sheet_name, headers, required_headers, item_kind, nomor_set):
    if sheet_name not in workbook.sheetnames:
        return [], []
    worksheet = workbook[sheet_name]
    normalized_headers, _raw_headers = _sheet_headers(worksheet)
    missing, header_aliases = _missing_headers(normalized_headers, headers, required_headers)
    if missing:
        return [], [f'Sheet "{sheet_name}": Kolom wajib belum tersedia: {", ".join(missing)}.']

    status_values = _choice_values(StatusBarangChoices.choices)
    satuan_aset_values = _choice_values(SatuanAsetChoices.choices)
    satuan_bahan_values = _choice_values(SatuanBahanChoices.choices)
    kondisi_values = _choice_values(KondisiBarangChoices.choices)
    kategori_penunjang_values = _choice_values(KategoriBarangPenunjangChoices.choices)
    kategori_bahan_values = _choice_values(KategoriBahanOperasionalChoices.choices)
    rows = []
    errors = []

    def cell(row, header):
        return _cell(row, header, normalized_headers, header_aliases)

    for excel_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_string_cell(value) for value in row):
            continue
        row_errors = []
        volume_header = "Volume Dipinjam"
        data = {
            "nomor_pengajuan": _normalize_nomor_pengajuan(cell(row, "Nomor Pengajuan")),
            "nama_barang": cell(row, "Nama Barang"),
            "kategori_barang": cell(row, "Kategori Barang") if "Kategori Barang" in headers else "-",
            "satuan": cell(row, "Satuan") or ("Buah" if item_kind == "bahan" else "Unit"),
            "asal_peminjaman": cell(row, "Asal Peminjaman") or "Laboratorium",
            "tujuan_transfer": cell(row, "Tujuan Transfer") or "-",
            "catatan_pengembalian": cell(row, "Catatan Pengembalian") or "-",
        }
        if item_kind in {"penunjang", "peralatan_lab"}:
            data["tipe_merek_barang"] = cell(row, "Tipe / Merek Barang")
        if item_kind == "peralatan_lab":
            data.update({
                "jenis_barang": cell(row, "Jenis Barang"),
                "status_barang": cell(row, "Status Barang") or "-",
                "kode_aset_bmn": cell(row, "Kode Aset BMN"),
                "kode_laboratorium": cell(row, "Kode Laboratorium"),
                "kondisi_barang": cell(row, "Kondisi Barang") or "Baik",
            })

        for header in required_headers:
            if not cell(row, header):
                row_errors.append(f"{header} wajib diisi.")
        if data["nomor_pengajuan"] and data["nomor_pengajuan"] not in nomor_set:
            row_errors.append("Nomor Pengajuan tidak ditemukan pada sheet Peminjaman.")

        try:
            data["volume_dipinjam"] = _parse_int(cell(row, volume_header), volume_header, min_value=1, blank_value=None)
        except ValueError as exc:
            row_errors.append(str(exc))
            data["volume_dipinjam"] = 0

        if item_kind == "penunjang" and data.get("kategori_barang") not in ("", "-"):
            if data["kategori_barang"] not in kategori_penunjang_values:
                row_errors.append("Kategori Barang tidak sesuai pilihan yang tersedia.")

        if item_kind == "bahan" and data.get("kategori_barang") not in ("", "-"):
            if data["kategori_barang"] not in kategori_bahan_values:
                row_errors.append("Kategori Barang tidak sesuai pilihan yang tersedia.")

        if item_kind == "bahan":
            for qty_header, key in (("Sisa", "qty_sisa"), ("Transfer", "qty_transfer")):
                try:
                    data[key] = _parse_int(cell(row, qty_header), qty_header, min_value=0, blank_value=0)
                except ValueError as exc:
                    row_errors.append(str(exc))
                    data[key] = 0
            if data["satuan"] and data["satuan"] not in satuan_bahan_values:
                row_errors.append("Satuan tidak sesuai pilihan yang tersedia.")
        else:
            for qty_header, key in (("Dikembalikan", "qty_dikembalikan"), ("Rusak", "qty_rusak"), ("Hilang", "qty_hilang"), ("Transfer", "qty_transfer")):
                try:
                    data[key] = _parse_int(cell(row, qty_header), qty_header, min_value=0, blank_value=0)
                except ValueError as exc:
                    row_errors.append(str(exc))
                    data[key] = 0
            if data["satuan"] and data["satuan"] not in satuan_aset_values:
                row_errors.append("Satuan tidak sesuai pilihan yang tersedia.")

        if item_kind == "peralatan_lab":
            if data["status_barang"] != "-" and data["status_barang"] not in status_values:
                row_errors.append("Status Barang tidak sesuai pilihan yang tersedia.")
            if data["kondisi_barang"] and data["kondisi_barang"] not in kondisi_values:
                row_errors.append("Kondisi Barang tidak sesuai pilihan yang tersedia.")
            try:
                data["tahun_perolehan"] = _parse_year(cell(row, "Tahun Perolehan"))
            except ValueError as exc:
                row_errors.append(str(exc))
                data["tahun_perolehan"] = None

        if item_kind == "bahan":
            total_processed = data.get("qty_sisa", 0) + data.get("qty_transfer", 0)
        else:
            total_processed = data.get("qty_dikembalikan", 0) + data.get("qty_rusak", 0) + data.get("qty_hilang", 0) + data.get("qty_transfer", 0)
        if total_processed > (data.get("volume_dipinjam") or 0):
            row_errors.append("Total nilai pengembalian tidak boleh lebih besar dari Volume Dipinjam.")

        if row_errors:
            errors.append(f'Sheet "{sheet_name}" baris {excel_row_number}: ' + " ".join(row_errors))
            continue
        rows.append(data)
    return rows, errors


def _validate_import_file(file_obj):
    workbook, errors = _load_workbook(file_obj)
    if errors:
        return {}, errors

    requests, request_errors = _validate_peminjaman_sheet(workbook)
    errors.extend(request_errors)
    nomor_set = {row["nomor_pengajuan"] for row in requests}

    lab_items, lab_errors = _validate_lab_sheet(workbook, nomor_set)
    penunjang_items, penunjang_errors = _validate_penunjang_sheet(workbook, nomor_set)
    bahan_items, bahan_errors = _validate_bahan_sheet(workbook, nomor_set)
    peralatan_lab_items, peralatan_lab_errors = _validate_peralatan_lab_sheet(workbook, nomor_set)
    errors.extend(lab_errors + penunjang_errors + bahan_errors + peralatan_lab_errors)

    payload = {
        "requests": requests,
        "lab_items": lab_items,
        "penunjang_items": penunjang_items,
        "bahan_items": bahan_items,
        "peralatan_lab_items": peralatan_lab_items,
    }
    return payload, errors


# ============================================================
# Save helpers.
# ============================================================


def _get_user_for_row(row, fallback_user):
    username = row.get("username_peminjam")
    if username:
        found = get_user_model().objects.filter(username__iexact=username).first()
        if found:
            return found
    return fallback_user


def _get_or_create_layanan(name):
    if not name:
        return None
    obj, _created = LayananKegiatan.objects.get_or_create(jenis_layanan=name)
    return obj


def _get_or_create_tim(name):
    if not name:
        return None
    normalized = normalize_tim_kegiatan_name(name)
    obj, _created = TimKegiatan.objects.get_or_create(nama_tim=normalized)
    return obj


def _get_or_create_instansi(name, organisasi, alamat=None):
    if not name:
        return None
    alamat_bersih = (alamat or "").strip() or "-"
    obj, created = InstansiKlien.objects.get_or_create(
        nama_instansi=name,
        defaults={
            "alamat_instansi": alamat_bersih,
            "organisasi": organisasi or InstansiKlien.OrganisasiChoices.EKSTERNAL_PU,
        },
    )
    update_fields = []
    if not created and organisasi and obj.organisasi != organisasi:
        obj.organisasi = organisasi
        update_fields.append("organisasi")
    if not created and alamat and obj.alamat_instansi in ("", "-"):
        obj.alamat_instansi = alamat_bersih
        update_fields.append("alamat_instansi")
    if update_fields:
        obj.save(update_fields=update_fields)
    return obj


def _get_or_create_survei(labels):
    objects = []
    for label in labels or []:
        label = normalize_survei_name(label)
        if not label:
            continue
        obj, _created = SurveiKegiatan.objects.get_or_create(jenis_survei=label)
        objects.append(obj)
    return objects


def _find_lab_master(row):
    kode_lab = row.get("kode_laboratorium")
    kode_bmn = row.get("kode_aset_bmn")
    nama = row.get("nama_barang")
    if kode_lab:
        found = BarangLaboratorium.objects.filter(kode_laboratorium__iexact=kode_lab).first()
        if found:
            return found
    if kode_bmn:
        found = BarangLaboratorium.objects.filter(kode_aset_bmn__iexact=kode_bmn).first()
        if found:
            return found
    if nama:
        return BarangLaboratorium.objects.filter(nama_barang__iexact=nama).first()
    return None


def _find_peralatan_lab_master(row):
    kode_lab = row.get("kode_laboratorium")
    kode_bmn = row.get("kode_aset_bmn")
    nama = row.get("nama_barang")
    if kode_lab:
        found = PeralatanLaboratorium.objects.filter(kode_laboratorium__iexact=kode_lab).first()
        if found:
            return found
    if kode_bmn:
        found = PeralatanLaboratorium.objects.filter(kode_aset_bmn__iexact=kode_bmn).first()
        if found:
            return found
    if nama:
        return PeralatanLaboratorium.objects.filter(nama_barang__iexact=nama).first()
    return None


def _find_penunjang_master(row):
    nama = row.get("nama_barang")
    if nama:
        return BarangPenunjangOperasional.objects.filter(nama_barang__iexact=nama).first()
    return None


def _find_bahan_master(row):
    nama = row.get("nama_barang")
    if nama:
        return BahanOperasional.objects.filter(nama_barang__iexact=nama).first()
    return None


def _update_snapshot(model, pk, **fields):
    clean_fields = {key: value for key, value in fields.items() if hasattr(model, key)}
    if clean_fields:
        model.objects.filter(pk=pk).update(**clean_fields)


def _snapshot_pengukuran(row):
    return [
        {
            "key": key,
            "label": label,
            "value": row.get(key),
            "display": format_optional_numeric_display(row.get(key)),
        }
        for key, label in PENGUKURAN_CONFIG
    ]


def _report_status_text(lab_items, penunjang_items, peralatan_lab_items, bahan_items):
    has_dikembalikan = False
    has_transfer = False
    has_hilang = False
    has_rusak = False
    for item in lab_items:
        has_dikembalikan = has_dikembalikan or item.get("status_pengembalian") == ReturnItemStatusChoices.DIKEMBALIKAN
        has_transfer = has_transfer or item.get("status_pengembalian") == ReturnItemStatusChoices.TRANSFER
        has_hilang = has_hilang or item.get("status_pengembalian") == ReturnItemStatusChoices.HILANG
        has_rusak = has_rusak or item.get("status_pengembalian") == ReturnItemStatusChoices.RUSAK
    for item in penunjang_items + peralatan_lab_items:
        has_dikembalikan = has_dikembalikan or item.get("qty_dikembalikan", 0) > 0
        has_transfer = has_transfer or item.get("qty_transfer", 0) > 0
        has_hilang = has_hilang or item.get("qty_hilang", 0) > 0
        has_rusak = has_rusak or item.get("qty_rusak", 0) > 0
    for item in bahan_items:
        has_dikembalikan = has_dikembalikan or item.get("qty_sisa", 0) > 0
        has_transfer = has_transfer or item.get("qty_transfer", 0) > 0
    tags = []
    if has_dikembalikan:
        tags.append("Dikembalikan")
    if has_transfer:
        tags.append("Transfer")
    if has_hilang:
        tags.append("Hilang")
    if has_rusak:
        tags.append("Rusak")
    return ", ".join(tags) if tags else "Belum ada data pengembalian"


def _build_report_snapshot(row, lab_items, penunjang_items, bahan_items, peralatan_lab_items):
    mulai = date.fromisoformat(row["tanggal_mulai"])
    selesai = date.fromisoformat(row["tanggal_selesai"])
    kembali = date.fromisoformat(row["tanggal_pengembalian"])
    return {
        "nomor_pengajuan": row.get("nomor_pengajuan") or "-",
        "submitted_at": f"{format_date_id(mulai)} | 08:00 WIB",
        "return_started_at": f"{format_date_id(kembali)} | 08:00 WIB",
        "return_completed_at": f"{format_date_id(kembali)} | 16:00 WIB",
        "proses_peminjaman_label": "Disetujui",
        "proses_pengembalian_label": "Pengembalian Selesai",
        "pengembalian_status_text": _report_status_text(lab_items, penunjang_items, peralatan_lab_items, bahan_items),
        "peminjam": {
            "nama": row.get("nama_peminjam") or "-",
            "nomor_telepon": row.get("no_hp_peminjam") or "-",
            "email": row.get("email_peminjam") or "-",
            "nip": row.get("nip_peminjam") or "-",
            "alamat": row.get("alamat_peminjam") or "-",
        },
        "kegiatan": {
            "layanan_kegiatan": row.get("layanan_kegiatan") or "-",
            "kegiatan_survei": list(row.get("kegiatan_survei") or []) + ([f"Lainnya: {row.get('survei_lainnya')}"] if row.get("survei_lainnya") else []),
            "tim_kegiatan": row.get("tim_kegiatan") or "-",
            "instansi_tujuan": row.get("instansi_tujuan") or "-",
            "alamat_instansi": row.get("alamat_instansi") or "-",
            "organisasi_instansi": row.get("organisasi_instansi") or "-",
            "mulai_tanggal": format_date_id(mulai),
            "selesai_tanggal": format_date_id(selesai),
            "total_hari": row.get("total_hari") or 1,
            "berkas_pendukung": "",
        },
        "pengukuran": _snapshot_pengukuran(row),
        "items": {
            "lab": lab_items,
            "penunjang": penunjang_items,
            "peralatan_lab": peralatan_lab_items,
            "bahan": bahan_items,
        },
    }


def _prepare_grouped_items(payload):
    grouped = {"lab": {}, "penunjang": {}, "bahan": {}, "peralatan_lab": {}}
    for section, key in (
        ("lab", "lab_items"),
        ("penunjang", "penunjang_items"),
        ("bahan", "bahan_items"),
        ("peralatan_lab", "peralatan_lab_items"),
    ):
        for item in payload.get(key) or []:
            grouped[section].setdefault(item.get("nomor_pengajuan"), []).append(item)
    return grouped


def _create_lab_items(pengajuan, rows):
    report_items = []
    for row in rows:
        barang = _find_master_asset(BarangLaboratorium, row)
        if barang and PeminjamanBarangLaboratorium.objects.filter(pengajuan=pengajuan, barang=barang).exists():
            barang = None
        borrowed = PeminjamanBarangLaboratorium.objects.create(
            pengajuan=pengajuan,
            barang=barang,
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_tipe_merek_barang=row.get("tipe_merek_barang") or "-",
            snapshot_jenis_barang=row.get("jenis_barang") or "-",
            snapshot_status_barang=row.get("status_barang") or "-",
            snapshot_kode_aset_bmn=row.get("kode_aset_bmn") or "-",
            snapshot_kode_laboratorium=row.get("kode_laboratorium") or "-",
            snapshot_volume=row.get("volume"),
            snapshot_satuan=row.get("satuan") or "-",
            snapshot_kondisi_barang=row.get("kondisi_barang") or "-",
            snapshot_tahun_perolehan=row.get("tahun_perolehan"),
        )
        _update_snapshot(
            PeminjamanBarangLaboratorium,
            borrowed.pk,
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_tipe_merek_barang=row.get("tipe_merek_barang") or "-",
            snapshot_jenis_barang=row.get("jenis_barang") or "-",
            snapshot_status_barang=row.get("status_barang") or "-",
            snapshot_kode_aset_bmn=row.get("kode_aset_bmn") or "-",
            snapshot_kode_laboratorium=row.get("kode_laboratorium") or "-",
            snapshot_volume=row.get("volume"),
            snapshot_satuan=row.get("satuan") or "-",
            snapshot_kondisi_barang=row.get("kondisi_barang") or "-",
            snapshot_tahun_perolehan=row.get("tahun_perolehan"),
        )
        returned = PengembalianBarangLaboratorium.objects.create(
            pengajuan=pengajuan,
            barang=barang,
            status=row.get("status_pengembalian") or ReturnItemStatusChoices.DIKEMBALIKAN,
            note="" if row.get("catatan_pengembalian") == "-" else row.get("catatan_pengembalian") or "",
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_tipe_merek_barang=row.get("tipe_merek_barang") or "-",
            snapshot_jenis_barang=row.get("jenis_barang") or "-",
            snapshot_status_barang=row.get("status_barang") or "-",
            snapshot_kode_aset_bmn=row.get("kode_aset_bmn") or "-",
            snapshot_kode_laboratorium=row.get("kode_laboratorium") or "-",
            snapshot_volume=row.get("volume"),
            snapshot_satuan=row.get("satuan") or "-",
            snapshot_kondisi_barang=row.get("kondisi_barang") or "-",
            snapshot_tahun_perolehan=row.get("tahun_perolehan"),
        )
        _update_snapshot(
            PengembalianBarangLaboratorium,
            returned.pk,
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_tipe_merek_barang=row.get("tipe_merek_barang") or "-",
            snapshot_jenis_barang=row.get("jenis_barang") or "-",
            snapshot_status_barang=row.get("status_barang") or "-",
            snapshot_kode_aset_bmn=row.get("kode_aset_bmn") or "-",
            snapshot_kode_laboratorium=row.get("kode_laboratorium") or "-",
            snapshot_volume=row.get("volume"),
            snapshot_satuan=row.get("satuan") or "-",
            snapshot_kondisi_barang=row.get("kondisi_barang") or "-",
            snapshot_tahun_perolehan=row.get("tahun_perolehan"),
        )
        report_items.append({
            "nama_barang": row.get("nama_barang") or "-",
            "tipe_merek_barang": row.get("tipe_merek_barang") or "-",
            "jenis_barang": row.get("jenis_barang") or "-",
            "status_barang": row.get("status_barang") or "-",
            "kode_aset_bmn": row.get("kode_aset_bmn") or "-",
            "kode_laboratorium": row.get("kode_laboratorium") or "-",
            "volume": row.get("volume") if row.get("volume") not in (None, "") else "-",
            "satuan": row.get("satuan") or "-",
            "kondisi_barang": row.get("kondisi_barang") or "-",
            "tahun_perolehan": row.get("tahun_perolehan") if row.get("tahun_perolehan") not in (None, "") else "-",
            "asal_peminjaman": row.get("asal_peminjaman") or "Laboratorium",
            "status_pengembalian": row.get("status_pengembalian_display") or _return_status_label(row.get("status_pengembalian")),
            "tujuan_transfer": row.get("tujuan_transfer") or "-",
            "catatan_pengembalian": row.get("catatan_pengembalian") or "-",
        })
    return report_items


def _create_penunjang_items(pengajuan, rows):
    report_items = []
    for row in rows:
        barang = None  # Riwayat import disimpan sebagai snapshot agar tidak mengubah/bergantung pada master data aktif.
        borrowed = PeminjamanBarangPenunjang.objects.create(
            pengajuan=pengajuan,
            barang=barang,
            volume=row.get("volume_dipinjam") or 1,
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_tipe_merek_barang=row.get("tipe_merek_barang") or "-",
            snapshot_kategori_barang=row.get("kategori_barang") or "-",
            snapshot_satuan=row.get("satuan") or "-",
        )
        _update_snapshot(
            PeminjamanBarangPenunjang,
            borrowed.pk,
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_tipe_merek_barang=row.get("tipe_merek_barang") or "-",
            snapshot_kategori_barang=row.get("kategori_barang") or "-",
            snapshot_satuan=row.get("satuan") or "-",
        )
        returned = PengembalianBarangPenunjang.objects.create(
            pengajuan=pengajuan,
            barang=barang,
            qty_dikembalikan=row.get("qty_dikembalikan") or 0,
            qty_rusak=row.get("qty_rusak") or 0,
            qty_hilang=row.get("qty_hilang") or 0,
            qty_transfer=row.get("qty_transfer") or 0,
            note="" if row.get("catatan_pengembalian") == "-" else row.get("catatan_pengembalian") or "",
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_tipe_merek_barang=row.get("tipe_merek_barang") or "-",
            snapshot_kategori_barang=row.get("kategori_barang") or "-",
            snapshot_satuan=row.get("satuan") or "-",
        )
        _update_snapshot(
            PengembalianBarangPenunjang,
            returned.pk,
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_tipe_merek_barang=row.get("tipe_merek_barang") or "-",
            snapshot_kategori_barang=row.get("kategori_barang") or "-",
            snapshot_satuan=row.get("satuan") or "-",
        )
        report_items.append({
            "nama_barang": row.get("nama_barang") or "-",
            "tipe_merek_barang": row.get("tipe_merek_barang") or "-",
            "kategori_barang": row.get("kategori_barang") or "-",
            "volume_dipinjam": row.get("volume_dipinjam") or 0,
            "satuan": row.get("satuan") or "-",
            "asal_peminjaman": row.get("asal_peminjaman") or "Laboratorium",
            "qty_dikembalikan": row.get("qty_dikembalikan") or 0,
            "qty_rusak": row.get("qty_rusak") or 0,
            "qty_hilang": row.get("qty_hilang") or 0,
            "qty_transfer": row.get("qty_transfer") or 0,
            "tujuan_transfer": row.get("tujuan_transfer") or "-",
            "catatan_pengembalian": row.get("catatan_pengembalian") or "-",
        })
    return report_items


def _create_bahan_items(pengajuan, rows):
    report_items = []
    for row in rows:
        bahan = None  # Riwayat import disimpan sebagai snapshot agar tidak mengubah/bergantung pada master data aktif.
        borrowed = PeminjamanBahanOperasional.objects.create(
            pengajuan=pengajuan,
            bahan=bahan,
            volume=row.get("volume_dipinjam") or 1,
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_kategori_barang=row.get("kategori_barang") or "-",
            snapshot_satuan=row.get("satuan") or "-",
        )
        _update_snapshot(
            PeminjamanBahanOperasional,
            borrowed.pk,
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_kategori_barang=row.get("kategori_barang") or "-",
            snapshot_satuan=row.get("satuan") or "-",
        )
        returned = PengembalianBahanOperasional.objects.create(
            pengajuan=pengajuan,
            bahan=bahan,
            qty_sisa=row.get("qty_sisa") or 0,
            qty_transfer=row.get("qty_transfer") or 0,
            note="" if row.get("catatan_pengembalian") == "-" else row.get("catatan_pengembalian") or "",
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_kategori_barang=row.get("kategori_barang") or "-",
            snapshot_satuan=row.get("satuan") or "-",
        )
        _update_snapshot(
            PengembalianBahanOperasional,
            returned.pk,
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_kategori_barang=row.get("kategori_barang") or "-",
            snapshot_satuan=row.get("satuan") or "-",
        )
        report_items.append({
            "nama_barang": row.get("nama_barang") or "-",
            "volume_dipinjam": row.get("volume_dipinjam") or 0,
            "satuan": row.get("satuan") or "-",
            "asal_peminjaman": row.get("asal_peminjaman") or "Laboratorium",
            "qty_sisa": row.get("qty_sisa") or 0,
            "qty_transfer": row.get("qty_transfer") or 0,
            "tujuan_transfer": row.get("tujuan_transfer") or "-",
            "catatan_pengembalian": row.get("catatan_pengembalian") or "-",
        })
    return report_items


def _create_peralatan_lab_items(pengajuan, rows):
    report_items = []
    for row in rows:
        barang = _find_master_asset(PeralatanLaboratorium, row)
        if barang and PeminjamanPeralatanLaboratorium.objects.filter(pengajuan=pengajuan, barang=barang).exists():
            barang = None
        borrowed = PeminjamanPeralatanLaboratorium.objects.create(
            pengajuan=pengajuan,
            barang=barang,
            volume=row.get("volume_dipinjam") or 1,
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_tipe_merek_barang=row.get("tipe_merek_barang") or "-",
            snapshot_jenis_barang=row.get("jenis_barang") or "-",
            snapshot_status_barang=row.get("status_barang") or "-",
            snapshot_kode_aset_bmn=row.get("kode_aset_bmn") or "-",
            snapshot_kode_laboratorium=row.get("kode_laboratorium") or "-",
            snapshot_volume=row.get("volume_dipinjam"),
            snapshot_satuan=row.get("satuan") or "-",
            snapshot_kondisi_barang=row.get("kondisi_barang") or "-",
            snapshot_tahun_perolehan=row.get("tahun_perolehan"),
        )
        _update_snapshot(
            PeminjamanPeralatanLaboratorium,
            borrowed.pk,
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_tipe_merek_barang=row.get("tipe_merek_barang") or "-",
            snapshot_jenis_barang=row.get("jenis_barang") or "-",
            snapshot_status_barang=row.get("status_barang") or "-",
            snapshot_kode_aset_bmn=row.get("kode_aset_bmn") or "-",
            snapshot_kode_laboratorium=row.get("kode_laboratorium") or "-",
            snapshot_volume=row.get("volume_dipinjam"),
            snapshot_satuan=row.get("satuan") or "-",
            snapshot_kondisi_barang=row.get("kondisi_barang") or "-",
            snapshot_tahun_perolehan=row.get("tahun_perolehan"),
        )
        returned = PengembalianPeralatanLaboratorium.objects.create(
            pengajuan=pengajuan,
            barang=barang,
            qty_dikembalikan=row.get("qty_dikembalikan") or 0,
            qty_rusak=row.get("qty_rusak") or 0,
            qty_hilang=row.get("qty_hilang") or 0,
            qty_transfer=row.get("qty_transfer") or 0,
            note="" if row.get("catatan_pengembalian") == "-" else row.get("catatan_pengembalian") or "",
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_tipe_merek_barang=row.get("tipe_merek_barang") or "-",
            snapshot_jenis_barang=row.get("jenis_barang") or "-",
            snapshot_status_barang=row.get("status_barang") or "-",
            snapshot_kode_aset_bmn=row.get("kode_aset_bmn") or "-",
            snapshot_kode_laboratorium=row.get("kode_laboratorium") or "-",
            snapshot_volume=row.get("volume_dipinjam"),
            snapshot_satuan=row.get("satuan") or "-",
            snapshot_kondisi_barang=row.get("kondisi_barang") or "-",
            snapshot_tahun_perolehan=row.get("tahun_perolehan"),
        )
        _update_snapshot(
            PengembalianPeralatanLaboratorium,
            returned.pk,
            snapshot_nama_barang=row.get("nama_barang") or "-",
            snapshot_tipe_merek_barang=row.get("tipe_merek_barang") or "-",
            snapshot_jenis_barang=row.get("jenis_barang") or "-",
            snapshot_status_barang=row.get("status_barang") or "-",
            snapshot_kode_aset_bmn=row.get("kode_aset_bmn") or "-",
            snapshot_kode_laboratorium=row.get("kode_laboratorium") or "-",
            snapshot_volume=row.get("volume_dipinjam"),
            snapshot_satuan=row.get("satuan") or "-",
            snapshot_kondisi_barang=row.get("kondisi_barang") or "-",
            snapshot_tahun_perolehan=row.get("tahun_perolehan"),
        )
        report_items.append({
            "nama_barang": row.get("nama_barang") or "-",
            "tipe_merek_barang": row.get("tipe_merek_barang") or "-",
            "jenis_barang": row.get("jenis_barang") or "-",
            "status_barang": row.get("status_barang") or "-",
            "kode_aset_bmn": row.get("kode_aset_bmn") or "-",
            "kode_laboratorium": row.get("kode_laboratorium") or "-",
            "volume_dipinjam": row.get("volume_dipinjam") or 0,
            "satuan": row.get("satuan") or "-",
            "kondisi_barang": row.get("kondisi_barang") or "-",
            "tahun_perolehan": row.get("tahun_perolehan") if row.get("tahun_perolehan") not in (None, "") else "-",
            "asal_peminjaman": row.get("asal_peminjaman") or "Laboratorium",
            "qty_dikembalikan": row.get("qty_dikembalikan") or 0,
            "qty_rusak": row.get("qty_rusak") or 0,
            "qty_hilang": row.get("qty_hilang") or 0,
            "qty_transfer": row.get("qty_transfer") or 0,
            "tujuan_transfer": row.get("tujuan_transfer") or "-",
            "catatan_pengembalian": row.get("catatan_pengembalian") or "-",
        })
    return report_items


def _save_import_payload(payload, actor):
    requests = payload.get("requests") or []
    nomor_values = [row.get("nomor_pengajuan") for row in requests]
    existing = list(PeminjamanRequest.objects.filter(nomor_pengajuan__in=nomor_values).values_list("nomor_pengajuan", flat=True))
    if existing:
        return 0, ["Nomor Pengajuan sudah terdaftar di database: " + ", ".join(existing)]

    grouped = _prepare_grouped_items(payload)
    saved = 0

    with transaction.atomic():
        for row in requests:
            completed_at = _aware_datetime_from_iso(row["tanggal_pengembalian"], time(16, 0))
            submitted_at = _aware_datetime_from_iso(row["tanggal_mulai"], time(8, 0))
            peminjam = _get_user_for_row(row, actor)
            pengajuan = PeminjamanRequest.objects.create(
                nomor_pengajuan=row["nomor_pengajuan"],
                peminjam=peminjam,
                nama_peminjam=row.get("nama_peminjam") or "-",
                no_hp_peminjam=row.get("no_hp_peminjam") or "",
                email_peminjam=row.get("email_peminjam") or "",
                alamat_peminjam=row.get("alamat_peminjam") or "",
                nip_peminjam=row.get("nip_peminjam") or "",
                layanan_kegiatan=_get_or_create_layanan(row.get("layanan_kegiatan")),
                tim_kegiatan=_get_or_create_tim(row.get("tim_kegiatan")),
                instansi_tujuan=_get_or_create_instansi(
                    row.get("instansi_tujuan"),
                    row.get("organisasi_instansi"),
                    row.get("alamat_instansi"),
                ),
                instansi_tujuan_lainnya="" if row.get("instansi_tujuan") else row.get("instansi_tujuan", ""),
                tanggal_mulai=date.fromisoformat(row["tanggal_mulai"]),
                tanggal_selesai=date.fromisoformat(row["tanggal_selesai"]),
                total_hari=row.get("total_hari") or 1,
                current_step=StepChoices.APPROVED,
                aset_sudah_dialokasikan=False,
                admin_lab_status=DecisionChoices.APPROVED,
                teknisi_lab_status=DecisionChoices.APPROVED,
                user_verification_status=DecisionChoices.APPROVED,
                kepala_lab_status=DecisionChoices.APPROVED,
                pimpinan_status=DecisionChoices.APPROVED,
                admin_lab_by=actor,
                teknisi_lab_by=actor,
                kepala_lab_by=actor,
                pimpinan_by=actor,
                admin_lab_at=completed_at,
                teknisi_lab_at=completed_at,
                user_verification_at=completed_at,
                kepala_lab_at=completed_at,
                pimpinan_at=completed_at,
                return_current_step=ReturnStepChoices.COMPLETED,
                return_started_at=_aware_datetime_from_iso(row["tanggal_pengembalian"], time(8, 0)),
                return_user_verification_status=DecisionChoices.APPROVED,
                return_user_verification_at=completed_at,
                return_completed_at=completed_at,
                return_inventory_applied=True,
                survei_lainnya=row.get("survei_lainnya") or "",
                titik_geolistrik_1d=row.get("titik_geolistrik_1d"),
                lintasan_geolistrik_2d=row.get("lintasan_geolistrik_2d"),
                titik_kualitas_air=row.get("titik_kualitas_air"),
                titik_mat=row.get("titik_mat"),
                titik_pumping_test=row.get("titik_pumping_test"),
                titik_infiltrasi=row.get("titik_infiltrasi"),
                titik_debit_air=row.get("titik_debit_air"),
                lokasi_topografi=row.get("lokasi_topografi"),
                titik_borehole=row.get("titik_borehole"),
                titik_logging=row.get("titik_logging"),
            )
            PeminjamanRequest.objects.filter(pk=pengajuan.pk).update(submitted_at=submitted_at)
            pengajuan.kegiatan_survei.set(_get_or_create_survei(row.get("kegiatan_survei")))

            nomor = row["nomor_pengajuan"]
            lab_report = _create_lab_items(pengajuan, grouped["lab"].get(nomor, []))
            penunjang_report = _create_penunjang_items(pengajuan, grouped["penunjang"].get(nomor, []))
            bahan_report = _create_bahan_items(pengajuan, grouped["bahan"].get(nomor, []))
            peralatan_lab_report = _create_peralatan_lab_items(pengajuan, grouped["peralatan_lab"].get(nomor, []))

            pengajuan.report_snapshot = _build_report_snapshot(
                row,
                lab_report,
                penunjang_report,
                bahan_report,
                peralatan_lab_report,
            )
            pengajuan.save(update_fields=["report_snapshot", "updated_at"])
            PeminjamanTimeline.objects.create(
                pengajuan=pengajuan,
                stage="Import Riwayat",
                action="Riwayat peminjaman diimport dari Excel",
                actor=actor,
                note="Data historis diimport tanpa mengubah stok aktif master data.",
            )
            saved += 1

    return saved, []


# ============================================================
# Views.
# ============================================================


@login_required
def import_riwayat_peminjaman(request):
    if not _is_super_admin_user(request.user):
        return _deny_import_access(request)
    if request.method != "POST":
        return redirect("peminjaman:laporan")

    action = request.POST.get("import_action")

    if action == "cancel":
        request.session.pop(IMPORT_RIWAYAT_SESSION_KEY, None)
        return _json_or_redirect(request, {"ok": True, "cancelled": True})

    if action == "validate":
        payload, errors = _validate_import_file(request.FILES.get("file_import"))
        total_rows = len(payload.get("requests") or []) if payload else 0
        if errors:
            request.session.pop(IMPORT_RIWAYAT_SESSION_KEY, None)
        else:
            request.session[IMPORT_RIWAYAT_SESSION_KEY] = payload
            request.session.modified = True
        return _json_or_redirect(
            request,
            {
                "ok": not bool(errors),
                "validated": not bool(errors),
                "can_save": bool(total_rows) and not errors,
                "total_rows": total_rows,
                "errors": errors,
                "message": f"Validasi berhasil. {total_rows} riwayat peminjaman siap disimpan." if not errors else "Validasi belum berhasil.",
            },
        )

    if action == "save":
        payload = request.session.get(IMPORT_RIWAYAT_SESSION_KEY) or {}
        if not payload:
            return _json_or_redirect(
                request,
                {
                    "ok": False,
                    "saved": False,
                    "can_save": False,
                    "errors": ["Data validasi tidak ditemukan atau sudah kedaluwarsa. Lakukan Validasi Data terlebih dahulu."],
                    "message": "Data belum dapat disimpan.",
                },
            )
        total_saved, save_errors = _save_import_payload(payload, request.user)
        if save_errors:
            request.session.pop(IMPORT_RIWAYAT_SESSION_KEY, None)
            return _json_or_redirect(
                request,
                {
                    "ok": False,
                    "saved": False,
                    "can_save": False,
                    "errors": save_errors,
                    "message": "Data tidak disimpan karena ditemukan duplikasi terbaru di database.",
                },
            )
        request.session.pop(IMPORT_RIWAYAT_SESSION_KEY, None)
        return _json_or_redirect(
            request,
            {
                "ok": True,
                "saved": True,
                "can_save": False,
                "total_rows": total_saved,
                "errors": [],
                "message": f"{total_saved} riwayat peminjaman berhasil diimport.",
                "redirect_url": reverse("peminjaman:laporan"),
            },
        )

    return _json_or_redirect(
        request,
        {"ok": False, "errors": ["Aksi import tidak dikenali."], "message": "Aksi import tidak dikenali."},
        status=400,
    )


@login_required
def download_format_import_riwayat_peminjaman(request):
    if not _is_super_admin_user(request.user):
        return _deny_import_access(request)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError("Library openpyxl belum tersedia. Jalankan: pip install openpyxl") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="EAF2FF")

    def add_sheet(title, headers, sample=None, validations=None):
        worksheet = workbook.create_sheet(title)
        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            worksheet.column_dimensions[cell.column_letter].width = 28
        if sample:
            for column_index, value in enumerate(sample, start=1):
                worksheet.cell(row=2, column=column_index, value=value)
        for header, values in (validations or {}).items():
            if header not in headers or not values:
                continue
            column_index = headers.index(header) + 1
            column_letter = worksheet.cell(row=1, column=column_index).column_letter
            formula = '"' + ",".join(values) + '"'
            validator = DataValidation(type="list", formula1=formula, allow_blank=True)
            worksheet.add_data_validation(validator)
            validator.add(f"{column_letter}2:{column_letter}1000")
        return worksheet

    status_values = [value for value, _label in StatusBarangChoices.choices]
    satuan_aset_values = [value for value, _label in SatuanAsetChoices.choices]
    satuan_bahan_values = [value for value, _label in SatuanBahanChoices.choices]
    kondisi_values = [value for value, _label in KondisiBarangChoices.choices]
    return_values = [label for _value, label in ReturnItemStatusChoices.choices]
    organisasi_values = [value for value, _label in InstansiKlien.OrganisasiChoices.choices]
    kategori_penunjang_values = [value for value, _label in KategoriBarangPenunjangChoices.choices]
    kategori_bahan_values = [value for value, _label in KategoriBahanOperasionalChoices.choices]
    layanan_values = [item.jenis_layanan for item in LayananKegiatan.objects.order_by("jenis_layanan")]
    tim_values = [normalize_tim_kegiatan_name(item.nama_tim) for item in TimKegiatan.objects.order_by("nama_tim")]
    instansi_values = [item.nama_instansi for item in InstansiKlien.objects.order_by("nama_instansi")]
    survei_values = [item.jenis_survei for item in SurveiKegiatan.objects.order_by("jenis_survei")]

    add_sheet(
        PEMINJAMAN_SHEET,
        PEMINJAMAN_HEADERS,
        sample=[
            "PMJ-250115-001",
            "",
            "Budi Santoso",
            "081234567890",
            "budi@example.com",
            "199001012020011001",
            "Jl. Contoh No. 1",
            "Layanan Survei Air Tanah",
            "Tim Layanan Teknis",
            "Dinas PUPR Kabupaten Contoh",
            "Eksternal PU",
            "Jl. Contoh Instansi No. 10",
            "2025-01-15",
            "2025-01-17",
            "2025-01-18",
            "Geolistrik 1D, Kualitas Air",
            "",
            2,
            0,
            1,
            1,
            0,
            0,
            0,
            0,
            0,
            0,
        ],
        validations={"Organisasi Instansi": organisasi_values},
    )
    add_sheet(
        LAB_SHEET,
        LAB_HEADERS,
        sample=[
            "PMJ-250115-001",
            "GPS Garmin 64s",
            "Garmin 64s",
            "GPS Handheld",
            "BMN",
            "BMN-2025-0001",
            "LAB-PSL-0001",
            1,
            "Unit",
            "Baik",
            2024,
            "Laboratorium",
            "Dikembalikan",
            "",
            "",
        ],
        validations={
            "Status Barang": status_values,
            "Satuan": satuan_aset_values,
            "Kondisi Barang": kondisi_values,
            "Status Pengembalian": return_values,
        },
    )
    add_sheet(
        PENUNJANG_SHEET,
        PENUNJANG_HEADERS,
        sample=[
            "PMJ-250115-001",
            "Payung Lapangan",
            "Payung Lipat",
            "Penunjang Operasional Lapangan",
            2,
            "Buah",
            "Laboratorium",
            2,
            0,
            0,
            0,
            "",
            "",
        ],
        validations={
            "Kategori Barang": kategori_penunjang_values,
            "Satuan": satuan_aset_values,
        },
    )
    add_sheet(
        BAHAN_SHEET,
        BAHAN_HEADERS,
        sample=[
            "PMJ-250115-001",
            "Botol Sampel",
            "Bahan Lapangan",
            10,
            "Botol",
            "Laboratorium",
            2,
            0,
            "",
            "Sisa 2 botol dikembalikan.",
        ],
        validations={
            "Kategori Barang": kategori_bahan_values,
            "Satuan": satuan_bahan_values,
        },
    )
    add_sheet(
        PERALATAN_LAB_SHEET,
        PERALATAN_LAB_HEADERS,
        sample=[
            "PMJ-250115-001",
            "Pompa Uji Laboratorium",
            "Grundfos",
            "Pompa",
            "Non BMN",
            "",
            "LAB-PL-0001",
            1,
            "Unit",
            "Baik",
            2023,
            "Laboratorium",
            1,
            0,
            0,
            0,
            "",
            "",
        ],
        validations={
            "Status Barang": status_values,
            "Satuan": satuan_aset_values,
            "Kondisi Barang": kondisi_values,
        },
    )

    reference_sheet = workbook.create_sheet("Referensi Pilihan")
    references = [
        ("Organisasi Instansi", organisasi_values),
        ("Layanan Kegiatan", layanan_values),
        ("Tim Kegiatan", tim_values),
        ("Instansi Tujuan", instansi_values),
        ("Kegiatan Survei", survei_values),
        ("Kategori Penunjang", kategori_penunjang_values),
        ("Kategori Bahan", kategori_bahan_values),
        ("Status Barang", status_values),
        ("Satuan Aset", satuan_aset_values),
        ("Satuan Bahan", satuan_bahan_values),
        ("Kondisi Barang", kondisi_values),
        ("Status Pengembalian", return_values),
    ]
    for col_index, (title, values) in enumerate(references, start=1):
        reference_sheet.cell(row=1, column=col_index, value=title).font = Font(bold=True)
        reference_sheet.column_dimensions[reference_sheet.cell(row=1, column=col_index).column_letter].width = 34
        for row_index, value in enumerate(values, start=2):
            reference_sheet.cell(row=row_index, column=col_index, value=value)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="format_import_riwayat_peminjaman.xlsx"'
    return response
