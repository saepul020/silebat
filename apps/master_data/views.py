from datetime import date, datetime
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required

from apps.core.navigation import get_next_url, redirect_next
from apps.core.list_pagination import paginate_list
from apps.core.excel_utils import build_excel_response
from apps.core.import_utils import (
    choice_values as _choice_values,
    import_cell as _import_cell,
    json_response_or_none as _import_json_response,
    load_import_worksheet as _load_shared_import_worksheet,
    normalize_import_header as _normalize_import_header,
    string_cell as _string_cell,
)
from django.db import transaction
from django.db.models import Count, Q
from django.db.models.deletion import ProtectedError
from django.db.models.functions import Coalesce, ExtractYear, TruncMonth
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.formats import date_format
from django.urls import reverse

from .forms import (
    BahanOperasionalForm,
    BarangLaboratoriumForm,
    BarangPenunjangOperasionalForm,
    FasilitasRuanganForm,
    PeralatanLaboratoriumForm,
)
from .models import (
    BahanOperasional,
    BarangLaboratorium,
    BarangPenunjangOperasional,
    FasilitasRuangan,
    PeralatanLaboratorium,
    KategoriBahanOperasionalChoices,
    KategoriBarangLaboratoriumChoices,
    KategoriBarangPenunjangChoices,
    KategoriSaranaPrasaranaChoices,
    KondisiBarangChoices,
    SatuanAsetChoices,
    SatuanBahanChoices,
    StatusBarangChoices,
)
from .qr_utils import ensure_master_qr_code, ensure_master_qr_codes


IMPORT_BARANG_LAB_SESSION_KEY = "barang_laboratorium_import_validated_rows"
IMPORT_BARANG_LAB_HEADERS = [
    "Status Barang",
    "Nama Barang",
    "Tipe / Merek Barang",
    "Jenis Barang",
    "Kode Aset BMN",
    "Kode Laboratorium",
    "Satuan",
    "Tahun Perolehan",
    "Kondisi Barang",
    "Lokasi Barang",
    "Kategori Barang",
    "Tanggal Pemeliharaan",
    "Tanggal Perbaikan",
    "Catatan",
]
IMPORT_BARANG_LAB_REQUIRED_HEADERS = [
    "Status Barang",
    "Nama Barang",
    "Tipe / Merek Barang",
    "Jenis Barang",
    "Kode Laboratorium",
    "Satuan",
    "Tahun Perolehan",
    "Kondisi Barang",
    "Lokasi Barang",
    "Kategori Barang",
]
IMPORT_BARANG_LAB_MAX_SIZE = 7 * 1024 * 1024

FORM_TEMPLATE = "master_data/master_form.html"
PINJAM_YEAR_ALL = "all"
ASSET_LIST_SEARCH_FIELDS = (
    "nama_barang",
    "tipe_merek_barang",
    "jenis_barang",
    "status_barang",
    "kode_aset_bmn",
    "kode_laboratorium",
    "satuan",
    "ketersediaan",
    "kondisi_barang",
    "lokasi_barang",
    "catatan",
)
PENUNJANG_LIST_SEARCH_FIELDS = (
    "nama_barang",
    "tipe_merek_barang",
    "satuan",
    "kategori_barang",
    "ketersediaan",
)
BAHAN_LIST_SEARCH_FIELDS = (
    "nama_barang",
    "kategori_barang",
    "satuan",
    "ketersediaan",
)
MONTH_LABELS_ID = {
    1: "Jan",
    2: "Feb",
    3: "Mar",
    4: "Apr",
    5: "Mei",
    6: "Jun",
    7: "Jul",
    8: "Agu",
    9: "Sep",
    10: "Okt",
    11: "Nov",
    12: "Des",
}

ASSET_FIELD_GROUPS = [
    ("Status Barang", [["status_barang"]]),
    (
        "Detail Informasi Barang",
        [
            ["nama_barang", "tipe_merek_barang", "jenis_barang"],
            ["kode_aset_bmn", "kode_laboratorium"],
            ["volume", "satuan", "ketersediaan_info"],
            ["tahun_perolehan", "kondisi_barang"],
            ["lokasi_barang"],
        ],
    ),
    ("Lampiran", [["foto_barang", "ik_alat"]]),
    (
        "Riwayat Terakhir",
        [["tanggal_pemeliharaan_info", "tanggal_perbaikan_info"]],
    ),
    ("Catatan", [["catatan"]]),
]

BARANG_LAB_FIELD_GROUPS = [
    ("Status Barang", [["status_barang"]]),
    (
        "Detail Informasi Barang",
        [
            ["nama_barang", "tipe_merek_barang", "jenis_barang"],
            ["kode_aset_bmn", "kode_laboratorium"],
            ["volume", "satuan", "ketersediaan_info"],
            ["tahun_perolehan", "kondisi_barang"],
            ["kategori_barang", "lokasi_barang"],
        ],
    ),
    ("Lampiran", [["foto_barang", "ik_alat"]]),
    (
        "Riwayat Terakhir",
        [["tanggal_pemeliharaan_info", "tanggal_perbaikan_info"]],
    ),
    ("Catatan", [["catatan"]]),
]

PENUNJANG_FIELD_GROUPS = [
    (
        "Informasi Barang",
        [
            ["nama_barang", "tipe_merek_barang"],
            ["volume", "volume_rusak", "total_volume_info"],
            ["satuan", "kategori_barang", "ketersediaan_info"],
        ],
    ),
]

BAHAN_FIELD_GROUPS = [
    (
        "Informasi Bahan",
        [
            ["nama_barang", "kategori_barang"],
            ["volume", "satuan"],
            ["stok_minimum", "ketersediaan_info"],
        ],
    ),
]

SARANA_FIELD_GROUPS = [
    ("Status Barang", [["status_barang"]]),
    (
        "Detail Informasi Barang",
        [
            ["nama_barang", "tipe_merek_barang", "jenis_barang"],
            ["kode_aset_bmn", "kode_laboratorium"],
            ["volume", "volume_rusak", "total_volume_info"],
            ["satuan", "ketersediaan_info"],
            ["tahun_perolehan", "kategori_barang"],
            ["kondisi_barang", "lokasi_barang"],
        ],
    ),
    ("Lampiran", [["foto_barang", "ik_alat"]]),
    (
        "Riwayat Terakhir",
        [["tanggal_pemeliharaan_info", "tanggal_perbaikan_info"]],
    ),
    ("Catatan", [["catatan"]]),
]

PERALATAN_LAB_FIELD_GROUPS = [
    ("Status Barang", [["status_barang"]]),
    (
        "Detail Informasi Barang",
        [
            ["nama_barang", "tipe_merek_barang", "jenis_barang"],
            ["kode_aset_bmn", "kode_laboratorium"],
            ["volume", "volume_rusak", "total_volume_info"],
            ["satuan", "ketersediaan_info"],
            ["tahun_perolehan", "kondisi_barang"],
            ["lokasi_barang"],
        ],
    ),
    ("Lampiran", [["foto_barang", "ik_alat"]]),
    (
        "Riwayat Terakhir",
        [["tanggal_pemeliharaan_info", "tanggal_perbaikan_info"]],
    ),
    ("Catatan", [["catatan"]]),
]


def _display_value(value, default="-"):
    if value in (None, "", []):
        return default
    return value


def _display_date(value):
    if not value:
        return "-"
    return date_format(value, "d F Y")


def _safe_file_url(file_field):
    if not file_field:
        return None
    try:
        return file_field.url
    except (ValueError, AttributeError):
        return None


def _file_export_url(request, file_field):
    file_url = _safe_file_url(file_field)
    return request.build_absolute_uri(file_url) if file_url else "-"


def _ik_alat_detail_item(obj):
    file_url = _safe_file_url(getattr(obj, "ik_alat", None))
    return {
        "label": "IK Alat",
        "value": "Lihat Dokumen PDF" if file_url else "-",
        "url": file_url,
    }


def _boolean_display(value):
    return "Ya" if value else "Tidak"


def _public_detail_export_url(request, obj):
    path = obj.get_public_detail_url_path() if hasattr(obj, "get_public_detail_url_path") else ""
    if not path or path == "#":
        return "-"
    return request.build_absolute_uri(path)


def _export_master_queryset(request, *, filename, sheet_title, headers, queryset, row_builder):
    return build_excel_response(
        filename,
        [
            {
                "title": sheet_title,
                "headers": headers,
                "rows": [row_builder(obj) for obj in queryset],
            }
        ],
    )


def _render_form_page(
    request,
    *,
    form,
    field_groups,
    page_title,
    page_subtitle,
    submit_label,
    cancel_url,
    cancel_url_kwargs=None,
):
    next_url = get_next_url(request)
    return render(
        request,
        FORM_TEMPLATE,
        {
            "form": form,
            "field_groups": field_groups,
            "page_title": page_title,
            "page_subtitle": page_subtitle,
            "submit_label": submit_label,
            "cancel_url": cancel_url,
            "cancel_url_kwargs": cancel_url_kwargs or {},
            "next_url": next_url,
        },
    )


def _resolve_redirect_target(request, redirect_to, saved_obj):
    if isinstance(redirect_to, tuple):
        url_name, kwargs_factory = redirect_to
        kwargs = kwargs_factory(saved_obj) if callable(kwargs_factory) else {}
        return redirect_next(request, url_name, **kwargs)
    return redirect_next(request, redirect_to)


def _handle_form_page(
    request,
    *,
    form_class,
    field_groups,
    success_message,
    redirect_to,
    page_title,
    page_subtitle,
    submit_label,
    cancel_url,
    cancel_url_kwargs=None,
    instance=None,
):
    form_kwargs = {}
    if request.method == "POST":
        form_kwargs["data"] = request.POST
        form_kwargs["files"] = request.FILES
    if instance is not None:
        form_kwargs["instance"] = instance

    form = form_class(**form_kwargs)
    if request.method == "POST" and form.is_valid():
        saved_obj = form.save()
        messages.success(request, success_message)
        return _resolve_redirect_target(request, redirect_to, saved_obj)

    return _render_form_page(
        request,
        form=form,
        field_groups=field_groups,
        page_title=page_title,
        page_subtitle=page_subtitle,
        submit_label=submit_label,
        cancel_url=cancel_url,
        cancel_url_kwargs=cancel_url_kwargs,
    )


def _get_active_peminjaman_delete_block_message(obj):
    """
    Menolak hapus data master hanya jika item masih dipakai pada proses
    peminjaman/pengembalian aktif. Item yang sudah masuk Laporan Peminjaman
    tidak menjadi pengunci karena data laporannya disimpan sebagai snapshot.
    """
    try:
        from apps.peminjaman.models import (
            PeminjamanBahanOperasional,
            PeminjamanBarangLaboratorium,
            PeminjamanBarangPenunjang,
            PengembalianBahanOperasional,
            PengembalianBarangLaboratorium,
            PengembalianBarangPenunjang,
            PeminjamanRequest,
            ReturnStepChoices,
            StepChoices,
        )
    except Exception:
        return ""

    active_numbers = set()

    def collect_numbers(queryset):
        queryset = (
            queryset.select_related("pengajuan")
            .exclude(pengajuan__return_current_step=ReturnStepChoices.COMPLETED)
            .exclude(pengajuan__current_step=StepChoices.REJECTED)
        )
        for row in queryset[:10]:
            nomor = getattr(row.pengajuan, "nomor_pengajuan", "") or f"Pengajuan #{row.pengajuan_id}"
            active_numbers.add(nomor)

    if isinstance(obj, BarangLaboratorium):
        collect_numbers(PeminjamanBarangLaboratorium.objects.filter(barang=obj))
        collect_numbers(PengembalianBarangLaboratorium.objects.filter(barang=obj))
        if getattr(obj, "sedang_dipinjam", False):
            active_numbers.add("status barang sedang dipinjam")
    elif isinstance(obj, BarangPenunjangOperasional):
        collect_numbers(PeminjamanBarangPenunjang.objects.filter(barang=obj))
        collect_numbers(PengembalianBarangPenunjang.objects.filter(barang=obj))
        if (getattr(obj, "volume_dipinjam", 0) or 0) > 0:
            active_numbers.add("stok masih tercatat sedang dipinjam")
    elif isinstance(obj, BahanOperasional):
        collect_numbers(PeminjamanBahanOperasional.objects.filter(bahan=obj))
        collect_numbers(PengembalianBahanOperasional.objects.filter(bahan=obj))

    if not active_numbers:
        return ""

    shown = sorted(active_numbers)
    shown_text = ", ".join(shown[:5])
    if len(shown) > 5:
        shown_text += f", dan {len(shown) - 5} data lainnya"

    return (
        f'Data "{obj}" tidak dapat dihapus karena masih digunakan pada proses '
        f'peminjaman/pengembalian aktif ({shown_text}). '
        "Selesaikan proses pengembalian atau batalkan/hapus pengajuan aktif terlebih dahulu."
    )


def _handle_delete(request, *, model, pk, success_message, redirect_to):
    obj = get_object_or_404(model, pk=pk)
    if request.method == "POST":
        block_message = _get_active_peminjaman_delete_block_message(obj)
        if block_message:
            messages.error(request, block_message)
            return redirect(redirect_to)

        try:
            obj.delete()
        except ProtectedError:
            messages.error(
                request,
                (
                    f'Data "{obj}" belum dapat dihapus karena masih mempunyai relasi '
                    "aktif di modul lain. Pastikan proses peminjaman/pengembalian yang aktif sudah selesai, "
                    "lalu jalankan migrasi terbaru agar data laporan memakai snapshot dan tidak lagi mengunci data master."
                ),
            )
            return redirect(redirect_to)

        messages.success(request, success_message)
    return redirect(redirect_to)


def _ensure_qr_codes_for_context_items(context):
    items = context.get("items") or []
    ensure_master_qr_codes(items)


def _get_pinjam_year(request, available_years):
    current_year = timezone.localdate().year
    raw_year = (request.GET.get("tahun_pinjam") or "").strip().lower()
    valid_years = {int(year) for year in available_years if year}

    if raw_year == PINJAM_YEAR_ALL:
        return PINJAM_YEAR_ALL
    try:
        requested_year = int(raw_year)
    except (TypeError, ValueError):
        return current_year
    return requested_year if requested_year in valid_years else current_year


def _clean_match_value(value):
    text = str(value or "").strip()
    return text if text and text != "-" else ""


def _snapshot_asset_q(obj):
    """
    Data laporan lama/import riwayat disimpan sebagai snapshot dengan FK barang
    bernilai NULL. Chart harus tetap menghitung baris tersebut dengan cara
    mencocokkan kode aset/laboratorium, lalu fallback ke identitas nama+tipe.
    """
    query = Q(barang=obj)
    null_fk = Q(barang__isnull=True)

    kode_lab = _clean_match_value(getattr(obj, "kode_laboratorium", ""))
    if kode_lab:
        query |= null_fk & Q(snapshot_kode_laboratorium__iexact=kode_lab)

    kode_bmn = _clean_match_value(getattr(obj, "kode_aset_bmn", ""))
    if kode_bmn:
        query |= null_fk & Q(snapshot_kode_aset_bmn__iexact=kode_bmn)

    nama = _clean_match_value(getattr(obj, "nama_barang", ""))
    tipe = _clean_match_value(getattr(obj, "tipe_merek_barang", ""))
    jenis = _clean_match_value(getattr(obj, "jenis_barang", ""))
    if nama and tipe:
        fallback = null_fk & Q(snapshot_nama_barang__iexact=nama) & Q(snapshot_tipe_merek_barang__iexact=tipe)
        if jenis:
            fallback &= Q(snapshot_jenis_barang__iexact=jenis)
        query |= fallback

    return query


def _build_pinjam_chart(request, obj, *, item_model, label):
    try:
        from apps.peminjaman.models import StepChoices
    except Exception:
        return None

    current_year = timezone.localdate().year
    date_expr = Coalesce(
        "pengajuan__pimpinan_at",
        "pengajuan__return_completed_at",
        "pengajuan__submitted_at",
        "pengajuan__updated_at",
    )
    base_qs = item_model.objects.filter(
        _snapshot_asset_q(obj),
        pengajuan__current_step=StepChoices.APPROVED,
    )
    monthly_rows = list(
        base_qs.annotate(pinjam_at=date_expr)
        .annotate(month=TruncMonth("pinjam_at"))
        .values("month")
        .annotate(total=Count("id"))
        .order_by("month")
    )
    year_rows = list(
        base_qs.annotate(pinjam_at=date_expr)
        .annotate(year=ExtractYear("pinjam_at"))
        .values("year")
        .annotate(total=Count("id"))
        .order_by("year")
    )
    available_years = sorted(
        {
            int(row["year"])
            for row in year_rows
            if row.get("year")
        }
        | {current_year}
    )
    selected_year = _get_pinjam_year(request, available_years)
    return {
        "title": f"Riwayat Peminjaman {label}",
        "datasetLabel": "Jumlah Peminjaman",
        "chartTitle": "Rekap Peminjaman Alat",
        "scrollMobile": True,
        "availableYears": available_years,
        "selectedYear": selected_year,
        "currentYear": current_year,
        "allValue": PINJAM_YEAR_ALL,
        "monthLabels": MONTH_LABELS_ID,
        "rows": [
            {
                "year": row["month"].year,
                "month": row["month"].month,
                "total": int(row.get("total") or 0),
            }
            for row in monthly_rows
            if row.get("month")
        ],
        "yearRows": [
            {"year": int(row["year"]), "total": int(row.get("total") or 0)}
            for row in year_rows
            if row.get("year")
        ],
    }


def _build_detail_context(
    *,
    obj,
    page_title,
    page_subtitle,
    edit_url_name,
    edit_label,
    back_url_name,
    top_meta,
    chips,
    detail_items,
    loan_chart=None,
    top_icon="bi-box-seam",
    photo_attr="foto_barang",
):
    ensure_master_qr_code(obj)
    return {
        "obj": obj,
        "page_title": page_title,
        "page_subtitle": page_subtitle,
        "edit_url_name": edit_url_name,
        "edit_label": edit_label,
        "back_url_name": back_url_name,
        "top_meta": _display_value(top_meta),
        "chips": [chip for chip in chips if chip],
        "detail_items": detail_items,
        "loan_chart": loan_chart,
        "photo_url": _safe_file_url(getattr(obj, photo_attr, None))
        if photo_attr
        else None,
        "top_icon": top_icon,
        "qr_code_url": _safe_file_url(getattr(obj, "qr_code", None)),
        "public_detail_url": obj.get_public_detail_url()
        if hasattr(obj, "get_public_detail_url")
        else None,
    }


def _render_public_master_detail(
    request,
    *,
    obj,
    page_title,
    page_subtitle,
    top_meta,
    chips,
    detail_items,
    loan_chart=None,
    top_icon="bi-box-seam",
    photo_attr="foto_barang",
):
    return render(
        request,
        "master_data/public_detail_barang.html",
        {
            "obj": obj,
            "page_title": page_title,
            "page_subtitle": page_subtitle,
            "top_meta": _display_value(top_meta),
            "chips": [chip for chip in chips if chip],
            "detail_items": detail_items,
            "loan_chart": loan_chart,
            "photo_url": _safe_file_url(getattr(obj, photo_attr, None))
            if photo_attr
            else None,
            "top_icon": top_icon,
        },
    )


@login_required
def index(request):
    return redirect("master_data:data_barang_laboratorium")


def _render_barang_laboratorium_list(request, import_context=None):
    queryset = BarangLaboratorium.objects.order_by("nama_barang")
    context = paginate_list(
        request,
        queryset,
        search_fields=(*ASSET_LIST_SEARCH_FIELDS, "kategori_barang"),
    )
    _ensure_qr_codes_for_context_items(context)
    context.update(
        {
            "page_title": "Data Peralatan Survei Lapangan",
            "page_subtitle": "Kelola peralatan survei lapangan.",
            "import_context": import_context or {},
        }
    )
    return render(request, "master_data/data_barang_laboratorium_list.html", context)


@login_required
def data_barang_laboratorium(request):
    return _render_barang_laboratorium_list(request)


@login_required
def export_barang_laboratorium(request):
    headers = [
        "Nama Barang",
        "Status Barang",
        "Tipe / Merek Barang",
        "Jenis Barang",
        "Kode Aset BMN",
        "Kode Laboratorium",
        "Volume",
        "Satuan",
        "Tahun Perolehan",
        "Kondisi Barang",
        "Kategori Barang",
        "Ketersediaan",
        "Sedang Dipinjam",
        "Lokasi Barang",
        "Tanggal Pemeliharaan",
        "Tanggal Perbaikan",
        "Catatan",
        "IK Alat (PDF)",
        "URL QR Detail",
    ]
    queryset = BarangLaboratorium.objects.order_by("nama_barang")
    return _export_master_queryset(
        request,
        filename="export_data_peralatan_survei_lapangan.xlsx",
        sheet_title="Peralatan Survei",
        headers=headers,
        queryset=queryset,
        row_builder=lambda obj: [
            obj.nama_barang,
            obj.status_barang,
            obj.tipe_merek_barang,
            obj.jenis_barang,
            obj.kode_aset_bmn,
            obj.kode_laboratorium,
            obj.volume,
            obj.satuan,
            obj.tahun_perolehan,
            obj.kondisi_barang,
            obj.kategori_barang,
            obj.ketersediaan,
            _boolean_display(obj.sedang_dipinjam),
            obj.lokasi_barang,
            obj.tanggal_pemeliharaan,
            obj.tanggal_perbaikan,
            obj.catatan,
            _file_export_url(request, obj.ik_alat),
            _public_detail_export_url(request, obj),
        ],
    )


def _year_cell(value):
    if value in (None, ''):
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    raw_value = str(value).strip()
    if raw_value.endswith('.0'):
        raw_value = raw_value[:-2]
    if not raw_value.isdigit():
        raise ValueError('Tahun Perolehan harus berupa angka tahun yang valid.')
    year = int(raw_value)
    if year < 1900 or year > 2100:
        raise ValueError('Tahun Perolehan harus berada pada rentang 1900 sampai 2100.')
    return year


def _date_import_cell(value, label):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw_value = str(value).strip()
    if not raw_value:
        return None
    if raw_value.endswith('.0'):
        raw_value = raw_value[:-2]
    for date_format_pattern in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(raw_value, date_format_pattern).date()
        except ValueError:
            continue
    raise ValueError(f'{label} harus berupa tanggal valid dengan format YYYY-MM-DD.')


IMPORT_BARANG_PENUNJANG_SESSION_KEY = "barang_penunjang_import_validated_rows"
IMPORT_BAHAN_OPERASIONAL_SESSION_KEY = "bahan_operasional_import_validated_rows"
IMPORT_FASILITAS_RUANGAN_SESSION_KEY = "fasilitas_ruangan_import_validated_rows"
IMPORT_PERALATAN_LABORATORIUM_SESSION_KEY = "peralatan_laboratorium_import_validated_rows"
IMPORT_EXCEL_MAX_SIZE = IMPORT_BARANG_LAB_MAX_SIZE

IMPORT_BARANG_PENUNJANG_HEADERS = [
    "Nama Barang",
    "Tipe / Merek Barang",
    "Volume Baik",
    "Volume Rusak",
    "Satuan",
    "Kategori Barang",
]
IMPORT_BARANG_PENUNJANG_REQUIRED_HEADERS = IMPORT_BARANG_PENUNJANG_HEADERS[:]

IMPORT_BAHAN_OPERASIONAL_HEADERS = [
    "Nama Barang",
    "Kategori Barang",
    "Volume",
    "Satuan",
    "Stok Minimum",
]
IMPORT_BAHAN_OPERASIONAL_REQUIRED_HEADERS = IMPORT_BAHAN_OPERASIONAL_HEADERS[:]

IMPORT_FASILITAS_RUANGAN_HEADERS = [
    "Status Barang",
    "Nama Barang",
    "Tipe / Merek Barang",
    "Jenis Barang",
    "Kode Aset BMN",
    "Kode Laboratorium",
    "Volume Baik",
    "Volume Rusak",
    "Satuan",
    "Tahun Perolehan",
    "Kategori Barang",
    "Kondisi Barang",
    "Lokasi Barang",
    "Tanggal Pemeliharaan",
    "Tanggal Perbaikan",
    "Catatan",
]
IMPORT_FASILITAS_RUANGAN_REQUIRED_HEADERS = [
    "Status Barang",
    "Nama Barang",
    "Tipe / Merek Barang",
    "Jenis Barang",
    "Kode Laboratorium",
    "Satuan",
    "Tahun Perolehan",
    "Kategori Barang",
    "Lokasi Barang",
]

IMPORT_PERALATAN_LABORATORIUM_HEADERS = [
    "Status Barang",
    "Nama Barang",
    "Tipe / Merek Barang",
    "Jenis Barang",
    "Kode Aset BMN",
    "Kode Laboratorium",
    "Volume Baik",
    "Volume Rusak",
    "Satuan",
    "Tahun Perolehan",
    "Kondisi Barang",
    "Lokasi Barang",
    "Tanggal Pemeliharaan",
    "Tanggal Perbaikan",
    "Catatan",
]
IMPORT_PERALATAN_LABORATORIUM_REQUIRED_HEADERS = [
    "Status Barang",
    "Nama Barang",
    "Tipe / Merek Barang",
    "Jenis Barang",
    "Kode Laboratorium",
    "Satuan",
    "Tahun Perolehan",
    "Lokasi Barang",
]


def _load_import_worksheet(file_obj, headers, required_headers):
    return _load_shared_import_worksheet(
        file_obj,
        headers,
        required_headers,
        max_size_bytes=IMPORT_EXCEL_MAX_SIZE,
    )


def _positive_int_import_cell(value, label, min_value=0):
    if value in (None, ''):
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    raw_value = str(value).strip()
    if raw_value.endswith('.0'):
        raw_value = raw_value[:-2]
    if not raw_value.isdigit():
        raise ValueError(f'{label} harus berupa angka bulat yang valid.')
    number = int(raw_value)
    if number < min_value:
        raise ValueError(f'{label} minimal {min_value}.')
    return number


def _validate_unique_import_value(*, row_errors, seen, value, excel_row_number, duplicate_message):
    key = (value or '').strip().lower()
    if not key:
        return
    if key in seen:
        row_errors.append(f'{duplicate_message} duplikat dengan baris {seen[key]}.')
    else:
        seen[key] = excel_row_number


def _validate_barang_penunjang_import(file_obj):
    worksheet, normalized_headers, header_aliases, initial_errors = _load_import_worksheet(
        file_obj,
        IMPORT_BARANG_PENUNJANG_HEADERS,
        IMPORT_BARANG_PENUNJANG_REQUIRED_HEADERS,
    )
    if initial_errors:
        return [], initial_errors

    satuan_choices = _choice_values(SatuanAsetChoices.choices)
    kategori_choices = _choice_values(KategoriBarangPenunjangChoices.choices)
    rows = []
    errors = []
    seen_nama = {}

    def cell(row, header):
        return _import_cell(row, header, normalized_headers, header_aliases)

    for excel_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_string_cell(value) for value in row):
            continue

        row_errors = []
        data = {
            'nama_barang': cell(row, 'Nama Barang'),
            'tipe_merek_barang': cell(row, 'Tipe / Merek Barang'),
            'satuan': cell(row, 'Satuan'),
            'kategori_barang': cell(row, 'Kategori Barang'),
        }

        for header in IMPORT_BARANG_PENUNJANG_REQUIRED_HEADERS:
            if not cell(row, header):
                row_errors.append(f'{header} wajib diisi.')

        try:
            data['volume'] = _positive_int_import_cell(cell(row, 'Volume Baik'), 'Volume Baik', 0)
        except ValueError as exc:
            row_errors.append(str(exc))
            data['volume'] = None

        try:
            data['volume_rusak'] = _positive_int_import_cell(cell(row, 'Volume Rusak'), 'Volume Rusak', 0)
        except ValueError as exc:
            row_errors.append(str(exc))
            data['volume_rusak'] = None

        if data['satuan'] and data['satuan'] not in satuan_choices:
            row_errors.append('Satuan tidak sesuai pilihan yang tersedia.')

        if data['kategori_barang'] and data['kategori_barang'] not in kategori_choices:
            row_errors.append('Kategori Barang tidak sesuai pilihan yang tersedia.')

        _validate_unique_import_value(
            row_errors=row_errors,
            seen=seen_nama,
            value=data['nama_barang'],
            excel_row_number=excel_row_number,
            duplicate_message='Nama Barang',
        )

        if data['nama_barang'] and BarangPenunjangOperasional.objects.filter(nama_barang__iexact=data['nama_barang']).exists():
            row_errors.append('Nama Barang sudah terdaftar di database.')

        if row_errors:
            errors.append(f'Baris {excel_row_number}: ' + ' '.join(row_errors))
            continue

        rows.append(data)

    if not rows and not errors:
        errors.append('File Excel tidak memiliki data barang penunjang lapangan untuk diimport.')

    return rows, errors


def _validate_bahan_operasional_import(file_obj):
    worksheet, normalized_headers, header_aliases, initial_errors = _load_import_worksheet(
        file_obj,
        IMPORT_BAHAN_OPERASIONAL_HEADERS,
        IMPORT_BAHAN_OPERASIONAL_REQUIRED_HEADERS,
    )
    if initial_errors:
        return [], initial_errors

    satuan_choices = _choice_values(SatuanBahanChoices.choices)
    kategori_choices = _choice_values(KategoriBahanOperasionalChoices.choices)
    rows = []
    errors = []
    seen_nama = {}

    def cell(row, header):
        return _import_cell(row, header, normalized_headers, header_aliases)

    for excel_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_string_cell(value) for value in row):
            continue

        row_errors = []
        data = {
            'nama_barang': cell(row, 'Nama Barang'),
            'kategori_barang': cell(row, 'Kategori Barang'),
            'satuan': cell(row, 'Satuan'),
        }

        for header in IMPORT_BAHAN_OPERASIONAL_REQUIRED_HEADERS:
            if not cell(row, header):
                row_errors.append(f'{header} wajib diisi.')

        try:
            data['volume'] = _positive_int_import_cell(cell(row, 'Volume'), 'Volume', 0)
        except ValueError as exc:
            row_errors.append(str(exc))
            data['volume'] = None

        try:
            data['stok_minimum'] = _positive_int_import_cell(cell(row, 'Stok Minimum'), 'Stok Minimum', 1)
        except ValueError as exc:
            row_errors.append(str(exc))
            data['stok_minimum'] = None

        if data['kategori_barang'] and data['kategori_barang'] not in kategori_choices:
            row_errors.append('Kategori Barang tidak sesuai pilihan yang tersedia.')

        if data['satuan'] and data['satuan'] not in satuan_choices:
            row_errors.append('Satuan tidak sesuai pilihan yang tersedia.')

        _validate_unique_import_value(
            row_errors=row_errors,
            seen=seen_nama,
            value=data['nama_barang'],
            excel_row_number=excel_row_number,
            duplicate_message='Nama Barang',
        )

        if data['nama_barang'] and BahanOperasional.objects.filter(nama_barang__iexact=data['nama_barang']).exists():
            row_errors.append('Nama Barang sudah terdaftar di database.')

        if row_errors:
            errors.append(f'Baris {excel_row_number}: ' + ' '.join(row_errors))
            continue

        rows.append(data)

    if not rows and not errors:
        errors.append('File Excel tidak memiliki data bahan operasional untuk diimport.')

    return rows, errors



def _validate_asset_status_import(
    file_obj,
    *,
    headers,
    required_headers,
    model,
    empty_message,
    include_kategori=False,
    kategori_choices=None,
    bmn_required_message='Kode Aset BMN wajib diisi untuk barang berstatus BMN.',
):
    worksheet, normalized_headers, header_aliases, initial_errors = _load_import_worksheet(
        file_obj,
        headers,
        required_headers,
    )
    if initial_errors:
        return [], initial_errors

    status_choices = _choice_values(StatusBarangChoices.choices)
    satuan_choices = _choice_values(SatuanAsetChoices.choices)
    kondisi_choices = _choice_values(KondisiBarangChoices.choices)
    kategori_choices = set(kategori_choices or [])

    rows = []
    errors = []
    seen_kode_aset_bmn = {}
    seen_kode_laboratorium = {}

    def cell(row, header):
        return _import_cell(row, header, normalized_headers, header_aliases)

    for excel_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_string_cell(value) for value in row):
            continue

        row_errors = []
        status_barang = cell(row, 'Status Barang')
        data = {
            'status_barang': status_barang,
            'nama_barang': cell(row, 'Nama Barang'),
            'tipe_merek_barang': cell(row, 'Tipe / Merek Barang'),
            'jenis_barang': cell(row, 'Jenis Barang'),
            'kode_aset_bmn': cell(row, 'Kode Aset BMN'),
            'kode_laboratorium': cell(row, 'Kode Laboratorium'),
            'satuan': cell(row, 'Satuan'),
            'kondisi_barang': cell(row, 'Kondisi Barang'),
            'lokasi_barang': cell(row, 'Lokasi Barang'),
            'catatan': cell(row, 'Catatan'),
        }
        if include_kategori:
            data['kategori_barang'] = cell(row, 'Kategori Barang')

        for header in required_headers:
            if not cell(row, header):
                row_errors.append(f'{header} wajib diisi.')

        if data['status_barang'] and data['status_barang'] not in status_choices:
            row_errors.append('Status Barang tidak sesuai pilihan yang tersedia.')

        if data['satuan'] and data['satuan'] not in satuan_choices:
            row_errors.append('Satuan tidak sesuai pilihan yang tersedia.')

        if include_kategori and data.get('kategori_barang') and data['kategori_barang'] not in kategori_choices:
            row_errors.append('Kategori Barang tidak sesuai pilihan yang tersedia.')

        try:
            volume_raw = cell(row, 'Volume Baik')
            data['volume'] = _positive_int_import_cell(volume_raw, 'Volume Baik', 0) if volume_raw else None
        except ValueError as exc:
            row_errors.append(str(exc))
            data['volume'] = None

        try:
            volume_rusak_raw = cell(row, 'Volume Rusak')
            data['volume_rusak'] = _positive_int_import_cell(volume_rusak_raw, 'Volume Rusak', 0) if volume_rusak_raw else None
        except ValueError as exc:
            row_errors.append(str(exc))
            data['volume_rusak'] = None

        try:
            data['tahun_perolehan'] = _year_cell(cell(row, 'Tahun Perolehan'))
        except ValueError as exc:
            row_errors.append(str(exc))
            data['tahun_perolehan'] = None

        for field_name, header_label in (
            ('tanggal_pemeliharaan', 'Tanggal Pemeliharaan'),
            ('tanggal_perbaikan', 'Tanggal Perbaikan'),
        ):
            try:
                data[field_name] = _date_import_cell(cell(row, header_label), header_label)
            except ValueError as exc:
                row_errors.append(str(exc))
                data[field_name] = None

        if data['status_barang'] == StatusBarangChoices.BMN:
            if not data['kode_aset_bmn']:
                row_errors.append(bmn_required_message)
            if not data['kondisi_barang']:
                row_errors.append('Kondisi Barang wajib diisi untuk barang berstatus BMN.')
            elif data['kondisi_barang'] not in kondisi_choices:
                row_errors.append('Kondisi Barang tidak sesuai pilihan yang tersedia.')
            elif data['kondisi_barang'] == KondisiBarangChoices.BAIK:
                data['volume'] = 1
                data['volume_rusak'] = 0
            elif data['kondisi_barang'] == KondisiBarangChoices.HILANG:
                data['volume'] = 0
                data['volume_rusak'] = 0
            else:
                data['volume'] = 0
                data['volume_rusak'] = 1
        elif data['status_barang'] == StatusBarangChoices.NON_BMN:
            if data['volume'] is None:
                row_errors.append('Volume Baik wajib diisi untuk barang berstatus Non BMN.')
            if data['volume_rusak'] is None:
                row_errors.append('Volume Rusak wajib diisi untuk barang berstatus Non BMN.')
            data['kode_aset_bmn'] = None
            data['kondisi_barang'] = KondisiBarangChoices.BAIK
        elif data['kondisi_barang'] and data['kondisi_barang'] not in kondisi_choices:
            row_errors.append('Kondisi Barang tidak sesuai pilihan yang tersedia.')

        if data['volume'] is None:
            data['volume'] = 0
        if data['volume_rusak'] is None:
            data['volume_rusak'] = 0

        enforce_asset_code_unique = model in (BarangLaboratorium, PeralatanLaboratorium)
        if enforce_asset_code_unique:
            _validate_unique_import_value(
                row_errors=row_errors,
                seen=seen_kode_laboratorium,
                value=data['kode_laboratorium'],
                excel_row_number=excel_row_number,
                duplicate_message='Kode Laboratorium',
            )
            if data.get('kode_aset_bmn'):
                _validate_unique_import_value(
                    row_errors=row_errors,
                    seen=seen_kode_aset_bmn,
                    value=data['kode_aset_bmn'],
                    excel_row_number=excel_row_number,
                    duplicate_message='Kode Aset BMN',
                )

            if data['kode_laboratorium'] and model.objects.filter(kode_laboratorium__iexact=data['kode_laboratorium']).exists():
                row_errors.append('Kode Laboratorium sudah terdaftar di database.')

            if data.get('kode_aset_bmn') and model.objects.filter(kode_aset_bmn__iexact=data['kode_aset_bmn']).exists():
                row_errors.append('Kode Aset BMN sudah terdaftar di database.')

        if row_errors:
            errors.append(f'Baris {excel_row_number}: ' + ' '.join(row_errors))
            continue

        rows.append(data)

    if not rows and not errors:
        errors.append(empty_message)

    return rows, errors


def _validate_fasilitas_ruangan_import(file_obj):
    return _validate_asset_status_import(
        file_obj,
        headers=IMPORT_FASILITAS_RUANGAN_HEADERS,
        required_headers=IMPORT_FASILITAS_RUANGAN_REQUIRED_HEADERS,
        model=FasilitasRuangan,
        empty_message='File Excel tidak memiliki data sarana prasarana ruangan untuk diimport.',
        include_kategori=True,
        kategori_choices=_choice_values(KategoriSaranaPrasaranaChoices.choices),
        bmn_required_message='Kode Aset BMN wajib diisi untuk sarana prasarana ruangan berstatus BMN.',
    )


def _validate_peralatan_laboratorium_import(file_obj):
    return _validate_asset_status_import(
        file_obj,
        headers=IMPORT_PERALATAN_LABORATORIUM_HEADERS,
        required_headers=IMPORT_PERALATAN_LABORATORIUM_REQUIRED_HEADERS,
        model=PeralatanLaboratorium,
        empty_message='File Excel tidak memiliki data peralatan laboratorium untuk diimport.',
        include_kategori=False,
        bmn_required_message='Kode Aset BMN wajib diisi untuk peralatan laboratorium berstatus BMN.',
    )


def _save_import_objects(rows, model, duplicate_checks=None):
    duplicate_errors = []
    duplicate_checks = duplicate_checks or []
    for index, row in enumerate(rows, start=1):
        for field_name, label in duplicate_checks:
            value = row.get(field_name)
            if value and model.objects.filter(**{f'{field_name}__iexact': value}).exists():
                duplicate_errors.append(f'Data valid nomor {index}: {label} sudah terdaftar di database.')

    if duplicate_errors:
        return 0, duplicate_errors

    with transaction.atomic():
        objects = [model(**row) for row in rows]
        for obj in objects:
            obj.full_clean()
            obj.save()

    return len(rows), []


def _handle_import_post(
    request,
    *,
    session_key,
    redirect_url_name,
    render_list_func,
    validate_func,
    save_func,
    item_label,
):
    if request.method != 'POST':
        return redirect(redirect_url_name)

    action = request.POST.get('import_action')

    if action == 'cancel':
        request.session.pop(session_key, None)
        response = _import_json_response(request, {'ok': True, 'cancelled': True})
        if response:
            return response
        return redirect(redirect_url_name)

    if action == 'validate':
        rows, errors = validate_func(request.FILES.get('file_import'))
        import_context = {
            'show_modal': True,
            'validated': not errors,
            'can_save': bool(rows) and not errors,
            'total_rows': len(rows),
            'errors': errors,
        }
        if errors:
            request.session.pop(session_key, None)
        else:
            request.session[session_key] = rows
            request.session.modified = True

        response = _import_json_response(request, {
            'ok': not bool(errors),
            'validated': not bool(errors),
            'can_save': bool(rows) and not errors,
            'total_rows': len(rows),
            'errors': errors,
            'message': f'Validasi berhasil. {len(rows)} data siap disimpan.' if not errors else 'Validasi belum berhasil.',
        })
        if response:
            return response
        return render_list_func(request, import_context=import_context)

    if action == 'save':
        rows = request.session.get(session_key) or []
        if not rows:
            payload = {
                'ok': False,
                'saved': False,
                'can_save': False,
                'errors': ['Data validasi tidak ditemukan atau sudah kedaluwarsa. Lakukan Validasi Data terlebih dahulu.'],
                'message': 'Data belum dapat disimpan.',
            }
            response = _import_json_response(request, payload)
            if response:
                return response
            return render_list_func(request, import_context={
                'show_modal': True,
                'validated': False,
                'can_save': False,
                'errors': payload['errors'],
            })

        total_saved, save_errors = save_func(rows)
        if save_errors:
            request.session.pop(session_key, None)
            payload = {
                'ok': False,
                'saved': False,
                'can_save': False,
                'errors': save_errors,
                'message': 'Data tidak disimpan karena ditemukan duplikasi terbaru di database.',
            }
            response = _import_json_response(request, payload)
            if response:
                return response
            return render_list_func(request, import_context={
                'show_modal': True,
                'validated': False,
                'can_save': False,
                'errors': save_errors,
            })

        request.session.pop(session_key, None)
        response = _import_json_response(request, {
            'ok': True,
            'saved': True,
            'can_save': False,
            'total_rows': total_saved,
            'errors': [],
            'message': f'{total_saved} data {item_label} berhasil diimport.',
            'redirect_url': reverse(redirect_url_name),
        })
        if response:
            return response
        return redirect(redirect_url_name)

    payload = {'ok': False, 'errors': ['Aksi import tidak dikenali.'], 'message': 'Aksi import tidak dikenali.'}
    response = _import_json_response(request, payload, status=400)
    if response:
        return response
    return redirect(redirect_url_name)


def _download_import_template(*, headers, sample, references, validations, worksheet_title, filename):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError('Library openpyxl belum tersedia. Jalankan: pip install openpyxl') from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_title

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='EAF2FF')
        worksheet.column_dimensions[cell.column_letter].width = 26

    for column_index, value in enumerate(sample, start=1):
        worksheet.cell(row=2, column=column_index, value=value)

    if references:
        lists_sheet = workbook.create_sheet('Referensi Pilihan')
        for col_index, (title, values) in enumerate(references, start=1):
            lists_sheet.cell(row=1, column=col_index, value=title).font = Font(bold=True)
            lists_sheet.column_dimensions[lists_sheet.cell(row=1, column=col_index).column_letter].width = 34
            for row_index, value in enumerate(values, start=2):
                lists_sheet.cell(row=row_index, column=col_index, value=value)

    for header, values in (validations or {}).items():
        if header not in headers:
            continue
        column_index = headers.index(header) + 1
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        formula = '"' + ','.join(values) + '"'
        validator = DataValidation(type='list', formula1=formula, allow_blank=False)
        worksheet.add_data_validation(validator)
        validator.add(f'{column_letter}2:{column_letter}1000')

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _validate_barang_laboratorium_import(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('Library openpyxl belum tersedia. Jalankan: pip install openpyxl') from exc

    if not file_obj:
        return [], ['File Excel wajib diupload.']

    filename = getattr(file_obj, 'name', '')
    if not filename.lower().endswith('.xlsx'):
        return [], ['Format file harus berupa Excel .xlsx.']

    if getattr(file_obj, 'size', 0) > IMPORT_BARANG_LAB_MAX_SIZE:
        return [], ['Ukuran file import maksimal 7 MB.']

    try:
        workbook = load_workbook(filename=BytesIO(file_obj.read()), data_only=True)
    except Exception:
        return [], ['File Excel tidak dapat dibaca. Pastikan file sesuai format .xlsx dan tidak rusak.']

    worksheet = workbook.active
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return [], ['Baris header pada file Excel tidak ditemukan.']

    normalized_headers = {_normalize_import_header(value): index for index, value in enumerate(header_row) if _string_cell(value)}
    header_aliases = {header: _normalize_import_header(header) for header in IMPORT_BARANG_LAB_HEADERS}
    missing_headers = [header for header in IMPORT_BARANG_LAB_REQUIRED_HEADERS if header_aliases[header] not in normalized_headers]
    if missing_headers:
        return [], [f'Kolom wajib belum tersedia: {", ".join(missing_headers)}.']

    status_choices = _choice_values(StatusBarangChoices.choices)
    satuan_choices = _choice_values(SatuanAsetChoices.choices)
    kondisi_choices = _choice_values(KondisiBarangChoices.choices)
    kategori_choices = _choice_values(KategoriBarangLaboratoriumChoices.choices)

    rows = []
    errors = []
    seen_kode_aset_bmn = {}
    seen_kode_laboratorium = {}

    def cell(row, header):
        key = header_aliases[header]
        index = normalized_headers.get(key)
        return _string_cell(row[index]) if index is not None and index < len(row) else ''

    for excel_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_string_cell(value) for value in row):
            continue

        row_errors = []
        data = {
            'status_barang': cell(row, 'Status Barang'),
            'nama_barang': cell(row, 'Nama Barang'),
            'tipe_merek_barang': cell(row, 'Tipe / Merek Barang'),
            'jenis_barang': cell(row, 'Jenis Barang'),
            'kode_aset_bmn': cell(row, 'Kode Aset BMN'),
            'kode_laboratorium': cell(row, 'Kode Laboratorium'),
            'satuan': cell(row, 'Satuan'),
            'kondisi_barang': cell(row, 'Kondisi Barang'),
            'lokasi_barang': cell(row, 'Lokasi Barang'),
            'kategori_barang': cell(row, 'Kategori Barang'),
            'catatan': cell(row, 'Catatan'),
        }

        for header in IMPORT_BARANG_LAB_REQUIRED_HEADERS:
            if not cell(row, header):
                row_errors.append(f'{header} wajib diisi.')

        if data['status_barang'] not in status_choices:
            row_errors.append('Status Barang tidak sesuai pilihan yang tersedia.')

        if data['satuan'] not in satuan_choices:
            row_errors.append('Satuan tidak sesuai pilihan yang tersedia.')

        if data['kondisi_barang'] not in kondisi_choices:
            row_errors.append('Kondisi Barang tidak sesuai pilihan yang tersedia.')

        if data['kategori_barang'] not in kategori_choices:
            row_errors.append('Kategori Barang tidak sesuai pilihan yang tersedia.')

        if data['status_barang'] == StatusBarangChoices.BMN and not data['kode_aset_bmn']:
            row_errors.append('Kode Aset BMN wajib diisi untuk barang berstatus BMN.')

        try:
            data['tahun_perolehan'] = _year_cell(cell(row, 'Tahun Perolehan'))
        except ValueError as exc:
            row_errors.append(str(exc))
            data['tahun_perolehan'] = None

        for field_name, header_label in (
            ('tanggal_pemeliharaan', 'Tanggal Pemeliharaan'),
            ('tanggal_perbaikan', 'Tanggal Perbaikan'),
        ):
            try:
                data[field_name] = _date_import_cell(cell(row, header_label), header_label)
            except ValueError as exc:
                row_errors.append(str(exc))
                data[field_name] = None

        kode_lab_key = data['kode_laboratorium'].lower()
        if kode_lab_key:
            if kode_lab_key in seen_kode_laboratorium:
                row_errors.append(f'Kode Laboratorium duplikat dengan baris {seen_kode_laboratorium[kode_lab_key]}.')
            else:
                seen_kode_laboratorium[kode_lab_key] = excel_row_number

        kode_bmn_key = data['kode_aset_bmn'].lower()
        if kode_bmn_key:
            if kode_bmn_key in seen_kode_aset_bmn:
                row_errors.append(f'Kode Aset BMN duplikat dengan baris {seen_kode_aset_bmn[kode_bmn_key]}.')
            else:
                seen_kode_aset_bmn[kode_bmn_key] = excel_row_number

        if data['kode_laboratorium'] and BarangLaboratorium.objects.filter(kode_laboratorium__iexact=data['kode_laboratorium']).exists():
            row_errors.append('Kode Laboratorium sudah terdaftar di database.')

        if data['kode_aset_bmn'] and BarangLaboratorium.objects.filter(kode_aset_bmn__iexact=data['kode_aset_bmn']).exists():
            row_errors.append('Kode Aset BMN sudah terdaftar di database.')

        if row_errors:
            errors.append(f'Baris {excel_row_number}: ' + ' '.join(row_errors))
            continue

        if data['status_barang'] != StatusBarangChoices.BMN:
            data['kode_aset_bmn'] = None

        data['volume'] = 0 if data['kondisi_barang'] == KondisiBarangChoices.HILANG else 1
        rows.append(data)

    if not rows and not errors:
        errors.append('File Excel tidak memiliki data barang untuk diimport.')

    return rows, errors


@login_required
def import_barang_laboratorium(request):
    return _handle_import_post(
        request,
        session_key=IMPORT_BARANG_LAB_SESSION_KEY,
        redirect_url_name='master_data:data_barang_laboratorium',
        render_list_func=_render_barang_laboratorium_list,
        validate_func=_validate_barang_laboratorium_import,
        save_func=lambda rows: _save_import_objects(
            rows,
            BarangLaboratorium,
            duplicate_checks=[
                ('kode_laboratorium', 'Kode Laboratorium'),
                ('kode_aset_bmn', 'Kode Aset BMN'),
            ],
        ),
        item_label='peralatan survei lapangan',
    )


@login_required
def download_format_import_barang_laboratorium(request):
    status_values = [value for value, _label in StatusBarangChoices.choices]
    satuan_values = [value for value, _label in SatuanAsetChoices.choices]
    kondisi_values = [value for value, _label in KondisiBarangChoices.choices]
    kategori_values = [value for value, _label in KategoriBarangLaboratoriumChoices.choices]
    return _download_import_template(
        headers=IMPORT_BARANG_LAB_HEADERS,
        sample=[
            'BMN',
            'Total Station',
            'Sokkia CX-105',
            'Alat Ukur Topografi',
            'BMN-2026-0001',
            'LAB-TS-0001',
            'Unit',
            2026,
            'Baik',
            'Gudang Laboratorium',
            'Topografi (TS)',
            '2026-03-01',
            '2026-04-01',
            'Contoh data, silakan hapus/ganti sebelum import.',
        ],
        references=[
            ('Status Barang', status_values),
            ('Satuan', satuan_values),
            ('Kondisi Barang', kondisi_values),
            ('Kategori Barang', kategori_values),
        ],
        validations={
            'Status Barang': status_values,
            'Satuan': satuan_values,
            'Kondisi Barang': kondisi_values,
            'Kategori Barang': kategori_values,
        },
        worksheet_title='Format Import Survei',
        filename='format_import_data_peralatan_survei_lapangan.xlsx',
    )


@login_required
def import_barang_penunjang(request):
    return _handle_import_post(
        request,
        session_key=IMPORT_BARANG_PENUNJANG_SESSION_KEY,
        redirect_url_name='master_data:data_barang_penunjang',
        render_list_func=_render_barang_penunjang_list,
        validate_func=_validate_barang_penunjang_import,
        save_func=lambda rows: _save_import_objects(
            rows,
            BarangPenunjangOperasional,
            duplicate_checks=[('nama_barang', 'Nama Barang')],
        ),
        item_label='barang penunjang lapangan',
    )


@login_required
def download_format_import_barang_penunjang(request):
    kategori_values = [value for value, _label in KategoriBarangPenunjangChoices.choices]
    satuan_values = [value for value, _label in SatuanAsetChoices.choices]
    return _download_import_template(
        headers=IMPORT_BARANG_PENUNJANG_HEADERS,
        sample=[
            'Tripod Kamera',
            'Universal Heavy Duty',
            10,
            0,
            'Unit',
            'Penunjang Operasional Alat Survei',
        ],
        references=[
            ('Satuan', satuan_values),
            ('Kategori Barang', kategori_values),
        ],
        validations={
            'Satuan': satuan_values,
            'Kategori Barang': kategori_values,
        },
        worksheet_title='Format Import Penunjang',
        filename='format_import_data_barang_penunjang_lapangan.xlsx',
    )


@login_required
def import_bahan_operasional(request):
    return _handle_import_post(
        request,
        session_key=IMPORT_BAHAN_OPERASIONAL_SESSION_KEY,
        redirect_url_name='master_data:data_bahan_operasional',
        render_list_func=_render_bahan_operasional_list,
        validate_func=_validate_bahan_operasional_import,
        save_func=lambda rows: _save_import_objects(
            rows,
            BahanOperasional,
            duplicate_checks=[('nama_barang', 'Nama Barang')],
        ),
        item_label='bahan operasional',
    )


@login_required
def download_format_import_bahan_operasional(request):
    kategori_values = [value for value, _label in KategoriBahanOperasionalChoices.choices]
    satuan_values = [value for value, _label in SatuanBahanChoices.choices]
    return _download_import_template(
        headers=IMPORT_BAHAN_OPERASIONAL_HEADERS,
        sample=[
            'Lakban Hitam',
            'Bahan Lapangan',
            25,
            'Rol',
            5,
        ],
        references=[
            ('Kategori Barang', kategori_values),
            ('Satuan', satuan_values),
        ],
        validations={
            'Kategori Barang': kategori_values,
            'Satuan': satuan_values,
        },
        worksheet_title='Format Import Bahan',
        filename='format_import_data_bahan_operasional.xlsx',
    )


@login_required
def import_fasilitas_ruangan(request):
    return _handle_import_post(
        request,
        session_key=IMPORT_FASILITAS_RUANGAN_SESSION_KEY,
        redirect_url_name='master_data:data_fasilitas_ruangan',
        render_list_func=_render_fasilitas_ruangan_list,
        validate_func=_validate_fasilitas_ruangan_import,
        save_func=lambda rows: _save_import_objects(
            rows,
            FasilitasRuangan,
        ),
        item_label='sarana prasarana ruangan',
    )


@login_required
def download_format_import_fasilitas_ruangan(request):
    status_values = [value for value, _label in StatusBarangChoices.choices]
    satuan_values = [value for value, _label in SatuanAsetChoices.choices]
    kategori_values = [value for value, _label in KategoriSaranaPrasaranaChoices.choices]
    kondisi_values = [value for value, _label in KondisiBarangChoices.choices]
    return _download_import_template(
        headers=IMPORT_FASILITAS_RUANGAN_HEADERS,
        sample=[
            'BMN',
            'Meja Laboratorium',
            'Meja Uji Beton',
            'Fasilitas Ruangan',
            'BMN-FAS-0001',
            'FAS-LAB-0001',
            1,
            0,
            'Unit',
            2026,
            'Fasilitas Ruangan',
            'Baik',
            'Ruang Laboratorium 1',
            '2026-03-01',
            '2026-04-01',
            'Contoh catatan sarana.',
        ],
        references=[
            ('Status Barang', status_values),
            ('Satuan', satuan_values),
            ('Kategori Barang', kategori_values),
            ('Kondisi Barang', kondisi_values),
        ],
        validations={
            'Status Barang': status_values,
            'Satuan': satuan_values,
            'Kategori Barang': kategori_values,
            'Kondisi Barang': kondisi_values,
        },
        worksheet_title='Format Import Sarana',
        filename='format_import_data_sarana_prasarana_ruangan.xlsx',
    )



@login_required
def export_barang_penunjang(request):
    headers = [
        "Nama Barang",
        "Tipe / Merek Barang",
        "Kategori Barang",
        "Volume Baik",
        "Volume Rusak",
        "Total Volume",
        "Volume Dipinjam",
        "Sisa Stok",
        "Satuan",
        "Ketersediaan",
        "URL QR Detail",
    ]
    queryset = BarangPenunjangOperasional.objects.order_by("nama_barang")
    return _export_master_queryset(
        request,
        filename="export_data_barang_penunjang_lapangan.xlsx",
        sheet_title="Barang Penunjang",
        headers=headers,
        queryset=queryset,
        row_builder=lambda obj: [
            obj.nama_barang,
            obj.tipe_merek_barang,
            obj.kategori_barang,
            obj.volume_baik,
            obj.volume_rusak,
            obj.total_volume,
            obj.volume_pinjam_aktif,
            obj.sisa_volume,
            obj.satuan,
            obj.ketersediaan,
            _public_detail_export_url(request, obj),
        ],
    )


@login_required
def export_bahan_operasional(request):
    headers = [
        "Nama Barang",
        "Kategori Barang",
        "Volume",
        "Satuan",
        "Stok Minimum",
        "Ketersediaan",
        "URL QR Detail",
    ]
    queryset = BahanOperasional.objects.order_by("nama_barang")
    return _export_master_queryset(
        request,
        filename="export_data_bahan_operasional.xlsx",
        sheet_title="Bahan Operasional",
        headers=headers,
        queryset=queryset,
        row_builder=lambda obj: [
            obj.nama_barang,
            obj.kategori_barang,
            obj.volume,
            obj.satuan,
            obj.stok_minimum,
            obj.ketersediaan,
            _public_detail_export_url(request, obj),
        ],
    )


@login_required
def export_fasilitas_ruangan(request):
    headers = [
        "Nama Barang",
        "Status Barang",
        "Tipe / Merek Barang",
        "Jenis Barang",
        "Kode Aset BMN",
        "Kode Laboratorium",
        "Kategori Barang",
        "Volume Baik",
        "Volume Rusak",
        "Total Volume",
        "Satuan",
        "Tahun Perolehan",
        "Kondisi Barang",
        "Ketersediaan",
        "Sedang Dipinjam",
        "Lokasi Barang",
        "Tanggal Pemeliharaan",
        "Tanggal Perbaikan",
        "Catatan",
        "IK Alat (PDF)",
        "URL QR Detail",
    ]
    queryset = FasilitasRuangan.objects.order_by("nama_barang")
    return _export_master_queryset(
        request,
        filename="export_data_sarana_prasarana_ruangan.xlsx",
        sheet_title="Sarana Prasarana",
        headers=headers,
        queryset=queryset,
        row_builder=lambda obj: [
            obj.nama_barang,
            obj.status_barang,
            obj.tipe_merek_barang,
            obj.jenis_barang,
            obj.kode_aset_bmn,
            obj.kode_laboratorium,
            obj.kategori_barang,
            obj.volume_baik,
            obj.volume_rusak,
            obj.total_volume,
            obj.satuan,
            obj.tahun_perolehan,
            obj.kondisi_barang,
            obj.ketersediaan,
            _boolean_display(obj.sedang_dipinjam),
            obj.lokasi_barang,
            obj.tanggal_pemeliharaan,
            obj.tanggal_perbaikan,
            obj.catatan,
            _file_export_url(request, obj.ik_alat),
            _public_detail_export_url(request, obj),
        ],
    )


@login_required
def export_peralatan_laboratorium(request):
    headers = [
        "Nama Barang",
        "Status Barang",
        "Tipe / Merek Barang",
        "Jenis Barang",
        "Kode Aset BMN",
        "Kode Laboratorium",
        "Volume Baik",
        "Volume Rusak",
        "Total Volume",
        "Volume Dipinjam",
        "Sisa Stok",
        "Satuan",
        "Tahun Perolehan",
        "Kondisi Barang",
        "Ketersediaan",
        "Sedang Dipinjam",
        "Lokasi Barang",
        "Tanggal Pemeliharaan",
        "Tanggal Perbaikan",
        "Catatan",
        "IK Alat (PDF)",
        "URL QR Detail",
    ]
    queryset = PeralatanLaboratorium.objects.order_by("nama_barang")
    return _export_master_queryset(
        request,
        filename="export_data_peralatan_laboratorium.xlsx",
        sheet_title="Peralatan Lab",
        headers=headers,
        queryset=queryset,
        row_builder=lambda obj: [
            obj.nama_barang,
            obj.status_barang,
            obj.tipe_merek_barang,
            obj.jenis_barang,
            obj.kode_aset_bmn,
            obj.kode_laboratorium,
            obj.volume_baik,
            obj.volume_rusak,
            obj.total_volume,
            obj.volume_pinjam_aktif,
            obj.sisa_volume,
            obj.satuan,
            obj.tahun_perolehan,
            obj.kondisi_barang,
            obj.ketersediaan,
            _boolean_display(obj.sedang_dipinjam),
            obj.lokasi_barang,
            obj.tanggal_pemeliharaan,
            obj.tanggal_perbaikan,
            obj.catatan,
            _file_export_url(request, obj.ik_alat),
            _public_detail_export_url(request, obj),
        ],
    )


def _get_status_peminjaman_barang_laboratorium(obj):
    from apps.peminjaman.models import (
        PeminjamanBarangLaboratorium,
        ReturnStepChoices,
        StepChoices,
    )

    active_items = (
        PeminjamanBarangLaboratorium.objects.filter(
            barang=obj,
            pengajuan__current_step=StepChoices.APPROVED,
        )
        .exclude(pengajuan__return_current_step=ReturnStepChoices.COMPLETED)
        .select_related("pengajuan")
        .order_by("-pengajuan__submitted_at", "-pengajuan__id")
    )
    labels = []
    for item in active_items:
        pengajuan = item.pengajuan
        nomor = pengajuan.nomor_pengajuan or "-"
        nama = pengajuan.nama_peminjam or "-"
        labels.append(f"{nomor} - {nama}")
    return "; ".join(labels) if labels else "Tidak sedang dipinjam"


def _get_barang_laboratorium_detail_items(obj):
    return [
        {"label": "Nama Barang", "value": _display_value(obj.nama_barang)},
        {
            "label": "Tipe / Merek Barang",
            "value": _display_value(obj.tipe_merek_barang),
        },
        {"label": "Jenis Barang", "value": _display_value(obj.jenis_barang)},
        {"label": "Status Barang", "value": _display_value(obj.status_barang)},
        {"label": "Kode Aset BMN", "value": _display_value(obj.kode_aset_bmn)},
        {
            "label": "Kode Laboratorium",
            "value": _display_value(obj.kode_laboratorium),
        },
        {"label": "Volume", "value": _display_value(obj.volume)},
        {"label": "Satuan", "value": _display_value(obj.satuan)},
        {"label": "Ketersediaan", "value": _display_value(obj.ketersediaan)},
        {"label": "Tahun Perolehan", "value": _display_value(obj.tahun_perolehan)},
        {"label": "Kondisi Barang", "value": _display_value(obj.kondisi_barang)},
        {"label": "Lokasi Barang", "value": _display_value(obj.lokasi_barang)},
        {"label": "Kategori Barang", "value": _display_value(obj.kategori_barang)},
        {
            "label": "Pemeliharaan Terakhir",
            "value": _display_date(obj.tanggal_pemeliharaan),
        },
        {
            "label": "Perbaikan Terakhir",
            "value": _display_date(obj.tanggal_perbaikan),
        },
        _ik_alat_detail_item(obj),
        {"label": "Catatan", "value": _display_value(obj.catatan), "full": True},
        {
            "label": "Status Peminjaman",
            "value": _get_status_peminjaman_barang_laboratorium(obj),
            "full": True,
        },
    ]


def _get_peralatan_laboratorium_detail_items(obj):
    return [
        {"label": "Nama Barang", "value": _display_value(obj.nama_barang)},
        {
            "label": "Tipe / Merek Barang",
            "value": _display_value(obj.tipe_merek_barang),
        },
        {"label": "Jenis Barang", "value": _display_value(obj.jenis_barang)},
        {"label": "Status Barang", "value": _display_value(obj.status_barang)},
        {"label": "Kode Aset BMN", "value": _display_value(obj.kode_aset_bmn)},
        {
            "label": "Kode Laboratorium",
            "value": _display_value(obj.kode_laboratorium),
        },
        {"label": "Volume Baik", "value": _display_value(obj.volume_baik)},
        {"label": "Volume Rusak", "value": _display_value(obj.volume_rusak)},
        {"label": "Total Volume", "value": _display_value(obj.total_volume)},
        {"label": "Satuan", "value": _display_value(obj.satuan)},
        {"label": "Ketersediaan", "value": _display_value(obj.ketersediaan)},
        {"label": "Tahun Perolehan", "value": _display_value(obj.tahun_perolehan)},
        {
            "label": "Kondisi Barang",
            "value": _display_value(
                obj.kondisi_barang if obj.status_barang == StatusBarangChoices.BMN else None
            ),
        },
        {"label": "Lokasi Barang", "value": _display_value(obj.lokasi_barang)},
        {
            "label": "Pemeliharaan Terakhir",
            "value": _display_date(obj.tanggal_pemeliharaan),
        },
        {
            "label": "Perbaikan Terakhir",
            "value": _display_date(obj.tanggal_perbaikan),
        },
        _ik_alat_detail_item(obj),
        {"label": "Catatan", "value": _display_value(obj.catatan), "full": True},
    ]


def public_barang_laboratorium(request, token):
    from apps.peminjaman.models import PeminjamanBarangLaboratorium

    obj = get_object_or_404(BarangLaboratorium, qr_token=token)
    return _render_public_master_detail(
        request,
        obj=obj,
        page_title="Detail Peralatan Survei Lapangan",
        page_subtitle="Informasi public peralatan survei lapangan Laboratorium.",
        top_meta=obj.kode_laboratorium,
        chips=[obj.status_barang, obj.ketersediaan],
        detail_items=_get_barang_laboratorium_detail_items(obj),
        loan_chart=_build_pinjam_chart(
            request,
            obj,
            item_model=PeminjamanBarangLaboratorium,
            label="Peralatan Survei Lapangan",
        ),
        top_icon="bi-tools",
    )


def public_barang_penunjang(request, token):
    obj = get_object_or_404(BarangPenunjangOperasional, qr_token=token)
    return _render_public_master_detail(
        request,
        obj=obj,
        page_title="Detail Barang Penunjang Lapangan",
        page_subtitle="Informasi public barang penunjang lapangan.",
        top_meta=obj.kategori_barang,
        chips=[obj.ketersediaan],
        detail_items=[
            {"label": "Nama Barang", "value": _display_value(obj.nama_barang)},
            {"label": "Tipe / Merek Barang", "value": _display_value(obj.tipe_merek_barang)},
            {"label": "Volume Baik", "value": _display_value(obj.volume_baik)},
            {"label": "Volume Rusak", "value": _display_value(obj.volume_rusak)},
            {"label": "Total Volume", "value": _display_value(obj.total_volume)},
            {"label": "Satuan", "value": _display_value(obj.satuan)},
            {"label": "Kategori Barang", "value": _display_value(obj.kategori_barang)},
            {"label": "Ketersediaan", "value": _display_value(obj.ketersediaan)},
            {"label": "Stok Tersedia", "value": _display_value(obj.sisa_volume)},
        ],
        top_icon="bi-box-seam",
        photo_attr=None,
    )


def public_bahan_operasional(request, token):
    obj = get_object_or_404(BahanOperasional, qr_token=token)
    return _render_public_master_detail(
        request,
        obj=obj,
        page_title="Detail Bahan Operasional",
        page_subtitle="Informasi public bahan operasional.",
        top_meta=obj.kategori_barang,
        chips=[obj.ketersediaan],
        detail_items=[
            {"label": "Nama Barang", "value": _display_value(obj.nama_barang)},
            {"label": "Kategori Barang", "value": _display_value(obj.kategori_barang)},
            {"label": "Volume", "value": _display_value(obj.volume)},
            {"label": "Satuan", "value": _display_value(obj.satuan)},
            {"label": "Stok Minimum", "value": _display_value(obj.stok_minimum)},
            {"label": "Ketersediaan", "value": _display_value(obj.ketersediaan)},
        ],
        top_icon="bi-droplet",
        photo_attr=None,
    )


def public_fasilitas_ruangan(request, token):
    obj = get_object_or_404(FasilitasRuangan, qr_token=token)
    return _render_public_master_detail(
        request,
        obj=obj,
        page_title="Detail Sarana Prasarana Ruangan",
        page_subtitle="Informasi public sarana prasarana ruangan Laboratorium.",
        top_meta=obj.kode_laboratorium,
        chips=[obj.status_barang, obj.ketersediaan],
        detail_items=[
            {"label": "Nama Barang", "value": _display_value(obj.nama_barang)},
            {"label": "Tipe / Merek Barang", "value": _display_value(obj.tipe_merek_barang)},
            {"label": "Jenis Barang", "value": _display_value(obj.jenis_barang)},
            {"label": "Status Barang", "value": _display_value(obj.status_barang)},
            {"label": "Kode Aset BMN", "value": _display_value(obj.kode_aset_bmn)},
            {"label": "Kode Laboratorium", "value": _display_value(obj.kode_laboratorium)},
            {"label": "Kategori Barang", "value": _display_value(obj.kategori_barang)},
            {"label": "Volume Baik", "value": _display_value(obj.volume_baik)},
            {"label": "Volume Rusak", "value": _display_value(obj.volume_rusak)},
            {"label": "Total Volume", "value": _display_value(obj.total_volume)},
            {"label": "Satuan", "value": _display_value(obj.satuan)},
            {"label": "Ketersediaan", "value": _display_value(obj.ketersediaan)},
            {"label": "Tahun Perolehan", "value": _display_value(obj.tahun_perolehan)},
            {"label": "Kondisi Barang", "value": _display_value(obj.kondisi_barang if obj.status_barang == StatusBarangChoices.BMN else None)},
            {"label": "Lokasi Barang", "value": _display_value(obj.lokasi_barang)},
            {"label": "Pemeliharaan Terakhir", "value": _display_date(obj.tanggal_pemeliharaan)},
            {"label": "Perbaikan Terakhir", "value": _display_date(obj.tanggal_perbaikan)},
            _ik_alat_detail_item(obj),
            {"label": "Catatan", "value": _display_value(obj.catatan), "full": True},
        ],
        top_icon="bi-building",
    )


def public_peralatan_laboratorium(request, token):
    from apps.peminjaman.models import PeminjamanPeralatanLaboratorium

    obj = get_object_or_404(PeralatanLaboratorium, qr_token=token)
    return _render_public_master_detail(
        request,
        obj=obj,
        page_title="Detail Peralatan Laboratorium",
        page_subtitle="Informasi public peralatan laboratorium.",
        top_meta=obj.kode_laboratorium,
        chips=[obj.status_barang, obj.ketersediaan],
        detail_items=_get_peralatan_laboratorium_detail_items(obj),
        loan_chart=_build_pinjam_chart(
            request,
            obj,
            item_model=PeminjamanPeralatanLaboratorium,
            label="Peralatan Laboratorium",
        ),
        top_icon="bi-pc-display",
    )


@login_required
def detail_barang_laboratorium(request, pk):
    from apps.peminjaman.models import PeminjamanBarangLaboratorium

    obj = get_object_or_404(BarangLaboratorium, pk=pk)
    context = _build_detail_context(
        obj=obj,
        page_title="Detail Peralatan Survei Lapangan",
        page_subtitle="Informasi lengkap data peralatan survei lapangan.",
        edit_url_name="master_data:edit_barang_laboratorium",
        edit_label="Edit Peralatan Survei Lapangan",
        back_url_name="master_data:data_barang_laboratorium",
        top_meta=obj.kode_laboratorium,
        chips=[obj.status_barang, obj.ketersediaan],
        detail_items=_get_barang_laboratorium_detail_items(obj),
        loan_chart=_build_pinjam_chart(
            request,
            obj,
            item_model=PeminjamanBarangLaboratorium,
            label="Peralatan Survei Lapangan",
        ),
        top_icon="bi-tools",
    )
    return render(request, "master_data/detail_barang.html", context)


@login_required
def tambah_barang_laboratorium(request):
    return _handle_form_page(
        request,
        form_class=BarangLaboratoriumForm,
        field_groups=BARANG_LAB_FIELD_GROUPS,
        success_message="Data peralatan survei lapangan berhasil ditambahkan.",
        redirect_to="master_data:data_barang_laboratorium",
        page_title="Tambah Data Peralatan Survei Lapangan",
        page_subtitle="Pilih status barang terlebih dahulu sebelum memasukan data lainnya.",
        submit_label="Simpan",
        cancel_url="master_data:data_barang_laboratorium",
    )


@login_required
def edit_barang_laboratorium(request, pk):
    obj = get_object_or_404(BarangLaboratorium, pk=pk)
    return _handle_form_page(
        request,
        form_class=BarangLaboratoriumForm,
        field_groups=BARANG_LAB_FIELD_GROUPS,
        instance=obj,
        success_message="Data peralatan survei lapangan berhasil diperbarui.",
        redirect_to=("master_data:detail_barang_laboratorium", lambda saved_obj: {"pk": saved_obj.pk}),
        page_title="Edit Data Peralatan Survei Lapangan",
        page_subtitle="Perbarui data inventaris peralatan survei lapangan.",
        submit_label="Update",
        cancel_url="master_data:detail_barang_laboratorium",
        cancel_url_kwargs={"pk": obj.pk},
    )


@login_required
def hapus_barang_laboratorium(request, pk):
    return _handle_delete(
        request,
        model=BarangLaboratorium,
        pk=pk,
        success_message="Data peralatan survei lapangan berhasil dihapus.",
        redirect_to="master_data:data_barang_laboratorium",
    )


def _render_barang_penunjang_list(request, import_context=None):
    queryset = BarangPenunjangOperasional.objects.order_by("nama_barang")
    context = paginate_list(request, queryset, search_fields=PENUNJANG_LIST_SEARCH_FIELDS)
    _ensure_qr_codes_for_context_items(context)
    context.update(
        {
            "page_title": "Data Barang Penunjang Lapangan",
            "page_subtitle": "Kelola barang penunjang lapangan.",
            "import_context": import_context or {},
        }
    )
    return render(request, "master_data/data_barang_penunjang_list.html", context)


@login_required
def data_barang_penunjang(request):
    return _render_barang_penunjang_list(request)


@login_required
def tambah_barang_penunjang(request):
    return _handle_form_page(
        request,
        form_class=BarangPenunjangOperasionalForm,
        field_groups=PENUNJANG_FIELD_GROUPS,
        success_message="Data barang penunjang lapangan berhasil ditambahkan.",
        redirect_to="master_data:data_barang_penunjang",
        page_title="Tambah Data Barang Penunjang Lapangan",
        page_subtitle="Tambahkan data barang penunjang lapangan.",
        submit_label="Simpan",
        cancel_url="master_data:data_barang_penunjang",
    )


@login_required
def edit_barang_penunjang(request, pk):
    obj = get_object_or_404(BarangPenunjangOperasional, pk=pk)
    return _handle_form_page(
        request,
        form_class=BarangPenunjangOperasionalForm,
        field_groups=PENUNJANG_FIELD_GROUPS,
        instance=obj,
        success_message="Data barang penunjang lapangan berhasil diperbarui.",
        redirect_to="master_data:data_barang_penunjang",
        page_title="Edit Data Barang Penunjang Lapangan",
        page_subtitle="Perbarui data barang penunjang lapangan.",
        submit_label="Update",
        cancel_url="master_data:data_barang_penunjang",
    )


@login_required
def hapus_barang_penunjang(request, pk):
    return _handle_delete(
        request,
        model=BarangPenunjangOperasional,
        pk=pk,
        success_message="Data barang penunjang lapangan berhasil dihapus.",
        redirect_to="master_data:data_barang_penunjang",
    )


def _render_bahan_operasional_list(request, import_context=None):
    queryset = BahanOperasional.objects.order_by("nama_barang")
    context = paginate_list(request, queryset, search_fields=BAHAN_LIST_SEARCH_FIELDS)
    _ensure_qr_codes_for_context_items(context)
    context.update(
        {
            "page_title": "Data Bahan Operasional",
            "page_subtitle": "Kelola bahan operasional dan status stok ketersediaan.",
            "import_context": import_context or {},
        }
    )
    return render(request, "master_data/data_bahan_operasional_list.html", context)


@login_required
def data_bahan_operasional(request):
    return _render_bahan_operasional_list(request)


@login_required
def tambah_bahan_operasional(request):
    return _handle_form_page(
        request,
        form_class=BahanOperasionalForm,
        field_groups=BAHAN_FIELD_GROUPS,
        success_message="Data bahan operasional berhasil ditambahkan.",
        redirect_to="master_data:data_bahan_operasional",
        page_title="Tambah Data Bahan Operasional",
        page_subtitle="Tambahkan data bahan operasional.",
        submit_label="Simpan",
        cancel_url="master_data:data_bahan_operasional",
    )


@login_required
def edit_bahan_operasional(request, pk):
    obj = get_object_or_404(BahanOperasional, pk=pk)
    return _handle_form_page(
        request,
        form_class=BahanOperasionalForm,
        field_groups=BAHAN_FIELD_GROUPS,
        instance=obj,
        success_message="Data bahan operasional berhasil diperbarui.",
        redirect_to="master_data:data_bahan_operasional",
        page_title="Edit Data Bahan Operasional",
        page_subtitle="Perbarui volume bahan operasional dan stok minimum.",
        submit_label="Update",
        cancel_url="master_data:data_bahan_operasional",
    )


@login_required
def hapus_bahan_operasional(request, pk):
    return _handle_delete(
        request,
        model=BahanOperasional,
        pk=pk,
        success_message="Data bahan operasional berhasil dihapus.",
        redirect_to="master_data:data_bahan_operasional",
    )


def _render_fasilitas_ruangan_list(request, import_context=None):
    queryset = FasilitasRuangan.objects.order_by("nama_barang")
    context = paginate_list(
        request,
        queryset,
        search_fields=(*ASSET_LIST_SEARCH_FIELDS, "kategori_barang"),
    )
    _ensure_qr_codes_for_context_items(context)
    context.update(
        {
            "page_title": "Data Sarana Prasarana Ruangan",
            "page_subtitle": "Kelola sarana prasarana ruangan Laboratorium.",
            "import_context": import_context or {},
        }
    )
    return render(request, "master_data/data_fasilitas_ruangan_list.html", context)


@login_required
def data_fasilitas_ruangan(request):
    return _render_fasilitas_ruangan_list(request)


@login_required
def detail_fasilitas_ruangan(request, pk):
    obj = get_object_or_404(FasilitasRuangan, pk=pk)
    context = _build_detail_context(
        obj=obj,
        page_title="Detail Sarana Prasarana Ruangan",
        page_subtitle="Informasi lengkap data sarana prasarana ruangan.",
        edit_url_name="master_data:edit_fasilitas_ruangan",
        edit_label="Edit Sarana Prasarana Ruangan",
        back_url_name="master_data:data_fasilitas_ruangan",
        top_meta=obj.kode_laboratorium,
        chips=[obj.status_barang, obj.ketersediaan],
        detail_items=[
            {"label": "Nama Barang", "value": _display_value(obj.nama_barang)},
            {
                "label": "Tipe / Merek Barang",
                "value": _display_value(obj.tipe_merek_barang),
            },
            {"label": "Jenis Barang", "value": _display_value(obj.jenis_barang)},
            {"label": "Status Barang", "value": _display_value(obj.status_barang)},
            {"label": "Kode Aset BMN", "value": _display_value(obj.kode_aset_bmn)},
            {
                "label": "Kode Laboratorium",
                "value": _display_value(obj.kode_laboratorium),
            },
            {"label": "Kategori Barang", "value": _display_value(obj.kategori_barang)},
            {"label": "Volume Baik", "value": _display_value(obj.volume_baik)},
            {"label": "Volume Rusak", "value": _display_value(obj.volume_rusak)},
            {"label": "Total Volume", "value": _display_value(obj.total_volume)},
            {"label": "Satuan", "value": _display_value(obj.satuan)},
            {"label": "Ketersediaan", "value": _display_value(obj.ketersediaan)},
            {"label": "Tahun Perolehan", "value": _display_value(obj.tahun_perolehan)},
            {"label": "Kondisi Barang", "value": _display_value(obj.kondisi_barang if obj.status_barang == StatusBarangChoices.BMN else None)},
            {"label": "Lokasi Barang", "value": _display_value(obj.lokasi_barang)},
            {
                "label": "Pemeliharaan Terakhir",
                "value": _display_date(obj.tanggal_pemeliharaan),
            },
            {
                "label": "Perbaikan Terakhir",
                "value": _display_date(obj.tanggal_perbaikan),
            },
            _ik_alat_detail_item(obj),
            {"label": "Catatan", "value": _display_value(obj.catatan), "full": True},
        ],
        top_icon="bi-building",
    )
    return render(request, "master_data/detail_barang.html", context)


@login_required
def tambah_fasilitas_ruangan(request):
    return _handle_form_page(
        request,
        form_class=FasilitasRuanganForm,
        field_groups=SARANA_FIELD_GROUPS,
        success_message="Data sarana prasarana ruangan berhasil ditambahkan.",
        redirect_to="master_data:data_fasilitas_ruangan",
        page_title="Tambah Data Sarana Prasarana Ruangan",
        page_subtitle="Pilih status barang terlebih dahulu sebelum memasukan data lainnya.",
        submit_label="Simpan",
        cancel_url="master_data:data_fasilitas_ruangan",
    )


@login_required
def edit_fasilitas_ruangan(request, pk):
    obj = get_object_or_404(FasilitasRuangan, pk=pk)
    return _handle_form_page(
        request,
        form_class=FasilitasRuanganForm,
        field_groups=SARANA_FIELD_GROUPS,
        instance=obj,
        success_message="Data sarana prasarana ruangan berhasil diperbarui.",
        redirect_to=("master_data:detail_fasilitas_ruangan", lambda saved_obj: {"pk": saved_obj.pk}),
        page_title="Edit Data Sarana Prasarana Ruangan",
        page_subtitle="Perbarui data inventaris sarana prasarana ruangan.",
        submit_label="Update",
        cancel_url="master_data:detail_fasilitas_ruangan",
        cancel_url_kwargs={"pk": obj.pk},
    )


@login_required
def hapus_fasilitas_ruangan(request, pk):
    return _handle_delete(
        request,
        model=FasilitasRuangan,
        pk=pk,
        success_message="Data sarana prasarana ruangan berhasil dihapus.",
        redirect_to="master_data:data_fasilitas_ruangan",
    )

def _render_peralatan_laboratorium_list(request, import_context=None):
    queryset = PeralatanLaboratorium.objects.order_by("nama_barang")
    context = paginate_list(request, queryset, search_fields=ASSET_LIST_SEARCH_FIELDS)
    _ensure_qr_codes_for_context_items(context)
    context.update(
        {
            "page_title": "Data Peralatan Laboratorium",
            "page_subtitle": "Kelola peralatan laboratorium.",
            "import_context": import_context or {},
        }
    )
    return render(request, "master_data/data_peralatan_laboratorium_list.html", context)


@login_required
def data_peralatan_laboratorium(request):
    return _render_peralatan_laboratorium_list(request)


@login_required
def detail_peralatan_laboratorium(request, pk):
    from apps.peminjaman.models import PeminjamanPeralatanLaboratorium

    obj = get_object_or_404(PeralatanLaboratorium, pk=pk)
    context = _build_detail_context(
        obj=obj,
        page_title="Detail Peralatan Laboratorium",
        page_subtitle="Informasi lengkap data peralatan laboratorium.",
        edit_url_name="master_data:edit_peralatan_laboratorium",
        edit_label="Edit Peralatan Laboratorium",
        back_url_name="master_data:data_peralatan_laboratorium",
        top_meta=obj.kode_laboratorium,
        chips=[obj.status_barang, obj.ketersediaan],
        detail_items=_get_peralatan_laboratorium_detail_items(obj),
        loan_chart=_build_pinjam_chart(
            request,
            obj,
            item_model=PeminjamanPeralatanLaboratorium,
            label="Peralatan Laboratorium",
        ),
        top_icon="bi-pc-display",
    )
    return render(request, "master_data/detail_barang.html", context)


@login_required
def tambah_peralatan_laboratorium(request):
    return _handle_form_page(
        request,
        form_class=PeralatanLaboratoriumForm,
        field_groups=PERALATAN_LAB_FIELD_GROUPS,
        success_message="Data peralatan laboratorium berhasil ditambahkan.",
        redirect_to="master_data:data_peralatan_laboratorium",
        page_title="Tambah Data Peralatan Laboratorium",
        page_subtitle="Pilih status barang terlebih dahulu sebelum memasukan data lainnya.",
        submit_label="Simpan",
        cancel_url="master_data:data_peralatan_laboratorium",
    )


@login_required
def edit_peralatan_laboratorium(request, pk):
    obj = get_object_or_404(PeralatanLaboratorium, pk=pk)
    return _handle_form_page(
        request,
        form_class=PeralatanLaboratoriumForm,
        field_groups=PERALATAN_LAB_FIELD_GROUPS,
        instance=obj,
        success_message="Data peralatan laboratorium berhasil diperbarui.",
        redirect_to=("master_data:detail_peralatan_laboratorium", lambda saved_obj: {"pk": saved_obj.pk}),
        page_title="Edit Data Peralatan Laboratorium",
        page_subtitle="Perbarui data inventaris peralatan laboratorium.",
        submit_label="Update",
        cancel_url="master_data:detail_peralatan_laboratorium",
        cancel_url_kwargs={"pk": obj.pk},
    )


@login_required
def hapus_peralatan_laboratorium(request, pk):
    return _handle_delete(
        request,
        model=PeralatanLaboratorium,
        pk=pk,
        success_message="Data peralatan laboratorium berhasil dihapus.",
        redirect_to="master_data:data_peralatan_laboratorium",
    )


@login_required
def import_peralatan_laboratorium(request):
    return _handle_import_post(
        request,
        session_key=IMPORT_PERALATAN_LABORATORIUM_SESSION_KEY,
        redirect_url_name='master_data:data_peralatan_laboratorium',
        render_list_func=_render_peralatan_laboratorium_list,
        validate_func=_validate_peralatan_laboratorium_import,
        save_func=lambda rows: _save_import_objects(
            rows,
            PeralatanLaboratorium,
            duplicate_checks=[
                ('kode_laboratorium', 'Kode Laboratorium'),
                ('kode_aset_bmn', 'Kode Aset BMN'),
            ],
        ),
        item_label='peralatan laboratorium',
    )


@login_required
def download_format_import_peralatan_laboratorium(request):
    status_values = [value for value, _label in StatusBarangChoices.choices]
    satuan_values = [value for value, _label in SatuanAsetChoices.choices]
    kondisi_values = [value for value, _label in KondisiBarangChoices.choices]
    return _download_import_template(
        headers=IMPORT_PERALATAN_LABORATORIUM_HEADERS,
        sample=[
            'BMN',
            'Mikroskop Laboratorium',
            'Olympus CX23',
            'Peralatan Laboratorium',
            'BMN-ALAT-LAB-0001',
            'ALAT-LAB-0001',
            1,
            0,
            'Unit',
            2026,
            'Baik',
            'Ruang Laboratorium 1',
            '2026-03-01',
            '2026-04-01',
            'Contoh catatan peralatan.',
        ],
        references=[
            ('Status Barang', status_values),
            ('Satuan', satuan_values),
            ('Kondisi Barang', kondisi_values),
        ],
        validations={
            'Status Barang': status_values,
            'Satuan': satuan_values,
            'Kondisi Barang': kondisi_values,
        },
        worksheet_title='Format Import Alat Lab',
        filename='format_import_data_peralatan_laboratorium.xlsx',
    )
