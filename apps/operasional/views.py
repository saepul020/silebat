from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.shortcuts import get_object_or_404, redirect, render

from apps.core.list_pagination import paginate_list
from apps.core.excel_utils import build_excel_response
from apps.core.import_utils import (
    choice_values as _choice_values,
    import_cell as _import_cell,
    json_response_or_none as _import_json_response,
    load_import_worksheet as _load_shared_import_worksheet,
    string_cell as _string_cell,
)
from apps.core.permissions import ROLE_SUPER_ADMIN, get_role_name
from apps.core.navigation import get_next_url, redirect_next

from .forms import (
    DataKopDokumenForm,
    InstansiKlienForm,
    LayananKegiatanForm,
    SurveiKegiatanForm,
    TimKegiatanForm,
)
from .models import DataKopDokumen, InstansiKlien, LayananKegiatan, SurveiKegiatan, TimKegiatan


FORM_TEMPLATE = "operasional/master_form.html"
TIM_LIST_SEARCH_FIELDS = (
    "nama_tim",
    "ketua_tim__username",
    "ketua_tim__first_name",
    "ketua_tim__last_name",
)
IMPORT_INSTANSI_SESSION_KEY = "instansi_klien_import_validated_rows"
IMPORT_INSTANSI_HEADERS = [
    "Nama Instansi",
    "Alamat Instansi",
    "Organisasi",
]
IMPORT_INSTANSI_REQUIRED_HEADERS = IMPORT_INSTANSI_HEADERS[:]
IMPORT_INSTANSI_MAX_SIZE = 7 * 1024 * 1024


def _is_super_admin_user(user):
    return get_role_name(user) == ROLE_SUPER_ADMIN


def _deny_import_access(request):
    messages.error(request, 'Fitur import data instansi hanya dapat diakses oleh Super Admin.')
    return redirect('operasional:data_instansi')


def _load_import_worksheet(file_obj, headers, required_headers):
    return _load_shared_import_worksheet(
        file_obj,
        headers,
        required_headers,
        max_size_bytes=IMPORT_INSTANSI_MAX_SIZE,
    )


def _validate_instansi_import(file_obj):
    worksheet, normalized_headers, header_aliases, initial_errors = _load_import_worksheet(
        file_obj,
        IMPORT_INSTANSI_HEADERS,
        IMPORT_INSTANSI_REQUIRED_HEADERS,
    )
    if initial_errors:
        return [], initial_errors

    organisasi_choices = _choice_values(InstansiKlien.OrganisasiChoices.choices)
    rows = []
    errors = []
    seen_nama = {}

    def cell(row, header):
        return _import_cell(row, header, normalized_headers, header_aliases)

    for excel_row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_string_cell(value) for value in row):
            continue

        row_errors = []
        data = {
            'nama_instansi': cell(row, 'Nama Instansi'),
            'alamat_instansi': cell(row, 'Alamat Instansi'),
            'organisasi': cell(row, 'Organisasi'),
        }

        for header in IMPORT_INSTANSI_REQUIRED_HEADERS:
            if not cell(row, header):
                row_errors.append(f'{header} wajib diisi.')

        if data['organisasi'] and data['organisasi'] not in organisasi_choices:
            row_errors.append('Organisasi tidak sesuai pilihan yang tersedia.')

        nama_key = data['nama_instansi'].lower()
        if nama_key:
            if nama_key in seen_nama:
                row_errors.append(f'Nama Instansi duplikat dengan baris {seen_nama[nama_key]}.')
            else:
                seen_nama[nama_key] = excel_row_number

        if data['nama_instansi'] and InstansiKlien.objects.filter(nama_instansi__iexact=data['nama_instansi']).exists():
            row_errors.append('Nama Instansi sudah terdaftar di database.')

        if row_errors:
            errors.append(f'Baris {excel_row_number}: ' + ' '.join(row_errors))
            continue

        rows.append(data)

    if not rows and not errors:
        errors.append('File Excel tidak memiliki data instansi untuk diimport.')

    return rows, errors


def _save_instansi_import(rows):
    duplicate_errors = []
    for index, row in enumerate(rows, start=1):
        if InstansiKlien.objects.filter(nama_instansi__iexact=row['nama_instansi']).exists():
            duplicate_errors.append(f'Data valid nomor {index}: Nama Instansi sudah terdaftar di database.')

    if duplicate_errors:
        return 0, duplicate_errors

    with transaction.atomic():
        objects = [InstansiKlien(**row) for row in rows]
        for obj in objects:
            obj.full_clean()
            obj.save()

    return len(rows), []


def _handle_instansi_import_post(request):
    if request.method != 'POST':
        return redirect('operasional:data_instansi')

    action = request.POST.get('import_action')

    if action == 'cancel':
        request.session.pop(IMPORT_INSTANSI_SESSION_KEY, None)
        response = _import_json_response(request, {'ok': True, 'cancelled': True})
        if response:
            return response
        return redirect('operasional:data_instansi')

    if action == 'validate':
        rows, errors = _validate_instansi_import(request.FILES.get('file_import'))
        import_context = {
            'show_modal': True,
            'validated': not errors,
            'can_save': bool(rows) and not errors,
            'total_rows': len(rows),
            'errors': errors,
        }
        if errors:
            request.session.pop(IMPORT_INSTANSI_SESSION_KEY, None)
        else:
            request.session[IMPORT_INSTANSI_SESSION_KEY] = rows
            request.session.modified = True

        response = _import_json_response(request, {
            'ok': not bool(errors),
            'validated': not bool(errors),
            'can_save': bool(rows) and not errors,
            'total_rows': len(rows),
            'errors': errors,
            'message': f'Validasi berhasil. {len(rows)} data siap disimpan.' if not errors else 'Validasi belum berhasil.',
        })
        if response:
            return response
        return _render_data_instansi_list(request, import_context=import_context)

    if action == 'save':
        rows = request.session.get(IMPORT_INSTANSI_SESSION_KEY) or []
        if not rows:
            payload = {
                'ok': False,
                'saved': False,
                'can_save': False,
                'errors': ['Data validasi tidak ditemukan atau sudah kedaluwarsa. Lakukan Validasi Data terlebih dahulu.'],
                'message': 'Data belum dapat disimpan.',
            }
            response = _import_json_response(request, payload)
            if response:
                return response
            return _render_data_instansi_list(request, import_context={'show_modal': True, 'errors': payload['errors']})

        total_saved, save_errors = _save_instansi_import(rows)
        if save_errors:
            request.session.pop(IMPORT_INSTANSI_SESSION_KEY, None)
            payload = {
                'ok': False,
                'saved': False,
                'can_save': False,
                'errors': save_errors,
                'message': 'Data tidak disimpan karena ditemukan duplikasi terbaru di database.',
            }
            response = _import_json_response(request, payload)
            if response:
                return response
            return _render_data_instansi_list(request, import_context={'show_modal': True, 'errors': save_errors})

        request.session.pop(IMPORT_INSTANSI_SESSION_KEY, None)
        response = _import_json_response(request, {
            'ok': True,
            'saved': True,
            'can_save': False,
            'total_rows': total_saved,
            'errors': [],
            'message': f'{total_saved} data instansi berhasil diimport.',
            'redirect_url': reverse('operasional:data_instansi'),
        })
        if response:
            return response
        return redirect('operasional:data_instansi')

    payload = {'ok': False, 'errors': ['Aksi import tidak dikenali.'], 'message': 'Aksi import tidak dikenali.'}
    response = _import_json_response(request, payload, status=400)
    if response:
        return response
    return redirect('operasional:data_instansi')


def _download_import_template(*, headers, sample, references, validations, worksheet_title, filename):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:
        raise RuntimeError('Library openpyxl belum tersedia. Jalankan: pip install openpyxl') from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_title

    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='EAF2FF')
        worksheet.column_dimensions[cell.column_letter].width = 28

    for column_index, value in enumerate(sample, start=1):
        worksheet.cell(row=2, column=column_index, value=value)

    if references:
        lists_sheet = workbook.create_sheet('Referensi Pilihan')
        for col_index, (title, values) in enumerate(references, start=1):
            lists_sheet.cell(row=1, column=col_index, value=title).font = Font(bold=True)
            lists_sheet.column_dimensions[lists_sheet.cell(row=1, column=col_index).column_letter].width = 34
            for row_index, value in enumerate(values, start=2):
                lists_sheet.cell(row=row_index, column=col_index, value=value)

    for header, values in (validations or {}).items():
        if header not in headers:
            continue
        column_index = headers.index(header) + 1
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        formula = '"' + ','.join(values) + '"'
        validator = DataValidation(type='list', formula1=formula, allow_blank=False)
        worksheet.add_data_validation(validator)
        validator.add(f'{column_letter}2:{column_letter}1000')

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response




@login_required
def export_instansi(request):
    if not _is_super_admin_user(request.user):
        return _deny_import_access(request)

    queryset = InstansiKlien.objects.order_by("nama_instansi")
    return build_excel_response(
        "export_data_instansi_klien.xlsx",
        [
            {
                "title": "Data Instansi",
                "headers": ["Nama Instansi", "Alamat Instansi", "Organisasi"],
                "rows": [
                    [item.nama_instansi, item.alamat_instansi, item.organisasi]
                    for item in queryset
                ],
            }
        ],
    )

@login_required
def import_instansi(request):
    if not _is_super_admin_user(request.user):
        return _deny_import_access(request)
    return _handle_instansi_import_post(request)


@login_required
def download_format_import_instansi(request):
    if not _is_super_admin_user(request.user):
        return _deny_import_access(request)

    organisasi_values = [value for value, _label in InstansiKlien.OrganisasiChoices.choices]
    return _download_import_template(
        headers=IMPORT_INSTANSI_HEADERS,
        sample=[
            'Dinas PUPR Provinsi Nusa Tenggara Timur',
            'Jl. Contoh No. 1, Kupang',
            'Eksternal PU',
        ],
        references=[
            ('Organisasi', organisasi_values),
        ],
        validations={
            'Organisasi': organisasi_values,
        },
        worksheet_title='Format Import Instansi',
        filename='format_import_data_instansi_klien.xlsx',
    )




def _render_form_page(
    request,
    *,
    form,
    page_title,
    page_subtitle,
    submit_label,
    cancel_url,
    template_name=FORM_TEMPLATE,
    extra_context=None,
):
    next_url = get_next_url(request)
    context = {
        "form": form,
        "page_title": page_title,
        "page_subtitle": page_subtitle,
        "submit_label": submit_label,
        "cancel_url": cancel_url,
        "next_url": next_url,
    }
    if extra_context:
        context.update(extra_context)
    return render(
        request,
        template_name,
        context,
    )


def _handle_form_page(
    request,
    *,
    form_class,
    success_message,
    redirect_to,
    page_title,
    page_subtitle,
    submit_label,
    cancel_url,
    instance=None,
    template_name=FORM_TEMPLATE,
    extra_context=None,
):
    form_kwargs = {}
    if request.method == "POST":
        form_kwargs["data"] = request.POST
        form_kwargs["files"] = request.FILES
    if instance is not None:
        form_kwargs["instance"] = instance

    form = form_class(**form_kwargs)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, success_message)
        return redirect_next(request, redirect_to)

    return _render_form_page(
        request,
        form=form,
        page_title=page_title,
        page_subtitle=page_subtitle,
        submit_label=submit_label,
        cancel_url=cancel_url,
        template_name=template_name,
        extra_context=extra_context,
    )


def _handle_delete(request, *, model, pk, success_message, redirect_to):
    obj = get_object_or_404(model, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, success_message)
    return redirect(redirect_to)


@login_required
def index(request):
    return redirect("operasional:data_tim")


@login_required
def data_tim(request):
    queryset = TimKegiatan.objects.select_related("ketua_tim").order_by("nama_tim")
    pagination_context = paginate_list(request, queryset, search_fields=TIM_LIST_SEARCH_FIELDS)
    context = {
        "items": pagination_context["items"],
        "page_title": "Data Tim Kegiatan",
        "page_subtitle": "Kelola data divisi tim kegiatan Balai Air Tanah.",
    }
    context.update(pagination_context)
    return render(request, "operasional/data_tim_list.html", context)


@login_required
def tambah_tim(request):
    return _handle_form_page(
        request,
        form_class=TimKegiatanForm,
        success_message="Data tim kegiatan berhasil ditambahkan.",
        redirect_to="operasional:data_tim",
        page_title="Tambah Data Tim Kegiatan",
        page_subtitle="Tambahkan nama divisi tim dan ketua divisi tim kegiatan Balai Air Tanah.",
        submit_label="Simpan",
        cancel_url="operasional:data_tim",
    )


@login_required
def edit_tim(request, pk):
    obj = get_object_or_404(TimKegiatan, pk=pk)
    return _handle_form_page(
        request,
        form_class=TimKegiatanForm,
        instance=obj,
        success_message="Data tim kegiatan berhasil diperbarui.",
        redirect_to="operasional:data_tim",
        page_title="Edit Data Tim Kegiatan",
        page_subtitle="Perbarui nama tim dan ketua tim kegiatan.",
        submit_label="Update",
        cancel_url="operasional:data_tim",
    )


@login_required
def hapus_tim(request, pk):
    return _handle_delete(
        request,
        model=TimKegiatan,
        pk=pk,
        success_message="Data tim kegiatan berhasil dihapus.",
        redirect_to="operasional:data_tim",
    )


@login_required
def data_layanan(request):
    queryset = LayananKegiatan.objects.order_by("jenis_layanan")
    pagination_context = paginate_list(request, queryset, search_fields=("jenis_layanan",))
    context = {
        "items": pagination_context["items"],
        "page_title": "Data Layanan Kegiatan",
        "page_subtitle": "Kelola jenis layanan kegiatan yang berlaku.",
    }
    context.update(pagination_context)
    return render(request, "operasional/data_layanan_list.html", context)


@login_required
def tambah_layanan(request):
    return _handle_form_page(
        request,
        form_class=LayananKegiatanForm,
        success_message="Data layanan kegiatan berhasil ditambahkan.",
        redirect_to="operasional:data_layanan",
        page_title="Tambah Data Layanan Kegiatan",
        page_subtitle="Tambahkan jenis layanan kegiatan baru.",
        submit_label="Simpan",
        cancel_url="operasional:data_layanan",
    )


@login_required
def edit_layanan(request, pk):
    obj = get_object_or_404(LayananKegiatan, pk=pk)
    return _handle_form_page(
        request,
        form_class=LayananKegiatanForm,
        instance=obj,
        success_message="Data layanan kegiatan berhasil diperbarui.",
        redirect_to="operasional:data_layanan",
        page_title="Edit Data Layanan Kegiatan",
        page_subtitle="Perbarui jenis layanan kegiatan.",
        submit_label="Update",
        cancel_url="operasional:data_layanan",
    )


@login_required
def hapus_layanan(request, pk):
    return _handle_delete(
        request,
        model=LayananKegiatan,
        pk=pk,
        success_message="Data layanan kegiatan berhasil dihapus.",
        redirect_to="operasional:data_layanan",
    )


@login_required
def data_survei(request):
    queryset = SurveiKegiatan.objects.order_by("jenis_survei")
    pagination_context = paginate_list(request, queryset, search_fields=("jenis_survei",))
    context = {
        "items": pagination_context["items"],
        "page_title": "Data Kegiatan Survei",
        "page_subtitle": "Kelola jenis kegiatan survei yang tersedia.",
    }
    context.update(pagination_context)
    return render(request, "operasional/data_survei_list.html", context)


@login_required
def tambah_survei(request):
    return _handle_form_page(
        request,
        form_class=SurveiKegiatanForm,
        success_message="Data kegiatan survei berhasil ditambahkan.",
        redirect_to="operasional:data_survei",
        page_title="Tambah Data Kegiatan Survei",
        page_subtitle="Tambahkan jenis kegiatan survei baru.",
        submit_label="Simpan",
        cancel_url="operasional:data_survei",
    )


@login_required
def edit_survei(request, pk):
    obj = get_object_or_404(SurveiKegiatan, pk=pk)
    return _handle_form_page(
        request,
        form_class=SurveiKegiatanForm,
        instance=obj,
        success_message="Data kegiatan survei berhasil diperbarui.",
        redirect_to="operasional:data_survei",
        page_title="Edit Data Kegiatan Survei",
        page_subtitle="Perbarui jenis kegiatan survei.",
        submit_label="Update",
        cancel_url="operasional:data_survei",
    )


@login_required
def hapus_survei(request, pk):
    return _handle_delete(
        request,
        model=SurveiKegiatan,
        pk=pk,
        success_message="Data kegiatan survei berhasil dihapus.",
        redirect_to="operasional:data_survei",
    )


def _render_data_instansi_list(request, import_context=None):
    queryset = InstansiKlien.objects.order_by("nama_instansi")
    pagination_context = paginate_list(
        request,
        queryset,
        search_fields=("nama_instansi", "alamat_instansi", "organisasi"),
    )
    context = {
        "items": pagination_context["items"],
        "page_title": "Data Instansi (Klien)",
        "page_subtitle": "Kelola data instansi atau klien yang terlibat dalam kegiatan.",
        "import_context": import_context or {},
    }
    context.update(pagination_context)
    return render(request, "operasional/data_instansi_list.html", context)


@login_required
def data_instansi(request):
    return _render_data_instansi_list(request)


@login_required
def tambah_instansi(request):
    return _handle_form_page(
        request,
        form_class=InstansiKlienForm,
        success_message="Data instansi berhasil ditambahkan.",
        redirect_to="operasional:data_instansi",
        page_title="Tambah Data Instansi (Klien)",
        page_subtitle="Tambahkan nama instansi, alamat, dan kategori organisasi.",
        submit_label="Simpan",
        cancel_url="operasional:data_instansi",
    )


@login_required
def edit_instansi(request, pk):
    obj = get_object_or_404(InstansiKlien, pk=pk)
    return _handle_form_page(
        request,
        form_class=InstansiKlienForm,
        instance=obj,
        success_message="Data instansi berhasil diperbarui.",
        redirect_to="operasional:data_instansi",
        page_title="Edit Data Instansi (Klien)",
        page_subtitle="Perbarui data instansi / klien.",
        submit_label="Update",
        cancel_url="operasional:data_instansi",
    )


@login_required
def hapus_instansi(request, pk):
    return _handle_delete(
        request,
        model=InstansiKlien,
        pk=pk,
        success_message="Data instansi berhasil dihapus.",
        redirect_to="operasional:data_instansi",
    )


@login_required
def data_kop_dokumen(request):
    queryset = DataKopDokumen.objects.order_by("id")
    pagination_context = paginate_list(request, queryset, search_fields=("kop_dokumen",))
    items = pagination_context["items"]
    item = queryset.first()
    context = {
        "item": item,
        "items": items,
        "page_title": "Data Kop Dokumen",
        "page_subtitle": "Kelola gambar kop dokumen yang digunakan pada dokumen sistem.",
    }
    context.update(pagination_context)
    return render(request, "operasional/data_kop_dokumen_list.html", context)


@login_required
def tambah_kop_dokumen(request):
    existing = DataKopDokumen.objects.first()
    if existing is not None:
        messages.warning(request, "Data Kop Dokumen sudah tersedia. Silakan edit data yang ada.")
        return redirect("operasional:edit_kop_dokumen", pk=existing.pk)

    return _handle_form_page(
        request,
        form_class=DataKopDokumenForm,
        success_message="Data Kop Dokumen berhasil ditambahkan.",
        redirect_to="operasional:data_kop_dokumen",
        page_title="Tambah Data Kop Dokumen",
        page_subtitle="Unggah gambar kop dokumen yang akan digunakan pada dokumen sistem.",
        submit_label="Simpan",
        cancel_url="operasional:data_kop_dokumen",
        template_name="operasional/data_kop_dokumen_form.html",
    )


@login_required
def edit_kop_dokumen(request, pk):
    obj = get_object_or_404(DataKopDokumen, pk=pk)
    return _handle_form_page(
        request,
        form_class=DataKopDokumenForm,
        instance=obj,
        success_message="Data Kop Dokumen berhasil diperbarui.",
        redirect_to="operasional:data_kop_dokumen",
        page_title="Edit Data Kop Dokumen",
        page_subtitle="Perbarui gambar kop dokumen yang digunakan pada dokumen sistem.",
        submit_label="Update",
        cancel_url="operasional:data_kop_dokumen",
        template_name="operasional/data_kop_dokumen_form.html",
    )
